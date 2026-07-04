"""Deterministic parse of a returned SoR sheet (xlsx) into a :class:`BidReply` — no model.

Dispatch's own per-firm attachment IS an xlsx we authored (``generate_sor_sheet``:
a title row, a note row, a blank row, then the header ``Item | Description | Unit |
Qty | Rate (HKD) | Amount (HKD)`` and one row per priceable item). The realistic reply
is that very sheet returned with the Rate column filled — parsing it needs no model at
all, so this path is pure openpyxl: local, offline, byte-exact.

The parser is layout-tolerant within that authored format: it finds the header row by
its column names (wherever the preamble ends), maps columns by header (so a re-saved or
lightly reshuffled sheet still parses), skips any row without an item ref (the preamble,
totals, blank padding), and leaves blank cells as ``None`` — the sparse-field contract
of :class:`~schemas.models.BidLineItem`. A workbook with no recognisable header row
raises ``ValueError`` so the API can return a clear 4xx rather than a crash.

Identity (``firm_id``, ``trade``) is NOT read from the sheet — the correlation ref (or
the operator's form) stays authoritative, exactly as on the model-parsed path.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Optional

from schemas.models import BidLineItem, BidReply

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Header-cell text -> BidLineItem field. Cells are normalised (lower-cased, any
# parenthesised suffix such as "(HKD)" stripped) before lookup.
_HEADER_FIELDS = {
    "item": "item_ref",
    "item ref": "item_ref",
    "item_ref": "item_ref",
    "ref": "item_ref",
    "description": "description",
    "unit": "unit",
    "qty": "qty",
    "quantity": "qty",
    "rate": "rate",
    "amount": "amount",
}
_PAREN_SUFFIX = re.compile(r"\s*\(.*\)\s*$")


def is_xlsx_upload(filename: Optional[str], content_type: Optional[str]) -> bool:
    """True when an upload is an xlsx (by content type or ``.xlsx`` extension)."""
    ct = (content_type or "").split(";")[0].strip().lower()
    return ct == XLSX_CONTENT_TYPE or (filename or "").strip().lower().endswith(".xlsx")


def _header_field(cell: object) -> Optional[str]:
    if not isinstance(cell, str):
        return None
    return _HEADER_FIELDS.get(_PAREN_SUFFIX.sub("", cell.strip().lower()))


def _map_header(row: tuple) -> Optional[dict[str, int]]:
    """Map a candidate header row to {field: column index}; None if it is not the header.

    The row qualifies only when it names both an item-ref column and a rate column —
    the two essentials of a Schedule of Rates. The title/note preamble never does.
    """
    columns: dict[str, int] = {}
    for index, cell in enumerate(row):
        field = _header_field(cell)
        if field is not None and field not in columns:
            columns[field] = index
    return columns if {"item_ref", "rate"} <= columns.keys() else None


def _text(value: object) -> Optional[str]:
    """A cell as trimmed text (``None`` when blank); numeric refs render cleanly."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _num(value: object) -> Optional[float]:
    """A cell as a number, tolerating blanks and typed-in text like ``2,450``."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None  # unparseable text stays unpriced — surfaced by leveling, never guessed


def parse_sor_xlsx(file_bytes: bytes, *, firm_id: str = "", trade: str = "") -> BidReply:
    """Parse a returned SoR sheet into a :class:`BidReply` deterministically (no LLM).

    Raises ``ValueError`` when the bytes are not a readable workbook or no sheet carries
    the SoR header row (Item + Rate columns) — the caller turns that into a 4xx.
    """
    from openpyxl import load_workbook  # lazy — mirrors export_xlsx / generate_sor_sheet

    if not file_bytes:
        raise ValueError("Empty file — nothing to parse.")
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:  # zip/xml errors from a mislabelled or corrupt upload
        raise ValueError(f"Could not read the xlsx file: {exc}") from exc

    try:
        for sheet in workbook.worksheets:
            columns: Optional[dict[str, int]] = None
            line_items: list[BidLineItem] = []
            for row in sheet.iter_rows(values_only=True):
                if columns is None:
                    columns = _map_header(row)
                    continue  # rows above/at the header are preamble
                cell = lambda field: row[columns[field]] if field in columns and columns[field] < len(row) else None  # noqa: E731
                item_ref = _text(cell("item_ref"))
                if item_ref is None:
                    continue  # totals / padding rows carry no ref — never invent one
                line_items.append(BidLineItem(
                    item_ref=item_ref,
                    description=_text(cell("description")),
                    unit=_text(cell("unit")),
                    qty=_num(cell("qty")),
                    rate=_num(cell("rate")),
                    amount=_num(cell("amount")),
                ))
            if columns is not None:
                return BidReply(firm_id=firm_id, trade=trade, line_items=line_items)
    finally:
        workbook.close()

    raise ValueError(
        "Not a Schedule of Rates sheet — no header row naming an Item and a Rate column "
        "was found. Return the dispatched SoR sheet with the Rate column filled."
    )
