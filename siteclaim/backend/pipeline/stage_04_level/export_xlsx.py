"""Excel export of the levelled comparison (openpyxl, imported lazily).

A multi-sheet workbook — ONE sheet per trade: each sheet carries a row per SoR item
with each of that trade's firms' rate and corrected amount, a totals row of each
firm's ``corrected_total``, and a Notes column calling out corrected lines and scope
gaps; stated exclusions are listed below the table. Two trades' items are never
merged into one table. Saved to ``backend/fixtures/out/leveling.xlsx`` by default.

openpyxl is imported inside the function so the leveling arithmetic (and DEMO_MODE)
never depend on it.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from rules_engine.leveling import computable_amount
from schemas.models import BidReply, LevelledBid, SorItem

OUT_PATH = Path(__file__).resolve().parents[2] / "fixtures" / "out" / "leveling.xlsx"


def sheet_title(trade: str) -> str:
    """A valid worksheet name for a package key ("mechanical_plumbing" -> "Mechanical
    Plumbing"; a section sub-package "ground_investigation:H" -> "Ground Investigation H"),
    capped at Excel's 31-char limit with its forbidden characters stripped."""
    base, _, section = (trade or "").partition(":")
    label = base.replace("_", " ").strip().title() or "Leveling"
    if section:
        label = f"{label} {section}"  # section code appended (upper-case preserved)
    for ch in ':\\/?*[]':  # Excel forbids these in a sheet name
        label = label.replace(ch, " ")
    return label[:31]


def export_leveling_xlsx(
    levelled: list[LevelledBid],
    replies: list[BidReply],
    item_order: Optional[list[str]] = None,
    path: Path | str = OUT_PATH,
    *,
    project_name: str = "",
    extras: Optional[list[str]] = None,
    awaiting: Optional[dict[str, list[str]]] = None,
    units: Optional[list[str]] = None,
    unit_items: Optional[dict[str, list[SorItem]]] = None,
) -> Path:
    """Write the levelled comparison to ``path`` and return it — one styled sheet per ROUTED UNIT
    (the dispatched enquiry), one rate column per firm whose active reply covers that unit, plus a
    Summary cover tab when more than one unit is compared.

    The sheet UNIVERSE is ``units`` when given (the tender's DISPATCHED units, so an enquiry whose
    return is empty or absent STILL gets its sheet), unioned with any levelled/reply unit; without
    ``units`` it is the levelled units (the reply-anchored behaviour). ``unit_items`` (unit key ->
    the unit's canonical SoR items in scope order) anchors each sheet on ITS canonical items — every
    one shown, a firm's rate where its return matched, a scope gap where not. ``awaiting`` (unit key
    -> firm labels enquired but not yet replied) is noted per sheet; ``extras`` (returned lines
    priced outside the tender's SoR — surfaced, never in a unit's totals) get an "Extras" tab."""
    from openpyxl import Workbook  # lazy — leveling math must not require openpyxl

    # The sheet universe: the dispatched units (so an unreplied enquiry still gets a sheet), then any
    # levelled/reply unit not already listed. Each unit is its own sheet — two units never merge.
    trades: list[str] = list(units or [])
    for b in levelled:
        if b.trade not in trades:
            trades.append(b.trade)

    wb = Workbook()
    first_used = False
    if not trades:  # nothing levelled and nothing dispatched — keep a single empty comparison sheet
        wb.active.title = "Leveling"
        first_used = True
    if len(trades) > 1:  # the cover tab: each trade's corrected totals at a glance
        ws = wb.active
        ws.title = "Summary"
        first_used = True
        _write_summary_sheet(ws, trades, levelled, project_name)
    for trade in trades:
        ws = wb.create_sheet() if first_used else wb.active
        first_used = True
        ws.title = sheet_title(trade)
        _write_trade_sheet(
            ws,
            [b for b in levelled if b.trade == trade],
            [r for r in replies if r.trade == trade],
            item_order,
            project_name,
            awaiting_firms=(awaiting or {}).get(trade),
            canonical_items=(unit_items or {}).get(trade),
        )

    if extras:  # out-of-scope returned lines — surfaced, never folded into a section
        ws = wb.create_sheet() if first_used else wb.active
        first_used = True
        ws.title = "Extras (out of scope)"
        _write_extras_sheet(ws, extras)

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _write_extras_sheet(ws, extras: list[str]) -> None:
    """A plain list of returned lines that matched no canonical SoR item — the operator sees the
    subcontractor priced something outside this tender; nothing here enters a comparison total."""
    from pipeline._xlsx_style import autofit, style_header, title_block

    title_block(ws, "Priced lines outside this tender's scope", [
        "These returned lines matched no Schedule-of-Rates item and are NOT included in any",
        "section's comparison or totals — review and assign or query with the subcontractor.",
    ])
    ws.append(["Out-of-scope returned line"])
    header_row = ws.max_row
    style_header(ws, header_row, 1)
    for note in extras:
        ws.append([note])
    autofit(ws, min_row=header_row)


def _write_summary_sheet(ws, trades: list[str], levelled: list[LevelledBid], project_name: str) -> None:
    """The multi-trade cover: one row per trade with its bid count and the lowest
    corrected total (every value straight from the Layer-1 leveling)."""
    import datetime as _dt

    from pipeline._xlsx_style import autofit, money_cell, style_body, style_header, title_block

    meta = [m for m in (
        f"Project: {project_name}" if project_name else "",
        f"Generated: {_dt.date.today().isoformat()}",
        "One comparison sheet per trade follows.",
    ) if m]
    title_block(ws, "Levelled bid comparison — summary", meta)

    ws.append(["Trade", "Bids", "Lowest corrected total", "Lowest bidder"])
    header_row = ws.max_row
    style_header(ws, header_row, 4)
    for trade in trades:
        bids = [b for b in levelled if b.trade == trade]
        if not bids:  # a dispatched unit with no return yet — listed, never crashed on min([])
            ws.append([sheet_title(trade), 0, "—", "awaiting return"])
            continue
        low = min(bids, key=lambda b: b.corrected_total)
        ws.append([sheet_title(trade), len(bids), low.corrected_total, low.firm_name])
        money_cell(ws.cell(row=ws.max_row, column=3))
    style_body(ws, header_row + 1, ws.max_row, 4)
    autofit(ws, min_row=header_row)


def _write_trade_sheet(
    ws,
    levelled: list[LevelledBid],
    replies: list[BidReply],
    item_order: Optional[list[str]],
    project_name: str = "",
    *,
    awaiting_firms: Optional[list[str]] = None,
    canonical_items: Optional[list[SorItem]] = None,
) -> None:
    """One routed unit's comparison table onto ``ws`` (this unit's firms/items only), one rate
    column per firm whose reply covers it — styled with the shared kit; the values are exactly the
    leveling output.

    When ``canonical_items`` is given (the unit's canonical SoR items) the sheet is ENQUIRY-anchored:
    every canonical item is a row, a firm's rate where its return matched (by NORMALISED ref — a
    routed line keeps its original ref form), a "scope gap (unpriced)" note where not — so a return
    that priced nothing still shows the full item set, and the sheet exists even with no return.
    Without it the sheet is reply-anchored (unchanged). ``awaiting_firms`` (enquired, not yet
    replied) is noted below the table."""
    import datetime as _dt

    from pipeline._xlsx_style import (
        autofit,
        footer_note,
        money_cell,
        style_body,
        style_header,
        style_totals,
        title_block,
    )
    from pipeline.stage_04_level.route_items import normalize_ref

    levelled_by_firm = {b.firm_id: b for b in levelled}
    firm_ids = [b.firm_id for b in levelled]

    if canonical_items is not None:
        # Enquiry-anchored: ALL of the unit's canonical items, in scope order, matched to a firm's
        # returned line by the SAME routing leveling uses — exact NORMALISED ref, then a description
        # fallback for a garbled/absent ref — so a line that priced a canonical item under a garbled
        # ref lands on its canonical row (not a phantom extra + a false scope-gap). A line matching no
        # canonical item is appended (defensive — real out-of-scope lines go to Extras).
        from pipeline.stage_04_level.route_items import DEFAULT_DESC_THRESHOLD, token_set_ratio

        canon_norms = {normalize_ref(it.item_ref) for it in canonical_items}
        desc_by_ref = {it.item_ref: (it.description or "") for it in canonical_items}

        def _resolved_norm(line) -> str:
            nr = normalize_ref(line.item_ref)
            if nr in canon_norms:
                return nr  # primary: exact normalised ref
            desc = line.description or ""
            if desc.strip():  # secondary: description token-set similarity (only when the ref missed)
                best, best_score = None, 0.0
                for it in canonical_items:
                    if not (it.description or "").strip():
                        continue
                    score = token_set_ratio(desc, it.description)
                    if score > best_score:
                        best, best_score = it, score
                if best is not None and best_score >= DEFAULT_DESC_THRESHOLD:
                    return normalize_ref(best.item_ref)
            return nr  # unmatched -> its own norm (rendered as an appended extra row)

        items = [it.item_ref for it in canonical_items]
        for reply in replies:
            for line in reply.line_items:
                if _resolved_norm(line) not in canon_norms and line.item_ref not in items:
                    items.append(line.item_ref)
        norm_by_ref = {ref: normalize_ref(ref) for ref in items}
        line_index = {
            (r.firm_id, _resolved_norm(line)): line for r in replies for line in r.line_items
        }
    else:
        # Reply-anchored (unchanged): the priced refs in scope order, then any extra the bids introduced.
        priced_refs = {line.item_ref for r in replies for line in r.line_items}
        items = [ref for ref in (item_order or []) if ref in priced_refs]
        for reply in replies:
            for line in reply.line_items:
                if line.item_ref not in items:
                    items.append(line.item_ref)
        canon_norms = set()
        desc_by_ref = {}
        norm_by_ref = {ref: ref for ref in items}  # identity keying
        line_index = {(r.firm_id, line.item_ref): line for r in replies for line in r.line_items}

    finding_items = {
        firm_id: {f.location.replace("line ", "") for f in b.arithmetic_findings}
        for firm_id, b in levelled_by_firm.items()
    }
    gap_items = {
        firm_id: {gap.split(" — ")[0] for gap in b.scope_gaps}
        for firm_id, b in levelled_by_firm.items()
    }

    label = sheet_title(ws.title if not levelled else levelled[0].trade)
    meta = [m for m in (
        f"Project: {project_name}" if project_name else "",
        f"Trade: {label}",
        f"Generated: {_dt.date.today().isoformat()}",
        "Rates are the primary comparison; amounts appear only where a quantity exists ('—' = rate-only line).",
    ) if m]
    title_block(ws, f"Levelled bid comparison — {label}", meta)

    header = ["Item", "Description"]
    for firm_id in firm_ids:
        name = levelled_by_firm[firm_id].firm_name
        header += [f"{name} — rate", f"{name} — corrected"]
    header.append("Notes")
    ws.append(header)
    header_row = ws.max_row
    style_header(ws, header_row, len(header))

    for item_ref in items:
        description = desc_by_ref.get(item_ref, "")
        is_canonical = normalize_ref(item_ref) in canon_norms if canon_norms else False
        row = [item_ref, ""]
        notes: list[str] = []
        for firm_id in firm_ids:
            line = line_index.get((firm_id, norm_by_ref.get(item_ref, item_ref)))
            if line is None:
                row += ["", ""]
                if is_canonical:  # a canonical item this firm did not return -> a scope gap on its sheet
                    notes.append(f"{levelled_by_firm[firm_id].firm_name}: scope gap (unpriced)")
                continue
            description = description or line.description or ""
            # Rate is the primary comparison and is always shown; the amount cell is filled
            # only where an amount is computable — a rate-only line shows the rate and "—".
            amount = computable_amount(line)
            row += [
                line.rate if line.rate is not None else "—",
                amount if amount is not None else "—",
            ]
            if line.item_ref in finding_items.get(firm_id, set()):
                notes.append(f"{levelled_by_firm[firm_id].firm_name}: arithmetic corrected")
            if line.item_ref in gap_items.get(firm_id, set()):
                notes.append(f"{levelled_by_firm[firm_id].firm_name}: scope gap (unpriced)")
        row[1] = description
        row.append("; ".join(notes))
        ws.append(row)
        for col in range(3, 3 + 2 * len(firm_ids)):  # rate + corrected columns are currency
            money_cell(ws.cell(row=ws.max_row, column=col))
    style_body(ws, header_row + 1, ws.max_row, len(header))

    totals = ["", "TOTAL (corrected)"]
    for firm_id in firm_ids:
        totals += ["", levelled_by_firm[firm_id].corrected_total]
    totals.append("")
    ws.append(totals)
    style_totals(ws, ws.max_row, len(header))
    for col in range(3, 3 + 2 * len(firm_ids)):
        money_cell(ws.cell(row=ws.max_row, column=col))

    ws.append([])
    footer_note(ws, "Exclusions (non-comparable — not deducted from price)")
    for firm_id in firm_ids:
        for exclusion in levelled_by_firm[firm_id].exclusions:
            ws.append([levelled_by_firm[firm_id].firm_name, exclusion])

    if awaiting_firms:  # enquired but not yet replied — coverage at a glance, not a priced column
        ws.append([])
        footer_note(ws, "Awaiting reply (enquired, no priced return yet)")
        for firm in awaiting_firms:
            ws.append([firm])

    autofit(ws, min_row=header_row)
