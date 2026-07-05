"""Shared Excel style kit (Prompt 4) — the Atlas palette on openpyxl.

Every workbook the platform writes (the levelled comparison, the outbound SoR sheet,
the benchmark templates) uses these helpers so the outputs read as one branded
document family: a navy header band with bold white text, a title/meta block above
each table, hairline borders, an emphasised totals row, HK$ currency cells, auto-fit
column widths and frozen panes below the header.

Presentation ONLY — nothing here adds, removes, or changes a data value. openpyxl is
imported lazily inside each helper so importing this module (and the modules that use
it) never requires openpyxl at runtime unless a workbook is actually written.
"""

from __future__ import annotations

# Atlas palette (hex without '#', as openpyxl expects).
NAVY = "0F1B2D"       # header fill — deep navy ink
INK_SOFT = "46566B"   # meta text
HAIRLINE = "E2E8F0"   # table borders
TINT = "EEF2F7"       # totals-row wash (light brand paper)
# Whole-dollar HKD. The literal must be quoted: an unquoted "H" is Excel's hour token,
# which turns the whole format into a date and corrupts the value on re-read.
MONEY_FORMAT = '"HK$"#,##0'


def thin_border():
    """A hairline border on all four sides."""
    from openpyxl.styles import Border, Side

    side = Side(style="thin", color=HAIRLINE)
    return Border(left=side, right=side, top=side, bottom=side)


def title_block(ws, title: str, meta: list[str]) -> None:
    """The document letterhead: a display title, small meta lines (project, reference,
    date, trade), then a blank spacer. Append-only — call before the table header."""
    from openpyxl.styles import Font

    ws.append([title])
    ws[ws.max_row][0].font = Font(bold=True, size=14, color=NAVY)
    for line in meta:
        ws.append([line])
        ws[ws.max_row][0].font = Font(size=10, color=INK_SOFT)
    ws.append([])


def style_header(ws, row_idx: int, n_cols: int) -> None:
    """The header band: navy fill, bold white text, hairline borders; panes freeze
    just below so the header stays visible while scrolling."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=NAVY)
    border = thin_border()
    for cell in ws[row_idx][:n_cols]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    # By coordinate string — ws.cell(...) would materialise the cell below the header
    # and advance the append cursor, injecting a phantom blank row into the table.
    ws.freeze_panes = f"A{row_idx + 1}"


def style_totals(ws, row_idx: int, n_cols: int) -> None:
    """The totals row: emphasised with a navy top rule, bold text, and the light
    brand tint."""
    from openpyxl.styles import Border, Font, PatternFill, Side

    hair = Side(style="thin", color=HAIRLINE)
    top = Side(style="medium", color=NAVY)
    fill = PatternFill("solid", fgColor=TINT)
    for cell in ws[row_idx][:n_cols]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = Border(top=top, left=hair, right=hair, bottom=hair)


def style_body(ws, first_row: int, last_row: int, n_cols: int) -> None:
    """Hairline borders across the table body rows."""
    border = thin_border()
    for row in ws.iter_rows(min_row=first_row, max_row=last_row, max_col=n_cols):
        for cell in row:
            cell.border = border


def money_cell(cell) -> None:
    """Format one cell as whole-dollar HKD currency (display only — the value is
    untouched; a '—' rate-only placeholder passes through as text)."""
    if isinstance(cell.value, (int, float)):
        cell.number_format = MONEY_FORMAT


def footer_note(ws, text: str) -> None:
    """A small footer note line (the notes/exclusions preface the source writes)."""
    from openpyxl.styles import Font

    ws.append([text])
    ws[ws.max_row][0].font = Font(bold=True, size=10, color=INK_SOFT)


def autofit(ws, *, min_row: int = 1, max_width: int = 58) -> None:
    """Fit each column to its longest cell value from ``min_row`` down (pass the header
    row so a long title/meta line never balloons column A), capped at ``max_width``."""
    from openpyxl.utils import get_column_letter

    widths: dict[int, int] = {}
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            if cell.value is None:
                continue
            longest = max(len(line) for line in str(cell.value).split("\n"))
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max_width, max(9, width + 2))
