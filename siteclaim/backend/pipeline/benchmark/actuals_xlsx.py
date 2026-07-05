"""The Final Account template + its deterministic parser (Phase B1c).

We author the template so parsing needs no model — pure openpyxl, mirroring the
reply-xlsx discipline (header located by column names, tolerant of blank cells and
typed-in numbers). It differs from the reply parser in two ways the benchmark needs:
it captures a **Section** column, and it detects **granularity** per row
(`item` | `section` | `project`) — a final account may be item-by-item,
section-totals-only, or a single project total. Cost data stays local; no network.
"""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Optional

# The authored column order (header row). "Section" is the benchmark addition.
TEMPLATE_HEADERS = ["Item", "Description", "Unit", "Qty", "Rate (HKD)", "Amount (HKD)", "Section"]

_HEADER_FIELDS = {
    "item": "item_ref", "item ref": "item_ref", "item_ref": "item_ref", "ref": "item_ref",
    "description": "description", "unit": "unit",
    "qty": "qty", "quantity": "qty",
    "rate": "rate", "amount": "amount",
    "section": "section", "part": "section",
}
_PAREN_SUFFIX = re.compile(r"\s*\(.*\)\s*$")


def _header_field(cell: object) -> Optional[str]:
    if not isinstance(cell, str):
        return None
    return _HEADER_FIELDS.get(_PAREN_SUFFIX.sub("", cell.strip().lower()))


def _map_header(row: tuple) -> Optional[dict[str, int]]:
    """A header row for a Final Account qualifies when it names an identifier column
    (item or section) and at least one numeric column (rate, amount, or qty)."""
    columns: dict[str, int] = {}
    for index, cell in enumerate(row):
        field = _header_field(cell)
        if field is not None and field not in columns:
            columns[field] = index
    has_identifier = bool(columns.keys() & {"item_ref", "section"})
    has_numeric = bool(columns.keys() & {"rate", "amount", "qty"})
    return columns if (has_identifier and has_numeric) else None


def _text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _num(value: object) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _granularity(item_ref: Optional[str], section: Optional[str]) -> str:
    if item_ref:
        return "item"
    if section:
        return "section"
    return "project"


def parse_actuals_xlsx(file_bytes: bytes) -> list[dict]:
    """Parse a Final Account xlsx into actual-item rows, each with a detected granularity.

    Raises ``ValueError`` if the bytes are not a readable workbook or no sheet carries the
    Final Account header (an identifier + a numeric column) — the caller turns that into a
    400. Blank/preamble rows (no identifier and no numbers) are skipped.
    """
    from openpyxl import load_workbook  # lazy

    if not file_bytes:
        raise ValueError("Empty file — nothing to parse.")
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:  # zip/xml errors from a mislabelled or corrupt upload
        raise ValueError(f"Could not read the xlsx file: {exc}") from exc

    try:
        for sheet in workbook.worksheets:
            columns: Optional[dict[str, int]] = None
            rows_out: list[dict] = []
            for row in sheet.iter_rows(values_only=True):
                if columns is None:
                    columns = _map_header(row)
                    continue
                cell = lambda f: row[columns[f]] if f in columns and columns[f] < len(row) else None  # noqa: E731
                item_ref = _text(cell("item_ref"))
                section = _text(cell("section"))
                qty, rate, amount = _num(cell("qty")), _num(cell("rate")), _num(cell("amount"))
                # A row with no identifier and no numbers is preamble/blank — skip it.
                if not item_ref and not section and qty is None and rate is None and amount is None:
                    continue
                rows_out.append({
                    "item_ref": item_ref or "", "description": _text(cell("description")) or "",
                    "unit": _text(cell("unit")) or "", "qty": qty, "rate": rate, "amount": amount,
                    "section": section or "", "granularity": _granularity(item_ref, section),
                })
            if columns is not None:
                return rows_out
    finally:
        workbook.close()

    raise ValueError(
        "Not a Final Account sheet — no header row naming an Item/Section column and a "
        "Qty/Rate/Amount column was found. Use the benchmark actuals template."
    )


def build_actuals_template(project_name: str, tender_items: list[dict], path: Path | str) -> Path:
    """Write the Final Account template to ``path``. When ``tender_items`` are given, the
    item_ref / description / unit / section columns are pre-filled so the operator only
    types the actual qty / rate / amount. Styled with the shared kit (presentation only —
    the filled template still parses deterministically via :func:`parse_actuals_xlsx`)."""
    from openpyxl import Workbook  # lazy

    from pipeline._xlsx_style import autofit, style_body, style_header, title_block

    wb = Workbook()
    ws = wb.active
    ws.title = "Final Account"

    title_block(ws, f"Final Account (actual outturn) — {project_name}", [
        f"Reference: {project_name}",
        "Enter the ACTUAL quantity, rate and amount for each item. Rate-only lines are fine "
        "(leave qty blank). For a section-totals-only account, leave Item blank and fill "
        "Section + Amount. Do not change the header row.",
    ])

    ws.append(TEMPLATE_HEADERS)
    header_row = ws.max_row
    style_header(ws, header_row, len(TEMPLATE_HEADERS))

    for it in tender_items:
        ws.append([
            it.get("item_ref", ""), it.get("description", ""), it.get("unit", ""),
            "", "", "",  # qty / rate / amount — the operator fills these
            it.get("section", ""),
        ])
    style_body(ws, header_row + 1, ws.max_row, len(TEMPLATE_HEADERS))
    autofit(ws, min_row=header_row)

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
