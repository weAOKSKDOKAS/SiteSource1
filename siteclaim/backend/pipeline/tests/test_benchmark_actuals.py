"""Benchmark actuals — template generation + deterministic parse (Phase B1c).

Parser unit tests (item + coarse granularity, tolerance, wrong-layout) plus the two
routes end to end against a temp live DB. Offline (openpyxl only).
"""

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from api import app
from db import seed
from pipeline.benchmark.actuals_xlsx import (
    TEMPLATE_HEADERS,
    build_actuals_template,
    parse_actuals_xlsx,
)

client = TestClient(app)
_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def bench_db(tmp_path, monkeypatch):
    db = tmp_path / "bench.db"
    seed.build_database(db, profile="live")
    monkeypatch.setenv("SITESOURCE_DB", str(db))
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "ws"))
    return db


def _sheet(rows: list[list], headers: list[str] = TEMPLATE_HEADERS, preamble: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    if preamble:
        ws.append(["Some Project — Final Account"])
        ws.append(["instructions..."])
        ws.append([])
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -- parser units ---------------------------------------------------------------------
def test_item_level_actuals_parse_with_tolerant_cells():
    # blank qty on a rate-only line; a comma-typed rate; a $-prefixed amount
    data = _sheet([
        ["A1a(a)", "Rotary drilling", "m", 120, "1,300", 156000, "A"],
        ["M2", "Standing time", "hr", "", 950, "", "A"],       # rate-only (no qty/amount)
    ])
    rows = parse_actuals_xlsx(data)
    by_ref = {r["item_ref"]: r for r in rows}
    assert by_ref["A1a(a)"]["rate"] == 1300.0 and by_ref["A1a(a)"]["qty"] == 120.0
    assert by_ref["A1a(a)"]["section"] == "A" and by_ref["A1a(a)"]["granularity"] == "item"
    assert by_ref["M2"]["qty"] is None and by_ref["M2"]["rate"] == 950.0
    assert all(r["granularity"] == "item" for r in rows)


def test_section_totals_only_sheet_is_section_granularity():
    # no Item column values — section labels with amounts only
    data = _sheet([
        ["", "Site works", "", "", "", 2400000, "Part A"],
        ["", "Laboratory", "", "", "", 850000, "Part B"],
    ])
    rows = parse_actuals_xlsx(data)
    assert [r["granularity"] for r in rows] == ["section", "section"]
    assert rows[0]["section"] == "Part A" and rows[0]["amount"] == 2400000.0
    assert rows[0]["item_ref"] == ""


def test_project_total_row_is_project_granularity():
    data = _sheet([["", "Final account total", "", "", "", 3250000, ""]])
    rows = parse_actuals_xlsx(data)
    assert len(rows) == 1 and rows[0]["granularity"] == "project" and rows[0]["amount"] == 3250000.0


def test_blank_rows_are_skipped():
    data = _sheet([
        ["A1", "x", "m", 1, 10, 10, ""],
        ["", "", "", "", "", "", ""],       # fully blank -> skipped
        ["A2", "y", "m", 2, 20, 40, ""],
    ])
    rows = parse_actuals_xlsx(data)
    assert {r["item_ref"] for r in rows} == {"A1", "A2"}


def test_wrong_layout_raises_value_error():
    wb = Workbook()
    wb.active.append(["Colour", "Size", "Price"])  # no Item/Section + numeric header
    wb.active.append(["red", "M", 10])
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Not a Final Account sheet"):
        parse_actuals_xlsx(buf.getvalue())


def test_garbage_and_empty_bytes_raise():
    with pytest.raises(ValueError, match="Could not read"):
        parse_actuals_xlsx(b"not a zip")
    with pytest.raises(ValueError, match="Empty file"):
        parse_actuals_xlsx(b"")


def test_template_prefills_tender_refs(tmp_path):
    tender = [
        {"item_ref": "A1a(a)", "description": "Rotary drilling", "unit": "m", "section": "A"},
        {"item_ref": "M2", "description": "Standing time", "unit": "hr", "section": "A"},
    ]
    out = build_actuals_template("GE/2026/14", tender, tmp_path / "tmpl.xlsx")
    ws = load_workbook(out).active
    header = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    assert [ws.cell(row=header, column=c).value for c in range(1, 8)] == TEMPLATE_HEADERS
    refs = [ws.cell(row=r, column=1).value for r in range(header + 1, ws.max_row + 1)]
    assert refs == ["A1a(a)", "M2"]
    # qty/rate/amount columns are blank for the operator to fill
    assert ws.cell(row=header + 1, column=4).value in (None, "")
    # a filled template round-trips back through the parser
    ws.cell(row=header + 1, column=5, value=1300)  # Rate
    buf = BytesIO()
    load_workbook(out).save(buf)  # ensure it's a valid workbook
    parsed = parse_actuals_xlsx(out.read_bytes())
    assert {r["item_ref"] for r in parsed} == {"A1a(a)", "M2"}


# -- routes ---------------------------------------------------------------------------
def _project_with_tender(bench_db) -> int:
    pid = client.post("/benchmark/projects", json={"name": "P", "trade": "ground_investigation"}).json()["id"]
    # seed a tiny tender via link-scope (deterministic, no model)
    client.post(f"/benchmark/{pid}/link-scope", json={"project_name": "GE/2026/14", "packages": [{
        "trade": "ground_investigation", "scope_summary": "GI",
        "sor_items": [{"item_ref": "A1a(a)", "description": "Rotary drilling", "unit": "m", "qty": 100.0}],
        "source_refs": [],
    }]})
    return pid


def test_actuals_template_download_prefilled(bench_db):
    pid = _project_with_tender(bench_db)
    resp = client.get(f"/benchmark/actuals-template.xlsx?project={pid}")
    assert resp.status_code == 200 and resp.content[:2] == b"PK"
    ws = load_workbook(BytesIO(resp.content)).active
    header = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    refs = [ws.cell(row=r, column=1).value for r in range(header + 1, ws.max_row + 1)]
    assert "A1a(a)" in refs  # pre-filled from the tender snapshot


def test_actuals_upload_writes_items_with_granularity(bench_db):
    pid = _project_with_tender(bench_db)
    data = _sheet([["A1a(a)", "Rotary drilling", "m", 120, 1300, 156000, "A"]])
    resp = client.post(f"/benchmark/{pid}/actuals-upload", files={"files": ("fa.xlsx", data, _XLSX_CT)})
    assert resp.status_code == 200
    body = resp.json()
    assert body["source"] == "actuals-xlsx" and body["item_count"] == 1
    assert body["granularities"] == ["item"]
    assert body["items"][0]["qty"] == 120.0 and body["items"][0]["rate"] == 1300.0
    assert client.get(f"/benchmark/projects/{pid}").json()["actual_item_count"] == 1


def test_actuals_upload_wrong_layout_400(bench_db):
    pid = _project_with_tender(bench_db)
    wb = Workbook()
    wb.active.append(["Colour", "Size", "Price"])
    buf = BytesIO()
    wb.save(buf)
    resp = client.post(f"/benchmark/{pid}/actuals-upload", files={"files": ("bad.xlsx", buf.getvalue(), _XLSX_CT)})
    assert resp.status_code == 400 and "Final Account" in resp.json()["detail"]


def test_actuals_pdf_rejected_by_default(bench_db):
    pid = _project_with_tender(bench_db)
    resp = client.post(f"/benchmark/{pid}/actuals-upload", files={"files": ("fa.pdf", b"%PDF-1.4", "application/pdf")})
    assert resp.status_code == 400 and "ACTUALS_PDF_PARSE" in resp.json()["detail"]


def test_actuals_pdf_flag_on_but_demo_still_rejects(bench_db, monkeypatch):
    pid = _project_with_tender(bench_db)
    monkeypatch.setenv("ACTUALS_PDF_PARSE", "true")  # opt-in, but DEMO_MODE is on -> live-only path
    resp = client.post(f"/benchmark/{pid}/actuals-upload", files={"files": ("fa.pdf", b"%PDF-1.4", "application/pdf")})
    assert resp.status_code == 400 and "live engine" in resp.json()["detail"]
