"""Prompt 1 — the per-section (per-trade) leveling + recommend path.

/level-all groups replies by trade and levels each trade ONLY against its own bids
(the peer item reference never crosses trades); /recommend-all produces one
risk-adjusted recommendation per trade. The Excel refresh becomes a multi-sheet
workbook, one sheet per trade. All offline under the DEMO autouse fixture.
"""

import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api import app
from pipeline.stage_04_level.export_xlsx import export_leveling_xlsx, sheet_title
from pipeline.stage_04_level.level import load_demo_replies

client = TestClient(app)

_FIXTURES = Path(__file__).resolve().parents[2] / "fixtures"


@pytest.fixture
def two_trade_replies() -> list[dict]:
    """A two-trade reply set: the hero electrical bids + the clean joinery bids."""
    replies = load_demo_replies("cases/scenarios/hero_replies.json") + load_demo_replies(
        "cases/scenarios/clean_replies.json"
    )
    return [r.model_dump() for r in replies]


def test_level_all_groups_by_trade_without_mixing(two_trade_replies):
    body = client.post("/level-all", json={"replies": two_trade_replies}).json()
    sections = body["sections"]
    assert [s["trade"] for s in sections] == ["electrical", "joinery_fitting_out"]

    for section in sections:
        # every bid in a section belongs to that section's trade — no cross-trade mixing
        assert all(b["trade"] == section["trade"] for b in section["levelled"])
        # the peer item reference never crossed trades: no scope gap cites the OTHER
        # trade's item refs (a mixed peer set would gap every firm on the other trade's refs)
        foreign_prefix = "J-" if section["trade"] == "electrical" else "E-"
        for bid in section["levelled"]:
            assert not any(gap.startswith(foreign_prefix) for gap in bid["scope_gaps"])

    # per-trade leveling equals leveling each trade alone (grouping changes nothing else)
    solo = client.post("/level", json={"replies": [r for r in two_trade_replies if r["trade"] == "electrical"]}).json()
    assert solo == sections[0]["levelled"]


def test_recommend_all_runs_per_trade(two_trade_replies):
    sections = client.post("/level-all", json={"replies": two_trade_replies}).json()["sections"]
    flat = [b for s in sections for b in s["levelled"]]

    hero_fixture = "cases/scenarios/hero_rationale.json"
    body = client.post(
        "/recommend-all",
        json={"levelled": flat, "demo_fixtures": {"electrical": hero_fixture}},
    ).json()
    recs = {s["trade"]: s["recommendation"] for s in body["sections"]}
    assert set(recs) == {"electrical", "joinery_fitting_out"}

    for trade, rec in recs.items():
        assert rec["trade"] == trade
        trade_firms = {b["firm_id"] for b in flat if b["trade"] == trade}
        assert {r["firm_id"] for r in rec["ranked"]} == trade_firms  # ranked within the trade only
        assert rec["recommended_firm_id"] in trade_firms
        assert rec["rationale"]  # narrated (fixture or offline template) — never empty

    # the per-trade fixture map lands on its trade: electrical narrates from the baked
    # rationale; joinery (no entry) falls back to the deterministic offline template
    baked = json.loads((_FIXTURES / hero_fixture).read_text(encoding="utf-8"))["text"]
    assert recs["electrical"]["rationale"] == baked
    assert recs["joinery_fitting_out"]["rationale"] != baked


def test_leveling_workbook_has_one_sheet_per_trade(tmp_path, two_trade_replies):
    sections = client.post("/level-all", json={"replies": two_trade_replies}).json()["sections"]
    # rebuild the workbook directly so the assertion targets the exported file
    from schemas.models import BidReply, LevelledBid

    levelled = [LevelledBid.model_validate(b) for s in sections for b in s["levelled"]]
    replies = [BidReply.model_validate(r) for r in two_trade_replies]
    out = export_leveling_xlsx(levelled, replies, path=tmp_path / "multi.xlsx")

    wb = load_workbook(out)
    assert wb.sheetnames == [sheet_title("electrical"), sheet_title("joinery_fitting_out")]
    # each sheet prices only its own trade's item refs
    for name, prefix in ((sheet_title("electrical"), "E-"), (sheet_title("joinery_fitting_out"), "J-")):
        refs = [row[0].value for row in wb[name].iter_rows(min_row=2) if row[0].value]
        item_refs = [r for r in refs if isinstance(r, str) and "-" in r and len(r) <= 6]
        assert item_refs and all(r.startswith(prefix) for r in item_refs)


def test_single_trade_scenarios_unchanged():
    # the existing single-trade path (/level + /recommend) is untouched
    replies = [r.model_dump() for r in load_demo_replies("cases/scenarios/hero_replies.json")]
    levelled = client.post("/level", json={"replies": replies}).json()
    assert {b["trade"] for b in levelled} == {"electrical"}
    rec = client.post(
        "/recommend",
        json={"levelled": levelled, "trade": "electrical", "demo_fixture": "cases/scenarios/hero_rationale.json"},
    ).json()
    assert rec["trade"] == "electrical" and rec["recommended_firm_id"]
