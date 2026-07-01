"""Document routing and the generated SoR sheet (Phase A, build plan §5)."""

from pathlib import Path

from openpyxl import load_workbook

from pipeline.stage_03_dispatch.attachments import (
    build_attachments,
    generate_sor_sheet,
    route_documents,
)
from pipeline.workspace import Workspace
from schemas.models import (
    AttachmentKind,
    DocType,
    ScopePackages,
    SorItem,
    TenderDocument,
    TenderPackage,
    TradeWorkPackage,
)


def _scope() -> ScopePackages:
    return ScopePackages(project_name="P", packages=[
        TradeWorkPackage(trade="electrical", scope_summary="LV works", sor_items=[
            SorItem(item_ref="E1", description="LV board", unit="no", qty=2),
            SorItem(item_ref="E2", description="Final circuits", unit="m", qty=100),
        ]),
        TradeWorkPackage(trade="fire_services", scope_summary="Sprinklers", sor_items=[
            SorItem(item_ref="F1", description="Sprinkler heads", unit="no", qty=50),
        ]),
    ])


def _tender() -> TenderPackage:
    return TenderPackage(project_name="P", documents=[
        TenderDocument(doc_type=DocType.METHOD_OF_MEASUREMENT, filename="mom.pdf"),  # general
        TenderDocument(doc_type=DocType.PARTICULAR_SPECIFICATION, filename="elec_spec.pdf", trades=["electrical"]),
    ])


def test_sor_sheet_is_a_real_labelled_xlsx(tmp_path):
    path = generate_sor_sheet(_scope().packages[0], "P", tmp_path / "sor.xlsx")
    assert path.is_file()
    assert path.read_bytes()[:2] == b"PK"  # xlsx is a zip
    rows = list(load_workbook(path).active.iter_rows(values_only=True))
    flat = [str(c) for r in rows for c in r if c is not None]
    assert "E1" in flat and "E2" in flat            # the priceable items are present
    assert any("excerpt" in cell.lower() for cell in flat)  # labelled an excerpt (§5 safety rule)
    header = next(r for r in rows if r and r[0] == "Item")
    assert "Rate (HKD)" in header and "Amount (HKD)" in header  # blank columns to price


def test_route_documents_splits_general_and_trade_specific():
    general, specific = route_documents(_tender(), "electrical")
    assert [d.filename for d in general] == ["mom.pdf"]
    assert [d.filename for d in specific] == ["elec_spec.pdf"]
    # fire services gets the general document but not the electrical-only spec
    g2, s2 = route_documents(_tender(), "fire_services")
    assert [d.filename for d in g2] == ["mom.pdf"]
    assert s2 == []


def test_build_attachments_without_workspace_describes_only():
    atts = build_attachments("electrical", _scope(), _tender(), project_name="P")
    kinds = {a.kind for a in atts}
    assert {AttachmentKind.GENERAL, AttachmentKind.TRADE_SPECIFIC, AttachmentKind.SOR_SHEET} <= kinds
    assert all(a.source_path is None for a in atts)  # nothing written to disk


def test_build_attachments_with_workspace_generates_the_sor_sheet(tmp_path):
    ws = Workspace(root=tmp_path)
    atts = build_attachments("electrical", _scope(), _tender(), project_name="P", tender_id="P", workspace=ws)
    sor = next(a for a in atts if a.kind is AttachmentKind.SOR_SHEET)
    assert sor.source_path and Path(sor.source_path).is_file()
    # a general doc with no uploaded original stays described but path-less
    general = next(a for a in atts if a.kind is AttachmentKind.GENERAL)
    assert general.source_path is None


def test_build_attachments_resolves_an_uploaded_original(tmp_path):
    ws = Workspace(root=tmp_path)
    ws.save_upload("P", "mom.pdf", b"%PDF-1.4 fake")
    atts = build_attachments("electrical", _scope(), _tender(), project_name="P", tender_id="P", workspace=ws)
    general = next(a for a in atts if a.kind is AttachmentKind.GENERAL and a.filename == "mom.pdf")
    assert general.source_path and Path(general.source_path).is_file()
