"""Benchmark estimator — project CRUD + tender snapshot (Phase B1b), through the API.

Every write targets a temp LIVE profile DB via SITESOURCE_DB, so the committed
sitesource.db is never touched. Runs under the DEMO autouse fixture (offline).
"""

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api import app
from db import seed

client = TestClient(app)


@pytest.fixture
def bench_db(tmp_path, monkeypatch):
    db = tmp_path / "bench.db"
    seed.build_database(db, profile="live")
    monkeypatch.setenv("SITESOURCE_DB", str(db))
    return db


def _sor_sheet_bytes(rows: list[tuple], with_rates: bool = True) -> bytes:
    """A minimal SoR-sheet xlsx (our layout): title/note/blank, header, then item rows."""
    from pipeline.stage_03_dispatch.attachments import generate_sor_sheet
    from schemas.models import SorItem, TradeWorkPackage

    pkg = TradeWorkPackage(
        trade="ground_investigation", scope_summary="GI",
        sor_items=[SorItem(item_ref=r[0], description=r[1], unit=r[2], qty=r[3]) for r in rows],
        source_refs=["SR-01"],
    )
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
        path = generate_sor_sheet(pkg, "GE/2026/14", fh.name)
    wb = load_workbook(path)
    ws = wb.active
    header = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    if with_rates:
        rates = {r[0]: r[4] for r in rows}
        for rr in range(header + 1, ws.max_row + 1):
            ref = ws.cell(row=rr, column=1).value
            if ref in rates:
                ws.cell(row=rr, column=5, value=rates[ref])  # "Rate (HKD)"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_benchmark_routes_are_registered():
    paths = {r.path for r in app.routes}
    assert {"/benchmark/projects", "/benchmark/projects/{project_id}",
            "/benchmark/{project_id}/tender-upload", "/benchmark/{project_id}/link-scope"} <= paths


def test_project_crud_lifecycle(bench_db):
    created = client.post("/benchmark/projects", json={
        "name": "GI Term Contract 2026", "trade": "ground_investigation", "contract_ref": "GE/2026/14",
    }).json()
    assert created["id"] >= 1 and created["status"] == "open" and created["provenance"] == "live"
    pid = created["id"]

    listing = client.get("/benchmark/projects").json()
    assert any(p["id"] == pid for p in listing)

    got = client.get(f"/benchmark/projects/{pid}").json()
    assert got["contract_ref"] == "GE/2026/14" and got["tender_item_count"] == 0

    # rename + close
    patched = client.patch(f"/benchmark/projects/{pid}", json={"name": "Renamed", "status": "closed"}).json()
    assert patched["name"] == "Renamed" and patched["status"] == "closed" and patched["closed_at"]

    assert client.get("/benchmark/projects/9999").status_code == 404


def test_patch_rejects_an_out_of_vocabulary_status(bench_db):
    created = client.post("/benchmark/projects", json={"name": "P"}).json()
    pid = created["id"]
    closed = client.patch(f"/benchmark/projects/{pid}", json={"status": "closed"}).json()
    assert closed["status"] == "closed" and closed["closed_at"]
    # a bogus status is rejected (400) and does not un-stamp closed_at
    bad = client.patch(f"/benchmark/projects/{pid}", json={"status": "banana"})
    assert bad.status_code == 400
    assert client.get(f"/benchmark/projects/{pid}").json()["closed_at"] == closed["closed_at"]
    # reopening clears the stamp
    reopened = client.patch(f"/benchmark/projects/{pid}", json={"status": "open"}).json()
    assert reopened["status"] == "open" and reopened["closed_at"] == ""


def test_tender_upload_xlsx_keeps_rates(bench_db):
    pid = client.post("/benchmark/projects", json={"name": "P", "trade": "ground_investigation"}).json()["id"]
    xlsx = _sor_sheet_bytes([("A1a(a)", "Rotary drilling", "m", 100.0, 1200.0),
                             ("M2", "Standing time", "hr", 40.0, 950.0)])
    resp = client.post(
        f"/benchmark/{pid}/tender-upload",
        files={"files": ("tender.xlsx", xlsx, _XLSX_CT)},
        data={"source_doc": "GE-2026-14-priced.xlsx"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["source"] == "tender-xlsx" and body["item_count"] == 2
    by_ref = {i["item_ref"]: i for i in body["items"]}
    assert by_ref["A1a(a)"]["rate"] == 1200.0 and by_ref["A1a(a)"]["qty"] == 100.0
    assert by_ref["A1a(a)"]["source"] == "tender-xlsx"
    # reflected on the project
    assert client.get(f"/benchmark/projects/{pid}").json()["tender_item_count"] == 2


def test_tender_upload_replaces_not_appends(bench_db):
    pid = client.post("/benchmark/projects", json={"name": "P"}).json()["id"]
    first = _sor_sheet_bytes([("A1", "x", "m", 1.0, 10.0)])
    client.post(f"/benchmark/{pid}/tender-upload", files={"files": ("t.xlsx", first, _XLSX_CT)})
    second = _sor_sheet_bytes([("B1", "y", "m", 2.0, 20.0), ("B2", "z", "m", 3.0, 30.0)])
    resp = client.post(f"/benchmark/{pid}/tender-upload", files={"files": ("t.xlsx", second, _XLSX_CT)}).json()
    assert resp["item_count"] == 2 and {i["item_ref"] for i in resp["items"]} == {"B1", "B2"}  # replaced


def test_tender_upload_pdf_in_demo_is_rejected(bench_db):
    pid = client.post("/benchmark/projects", json={"name": "P"}).json()["id"]
    resp = client.post(f"/benchmark/{pid}/tender-upload",
                       files={"files": ("t.pdf", b"%PDF-1.4", "application/pdf")})
    assert resp.status_code == 400 and "live engine" in resp.json()["detail"]


def test_link_scope_captures_unpriced_tender_items(bench_db):
    pid = client.post("/benchmark/projects", json={"name": "P"}).json()["id"]
    scope = {"project_name": "GE/2026/14", "packages": [{
        "trade": "ground_investigation", "scope_summary": "GI",
        "sor_items": [{"item_ref": "A1", "description": "drilling", "unit": "m", "qty": 100.0},
                      {"item_ref": "A2", "description": "sampling", "unit": "no"}],
        "source_refs": [],
    }]}
    resp = client.post(f"/benchmark/{pid}/link-scope", json=scope)
    assert resp.status_code == 200
    body = resp.json()
    assert body["source"] == "pipeline-link" and body["item_count"] == 2
    a1 = next(i for i in body["items"] if i["item_ref"] == "A1")
    assert a1["rate"] is None and a1["qty"] == 100.0 and a1["section"] == "ground_investigation"


def test_tender_upload_to_unknown_project_404s(bench_db):
    xlsx = _sor_sheet_bytes([("A1", "x", "m", 1.0, 10.0)])
    assert client.post("/benchmark/123456/tender-upload",
                       files={"files": ("t.xlsx", xlsx, _XLSX_CT)}).status_code == 404
