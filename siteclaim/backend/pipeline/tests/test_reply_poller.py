"""The Gmail reply poller (inbound n8n-trigger replacement) — fully offline: a stubbed Gmail
service, a scripted processor, no Google SDK, no socket. The poller is transport only: it feeds
the SAME shared processing path the /inbound-reply route runs (proven by the agreement test)."""

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from pipeline import reply_loop, reply_poller
from pipeline.tests.test_gmail_client import StubService, _b64url
from pipeline.workspace import Workspace


def _message(mid: str, ref: str, filename: str = "return.xlsx", data: bytes = b"PK-bytes") -> tuple[dict, dict]:
    """(list entry, get_results payload) for one reply message carrying ``ref`` + one attachment.
    The single payload serves both the metadata read (headers) and the attachment walk (parts)."""
    payload = {"payload": {
        "headers": [{"name": "Subject", "value": f"Re: RFQ [SiteSource Ref: {ref}]"}],
        "filename": "", "body": {},
        "parts": [{"filename": filename, "body": {"data": _b64url(data)}}],
    }}
    return {"id": mid}, payload


def _service(*messages: tuple[dict, dict]) -> StubService:
    return StubService(
        list_result={"messages": [m for m, _p in messages]},
        get_results={m["id"]: p for m, p in messages},
    )


class Recorder:
    """A scripted processor: records every (ref, attachments) call, returns a fixed status."""

    def __init__(self, status="matched"):
        self.calls: list[tuple[str, list]] = []
        self.status = status

    def __call__(self, ref, attachments):
        self.calls.append((ref, attachments))
        return self.status


def test_poll_once_feeds_each_new_message_to_the_processor_and_persists_the_id(tmp_path):
    ws = Workspace(tmp_path)
    svc = _service(_message("m1", "t.f1.gi:H", data=b"xlsx-1"))
    rec = Recorder()
    summary = reply_poller.poll_once(rec, workspace=ws, service=svc)
    assert summary == {"found": 1, "processed": 1, "skipped": 0, "unmatched": 0, "failed": 0}
    assert rec.calls == [("t.f1.gi:H", [("return.xlsx", b"xlsx-1")])]  # ref + real bytes handed over
    assert reply_poller.load_processed(ws)["m1"]["status"] == "matched"  # persisted for idempotency


def test_poll_once_dedupes_on_the_persisted_message_id(tmp_path):
    # The idempotency requirement: the poller re-reads the inbox, so a processed message_id is
    # skipped on every later sweep — including after a restart (the file persists).
    ws = Workspace(tmp_path)
    svc = _service(_message("m1", "t.f1.gi:H"))
    rec = Recorder()
    reply_poller.poll_once(rec, workspace=ws, service=svc)
    summary2 = reply_poller.poll_once(rec, workspace=ws, service=svc)
    assert len(rec.calls) == 1                                    # processed exactly once
    assert summary2 == {"found": 1, "processed": 0, "skipped": 1, "unmatched": 0, "failed": 0}


def test_an_unresolvable_ref_is_reported_never_dropped(tmp_path):
    ws = Workspace(tmp_path)
    svc = _service(_message("m9", "unknown.ref.x"))
    summary = reply_poller.poll_once(Recorder(status="unmatched"), workspace=ws, service=svc)
    assert summary["unmatched"] == 1 and summary["processed"] == 0
    assert reply_poller.load_processed(ws)["m9"]["status"] == "unmatched"   # surfaced on the record
    assert reply_poller.poller_state()["unmatched_total"] >= 1              # and on the status surface


def test_gmail_down_records_the_error_and_never_raises(tmp_path):
    class DownService(StubService):
        def list(self, userId, q, maxResults):
            return type("C", (), {"execute": lambda self_: (_ for _ in ()).throw(RuntimeError("offline"))})()

    ws = Workspace(tmp_path)
    summary = reply_poller.poll_once(Recorder(), workspace=ws, service=DownService())
    assert summary == {"found": 0, "processed": 0, "skipped": 0, "unmatched": 0, "failed": 0}
    assert "offline" in reply_poller.poller_state()["last_error"]


def test_a_poisoned_message_is_recorded_failed_and_does_not_stall_the_sweep(tmp_path):
    ws = Workspace(tmp_path)
    svc = _service(_message("bad", "t.f1.gi:H"), _message("good", "t.f2.gi:H"))

    def flaky(ref, attachments):
        if ref == "t.f1.gi:H":
            raise ValueError("unreadable attachment")
        return "matched"

    summary = reply_poller.poll_once(flaky, workspace=ws, service=svc)
    assert summary["failed"] == 1 and summary["processed"] == 1   # the good message still lands
    processed = reply_poller.load_processed(ws)
    assert processed["bad"]["status"].startswith("error:")        # visible reason, not silence
    assert processed["good"]["status"] == "matched"


def test_ref_extraction_from_a_reply_subject():
    assert reply_poller.ref_from_subject("Re: RFQ — GI [SiteSource Ref: ge-2026-14.f1.gi:H]") == "ge-2026-14.f1.gi:H"
    assert reply_poller.ref_from_subject("no tag here") == ""


def test_polling_is_off_in_demo_even_when_enabled(monkeypatch):
    # The autouse fixture forces DEMO_MODE=true: even an explicit enable must not poll (DEMO is
    # fully offline), and the default (no env) is off everywhere.
    monkeypatch.setenv("GMAIL_POLLING_ENABLED", "true")
    assert reply_poller.polling_enabled() is False
    monkeypatch.delenv("GMAIL_POLLING_ENABLED", raising=False)
    monkeypatch.setenv("DEMO_MODE", "false")
    assert reply_poller.polling_enabled() is False                # off by default outside DEMO too
    monkeypatch.setenv("GMAIL_POLLING_ENABLED", "true")
    assert reply_poller.polling_enabled() is True


# -- the route and the poller run the SAME processing path (one shared function) --------------
def _priced_return_xlsx(tmp_path) -> bytes:
    """Our own dispatched SoR sheet returned with rates filled — parses with NO model call."""
    from pipeline.stage_03_dispatch.attachments import generate_sor_sheet
    from schemas.models import SorItem, TradeWorkPackage

    pkg = TradeWorkPackage(
        trade="external_works", scope_summary="Landscape",
        sor_items=[SorItem(item_ref="R-01", description="Soft landscape", unit="m2", qty=200.0)],
        source_refs=["Schedule of Rates"],
    )
    path = generate_sor_sheet(pkg, "GE/2026/14", tmp_path / "reply.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "R-01":
            ws.cell(row=r, column=5, value=350)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_route_and_poller_agree_because_they_share_one_processing_function(tmp_path, monkeypatch):
    # The regression the refactor must hold: the HTTP route and the poller call the SAME
    # process_inbound_reply — the stored reply records are identical either way.
    import api

    monkeypatch.setenv("DEMO_MODE", "false")
    xlsx = _priced_return_xlsx(tmp_path)
    ref = "ge-2026-14.TGD-LAND.external_works"

    # (a) the HTTP route, in its own workspace
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "route-ws"))
    ws_a = Workspace()
    reply_loop.record_dispatch(ws_a, ref, "GE/2026/14", "TGD-LAND", "external_works")
    client = TestClient(api.app)
    resp = client.post("/inbound-reply", files={"files": ("return.xlsx", xlsx,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                       data={"ref": ref})
    assert resp.status_code == 200 and resp.json()["status"] == "matched"
    via_route = [r.model_dump() for r in reply_loop.tender_replies(ws_a, "GE/2026/14")]

    # (b) the poller, in a fresh workspace, same bytes arriving as a Gmail attachment
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "poller-ws"))
    ws_b = Workspace()
    reply_loop.record_dispatch(ws_b, ref, "GE/2026/14", "TGD-LAND", "external_works")
    svc = _service(_message("m1", ref, data=xlsx))
    summary = reply_poller.poll_once(api._poller_process_reply, workspace=ws_b, service=svc)
    assert summary["processed"] == 1
    via_poller = [r.model_dump() for r in reply_loop.tender_replies(ws_b, "GE/2026/14")]

    assert via_route == via_poller                                # identical stored replies
    assert via_route and via_route[0]["line_items"]               # and a real priced return landed
