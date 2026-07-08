"""Live-run contracts (Prompt 1): a live run never levels a scenario fixture, and a manual
priced return attaches to its package.

DEMO_MODE is forced true by the autouse fixture (offline). The live-path assertions opt out
per test with monkeypatch, and stay offline by uploading an xlsx (deterministic parse — no
model, no network).
"""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api import app
from pipeline.stage_03_dispatch.attachments import generate_sor_sheet
from schemas.models import SorItem, TradeWorkPackage

client = TestClient(app)

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_level_all_in_demo_falls_back_to_the_fixture():
    # Demo convenience is preserved: an empty replies set levels the baked scenario.
    body = client.post("/level-all", json={"replies": []}).json()
    assert body["sections"] and any(s["levelled"] for s in body["sections"])


def test_level_all_live_empty_never_returns_demo_bids(monkeypatch):
    # The live-run leak: with zero replies, /level-all must NOT fabricate demo bids.
    monkeypatch.setenv("DEMO_MODE", "false")
    assert client.post("/level-all", json={"replies": []}).json()["sections"] == []
    # the singular /level is gated the same way
    assert client.post("/level", json={"replies": []}).json() == []


def _reply_xlsx(tmp_path, *, items, rates) -> bytes:
    """A dispatched SoR sheet returned with rates filled — exactly what a subcontractor sends."""
    pkg = TradeWorkPackage(
        trade="external_works",
        scope_summary="Landscape works",
        sor_items=[SorItem(item_ref=r, description=d, unit=u, qty=q) for (r, d, u, q) in items],
        source_refs=["Schedule of Rates"],
    )
    path = generate_sor_sheet(pkg, "GE/2026/14", tmp_path / "reply.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    for r in range(header_row + 1, ws.max_row + 1):
        ref = ws.cell(row=r, column=1).value
        if ref in rates:
            ws.cell(row=r, column=5, value=rates[ref])  # "Rate (HKD)"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_manual_priced_return_attaches_to_its_package(tmp_path, monkeypatch):
    # A live manual upload levels ONE firm's return, attached to the right package (trade +
    # firm). The frontend merges exactly this into that package's section — no other section.
    monkeypatch.setenv("DEMO_MODE", "false")
    xlsx = _reply_xlsx(
        tmp_path,
        items=[("R-01", "Soft landscape", "m2", 200.0), ("R-02", "Hard paving", "m2", 120.0)],
        rates={"R-01": 350, "R-02": 900},
    )
    resp = client.post(
        "/level-upload",
        files={"files": ("return.xlsx", xlsx, XLSX_CT)},
        data={"firm_id": "TGD-LAND", "trade": "external_works"},
    )
    assert resp.status_code == 200
    body = resp.json()
    levelled = body["levelled"]
    assert len(levelled) == 1  # one firm's return, not a demo comparison
    assert levelled[0]["firm_id"] == "TGD-LAND" and levelled[0]["trade"] == "external_works"
    assert levelled[0]["corrected_total"] == 200.0 * 350 + 120.0 * 900  # Layer-1 qty x rate
    assert body["misdirected"] is None  # no tender/scope supplied -> no misdirect guard run
