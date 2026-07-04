"""Stage 04 leveling — corrects the messy reply's arithmetic, surfaces the missing
provisional sum and the exclusion, flips the clean-firm ranking, and exports Excel."""

import re
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from db import seed, store
from pipeline.stage_01_ingest.ingest import MAX_CHUNK_CHARS
from pipeline.stage_04_level.export_xlsx import OUT_PATH, export_leveling_xlsx
from pipeline.stage_04_level.level import (
    IMAGE_PAGES_PER_CHUNK,
    _chunk_pages,
    level_bids,
    load_demo_replies,
    merge_replies,
    parse_bid_reply,
)
from rules_engine.leveling import computable_amount, level_reply, peer_item_reference
from schemas.models import BidLineItem, BidReply, Severity

_REPLIES_FIXTURE = "cases/messy/bid_replies.json"


@pytest.fixture(scope="module")
def conn(tmp_path_factory):
    db_path = tmp_path_factory.mktemp("level") / "test.db"
    seed.build_database(db_path)
    connection = store.get_connection(db_path)
    yield connection
    connection.close()


@pytest.fixture
def replies():
    return load_demo_replies(_REPLIES_FIXTURE)


@pytest.fixture
def levelled(replies, conn):
    return level_bids(replies, conn=conn)


def _by_firm(levelled):
    return {b.firm_id: b for b in levelled}


def test_corrected_total_differs_from_claimed_on_the_messy_reply(replies, levelled):
    messy = _by_firm(levelled)["F-EL-03"]
    claimed = next(r.claimed_total for r in replies if r.firm_id == "F-EL-03")
    # the understated line is corrected upward, so the "cheap" bid is not cheap
    assert messy.corrected_total != claimed
    assert messy.corrected_total == 12272000.0
    assert claimed == 10272000.0


def test_arithmetic_error_is_caught_with_corrected_value(levelled):
    messy = _by_firm(levelled)["F-EL-03"]
    findings = messy.arithmetic_findings
    assert findings and any(f.location == "line E-03" for f in findings)
    e03 = next(f for f in findings if f.location == "line E-03")
    assert e03.corrected_value == 7740000.0
    assert e03.severity is Severity.WARNING


def test_missing_provisional_sum_is_a_scope_gap_not_zero(levelled):
    messy = _by_firm(levelled)["F-EL-03"]
    assert any("E-06" in gap and "provisional" in gap.lower() for gap in messy.scope_gaps)
    # the gap is surfaced in normalized_total (peer-priced), never silently filled
    assert messy.normalized_total > messy.corrected_total


def test_stated_exclusion_is_flagged_not_deducted(levelled):
    messy = _by_firm(levelled)["F-EL-03"]
    assert any("BWIC" in e or "Builder's work" in e for e in messy.exclusions)


def test_clean_arithmetic_bids_have_no_findings(levelled):
    for firm_id in ("F-EL-01", "F-EL-02", "F-EL-04"):
        assert _by_firm(levelled)[firm_id].arithmetic_findings == []


def test_leveling_changes_the_clean_firm_ranking(replies, levelled):
    clean = ["F-EL-02", "F-EL-03", "F-EL-04"]  # exclude the risk-flagged gotcha
    claimed = {r.firm_id: r.claimed_total for r in replies}
    corrected = {b.firm_id: b.corrected_total for b in levelled}
    cheapest_by_claimed = min(clean, key=lambda f: claimed[f])
    cheapest_by_corrected = min(clean, key=lambda f: corrected[f])
    assert cheapest_by_claimed == "F-EL-03"      # looks cheapest on paper
    assert cheapest_by_corrected == "F-EL-02"    # leveling reveals the real cheapest
    assert cheapest_by_claimed != cheapest_by_corrected


def test_excel_is_produced(replies, levelled):
    out = export_leveling_xlsx(levelled, replies, item_order=["E-01", "E-02", "E-03", "E-04", "E-05", "E-06"], path=OUT_PATH)
    assert out.is_file()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.title == "Leveling"
    # the totals row carries the corrected totals
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert 12272000.0 in flat and 12033000.0 in flat


# -- reply parsing is chunked, merged and deduped (no truncation on a large SoR) --------
#
# A returned SoR with many priced rows overruns max_tokens if parsed in one call — the JSON
# truncates mid-string and the whole reply fails. parse_bid_reply guards against this exactly
# as ingest does: chunk the document, parse each chunk into BidLineItems, merge into one
# BidReply deduped by item_ref, with firm_id/trade authoritative from the resolved ref. These
# tests inject a fake LLMClient (no network, no DB) that returns per-call replies.


def _item(ref: str) -> BidLineItem:
    return BidLineItem(item_ref=ref, description="row", unit="m", qty=1.0, rate=1.0, amount=1.0)


class _FakeClient:
    """Stand-in for LLMClient: records every complete_json call and delegates the return
    to ``responder(user, images, call_index)`` so a test controls what each chunk 'reads'."""

    def __init__(self, responder):
        self.calls: list[SimpleNamespace] = []
        self._responder = responder

    def complete_json(self, *, system, user, target_model, demo_fixture=None, images=None, **_):
        self.calls.append(SimpleNamespace(system=system, user=user, images=images, demo_fixture=demo_fixture))
        return self._responder(user, images, len(self.calls) - 1)


def _big_sor_text(n_items: int) -> tuple[str, list[str]]:
    """A multi-page priced SoR text with ``n_items`` uniquely-referenced rows — large enough
    that _chunk_text must split it into more than one chunk."""
    refs = [f"E-{i:03d}" for i in range(1, n_items + 1)]
    lines: list[str] = []
    for i, ref in enumerate(refs):
        if i % 5 == 0:
            lines.append(f"[page {i // 5 + 1}]")
        lines.append(
            f"{ref} | Rotary drilling and installation of steel casing through soil and rock "
            f"strata including all necessary plant, labour, supervision and disposal | m | {i + 1} | 1000 | {1000 * (i + 1)}"
        )
    return "\n".join(lines), refs


def _echo_refs(user, images, idx):
    """Return exactly the line items whose refs appear in this call's payload, plus a
    deliberately wrong identity — so a dropped chunk would drop its items and the merge
    must override firm_id/trade."""
    refs = re.findall(r"E-\d{3}", user)
    return BidReply(
        firm_id="MODEL-GUESSED-WRONG", trade="model_wrong_trade",
        line_items=[_item(r) for r in refs], exclusions=[], claimed_total=None,
    )


def test_a_many_item_reply_parses_across_bounded_calls_without_truncation():
    text, refs = _big_sor_text(120)
    assert len(text) > MAX_CHUNK_CHARS  # precondition: too big to parse in one call

    client = _FakeClient(_echo_refs)
    reply = parse_bid_reply(firm_id="drill-tech-hk", trade="field_installation", doc_text=text, client=client)

    assert len(client.calls) > 1                                  # chunked, not one oversized call
    assert [li.item_ref for li in reply.line_items] == refs       # every row survived the merge, in order
    assert reply.firm_id == "drill-tech-hk" and reply.trade == "field_installation"  # ref identity wins


def test_a_scanned_reply_is_chunked_into_small_vision_calls():
    images = [f"page-{i}" for i in range(16)]
    client = _FakeClient(lambda user, imgs, idx: BidReply(
        firm_id="?", trade="?", line_items=[_item(f"IMG-{idx:03d}")], exclusions=[], claimed_total=None,
    ))

    reply = parse_bid_reply(firm_id="drill-tech-hk", trade="field_installation", images=images, client=client)

    assert len(client.calls) == 6                                 # ceil(16 / IMAGE_PAGES_PER_CHUNK=3)
    assert all(len(c.images) <= IMAGE_PAGES_PER_CHUNK for c in client.calls)
    assert sum(len(c.images) for c in client.calls) == 16         # every page sent exactly once
    assert len(reply.line_items) == 6                             # one item per call, all merged


def test_a_small_reply_still_parses_in_a_single_call():
    client = _FakeClient(lambda user, imgs, idx: BidReply(
        firm_id="?", trade="?", line_items=[_item("E-001"), _item("E-002")], exclusions=["BWIC"], claimed_total=999.0,
    ))

    reply = parse_bid_reply(firm_id="f1", trade="field_installation", doc_text="one short priced page", client=client)

    assert len(client.calls) == 1                                 # small text -> a single chunk
    assert [li.item_ref for li in reply.line_items] == ["E-001", "E-002"]
    assert reply.exclusions == ["BWIC"] and reply.claimed_total == 999.0


def test_no_text_and_no_images_still_makes_one_call():
    # The DEMO / fixture path: parse_bid_reply is called with images=[] and doc_text="".
    client = _FakeClient(lambda user, imgs, idx: BidReply(firm_id="?", trade="?", line_items=[_item("E-001")]))

    reply = parse_bid_reply(firm_id="f1", trade="field_installation", client=client)

    assert len(client.calls) == 1 and client.calls[0].images is None
    assert [li.item_ref for li in reply.line_items] == ["E-001"]


def test_merge_replies_dedupes_by_item_ref_unions_exclusions_and_keeps_identity():
    r1 = BidReply(firm_id="x", trade="t", line_items=[_item("A-1"), _item("A-2")], exclusions=["e1"], claimed_total=None)
    r2 = BidReply(firm_id="y", trade="u", line_items=[_item("A-2"), _item("A-3")], exclusions=["e1", "e2"], claimed_total=500.0)

    merged = merge_replies([r1, r2], firm_id="firm-authoritative", trade="field_installation")

    assert [li.item_ref for li in merged.line_items] == ["A-1", "A-2", "A-3"]  # A-2 deduped, first wins
    assert merged.exclusions == ["e1", "e2"]                                    # unioned, order preserved
    assert merged.claimed_total == 500.0                                       # first stated total
    assert merged.firm_id == "firm-authoritative" and merged.trade == "field_installation"


def test_chunk_pages_groups_without_splitting_a_page():
    pages = [f"p{i}" for i in range(7)]
    groups = _chunk_pages(pages, 3)

    assert [len(g) for g in groups] == [3, 3, 1]
    assert [p for g in groups for p in g] == pages  # order and completeness preserved
    assert _chunk_pages([], 3) == []


# -- rate-aware leveling: rate-first comparison, amounts only where quantities exist -----
#
# A real Schedule of Rates (GE/2026/14) is rate-only — most lines have a unit rate and no
# quantity — so leveling compares rates as the foundation and computes amounts/totals only
# where a quantity is present, never evaluating qty*rate on a None. level_reply is pure
# (no DB, no network); these drive it directly with constructed replies.


def _rate_only(item_ref: str, rate: float, unit: str = "m") -> BidLineItem:
    return BidLineItem(item_ref=item_ref, description=f"{item_ref} works", unit=unit, rate=rate)  # qty/amount absent


def _boq(item_ref: str, qty: float, rate: float, amount: float) -> BidLineItem:
    return BidLineItem(item_ref=item_ref, description=f"{item_ref} works", unit="no", qty=qty, rate=rate, amount=amount)


def test_rate_only_reply_levels_by_rate_with_no_amount_total_and_no_crash():
    # (a) a rate-only reply (no qty anywhere) levels by rate, crashes nowhere, no bogus total
    reply = BidReply(firm_id="F-RO", trade="ground_investigation", line_items=[
        _rate_only("A-1", 1200.0), _rate_only("A-2", 450.0, unit="hr"), _rate_only("A-3", 800.0),
    ])
    lev = level_reply(reply, "GI Firm", peer_item_reference([reply]))

    assert lev.corrected_total == 0.0 and lev.normalized_total == 0.0  # no quantities -> no amount
    assert lev.arithmetic_findings == []                              # rate-only is not an error
    assert lev.scope_gaps == []                                       # a rate is a price, not a gap
    assert {ir.item_ref: ir.rate for ir in lev.item_rates} == {"A-1": 1200.0, "A-2": 450.0, "A-3": 800.0}
    assert all(ir.amount is None for ir in lev.item_rates)            # amounts stay uncomputed


def test_full_boq_reply_totals_by_amount_exactly():
    # (b) a full BoQ reply totals exactly (qty*rate per line), as before
    reply = BidReply(firm_id="F-BQ", trade="electrical", line_items=[
        _boq("E-01", 2.0, 1000.0, 2000.0), _boq("E-02", 100.0, 50.0, 5000.0),
    ])
    lev = level_reply(reply, "BoQ Firm", peer_item_reference([reply]))

    assert lev.corrected_total == 7000.0
    assert lev.arithmetic_findings == []
    assert [ir.amount for ir in lev.item_rates] == [2000.0, 5000.0]   # amount computed per line


def test_mixed_reply_totals_boq_lines_and_compares_rate_only_lines():
    # (c) a mixed reply handles both per line
    reply = BidReply(firm_id="F-MX", trade="ground_investigation", line_items=[
        _boq("M-1", 10.0, 300.0, 3000.0),                                        # BoQ line
        _rate_only("M-2", 1500.0),                                               # rate-only line
        BidLineItem(item_ref="M-3", description="Provisional sum — testing", unit="item"),  # unpriced -> gap
    ])
    lev = level_reply(reply, "Mixed Firm", peer_item_reference([reply]))

    assert lev.corrected_total == 3000.0                             # only the BoQ line contributes
    assert lev.arithmetic_findings == []
    assert any(g.startswith("M-3") and "provisional" in g.lower() for g in lev.scope_gaps)
    rates = {ir.item_ref: (ir.rate, ir.amount) for ir in lev.item_rates}
    assert rates["M-1"] == (300.0, 3000.0)                           # BoQ: rate and amount
    assert rates["M-2"] == (1500.0, None)                            # rate-only: rate, no amount
    assert rates["M-3"] == (None, None)                              # unpriced


def test_rate_present_but_qty_absent_is_not_flagged_as_an_arithmetic_error():
    # (d) a rate-present / qty-absent line is not an arithmetic error, even with a stray amount,
    #     and does not inject that amount into the total (compared by rate only).
    reply = BidReply(firm_id="F-Q", trade="ground_investigation", line_items=[
        BidLineItem(item_ref="Q-1", description="Rotary drilling", unit="m", rate=1000.0, amount=999999.0),  # no qty
    ])
    lev = level_reply(reply, "Q Firm", peer_item_reference([reply]))

    assert lev.arithmetic_findings == []      # no qty -> no qty*rate to check the stated amount against
    assert lev.corrected_total == 0.0         # excluded from the amount total (point 4)
    assert lev.scope_gaps == []               # it has a rate, so it is not a gap
    assert lev.item_rates[0].rate == 1000.0 and lev.item_rates[0].amount is None


def test_computable_amount_covers_boq_lumpsum_rateonly_and_unpriced():
    assert computable_amount(_boq("x", 3.0, 10.0, 999.0)) == 30.0            # qty*rate wins over stated amount
    assert computable_amount(BidLineItem(item_ref="x", amount=500.0)) == 500.0  # lump sum, no rate basis -> stated
    assert computable_amount(_rate_only("x", 42.0)) is None                  # rate-only -> no amount
    assert computable_amount(BidLineItem(item_ref="x")) is None              # wholly unpriced -> no amount


def test_demo_hero_corrected_totals_are_unchanged(replies, levelled):
    # The rate-aware refactor must not move the baked demo/hero figures.
    corrected = {b.firm_id: b.corrected_total for b in levelled}
    assert corrected == {"F-EL-01": 9877000.0, "F-EL-02": 12033000.0, "F-EL-03": 12272000.0, "F-EL-04": 13442000.0}
    f03 = _by_firm(levelled)["F-EL-03"]
    assert f03.normalized_total == 12572000.0  # E-06 gap valued at the peer median (300000)


def test_export_of_a_rate_only_reply_shows_rate_and_blank_amount(conn, tmp_path):
    # Guards the export arithmetic site: a rate-only line must not crash on qty*rate.
    reply = BidReply(firm_id="F-RO", trade="ground_investigation", line_items=[
        _rate_only("A-1", 1200.0), _rate_only("A-2", 450.0, unit="hr"),
    ])
    levelled = level_bids([reply], conn=conn)
    out = export_leveling_xlsx(levelled, [reply], path=tmp_path / "rate_only.xlsx")

    flat = [c.value for row in load_workbook(out).active.iter_rows() for c in row]
    assert 1200.0 in flat and 450.0 in flat   # rates are shown
    assert "—" in flat                         # the amount cells are blank for rate-only lines
