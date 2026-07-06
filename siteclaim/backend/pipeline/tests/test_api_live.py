"""Phase A live-engine API surface, exercised in DEMO_MODE (offline).

The routes exist and behave, but DEMO_MODE keeps every one of them off the network:
``/level-upload`` returns the baked levelling, ``/dispatch`` records to the mock
outbox, and ``/contacts`` reads the seeded address book.
"""

import time

from fastapi.testclient import TestClient

from api import app

client = TestClient(app)


def _ingest_and_wait(*, files, data=None, timeout_s: float = 10.0) -> dict:
    """POST /ingest-upload, then for a live job poll /ingest-status until terminal. Returns the
    terminal IngestJobState dict. DEMO returns ``done`` inline (no job), so no polling happens."""
    start = client.post("/ingest-upload", files=files, data=data or {})
    assert start.status_code == 200, start.text
    body = start.json()
    if body["status"] in ("done", "error"):
        return body
    job_id = body["job_id"]
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        state = client.get(f"/ingest-status/{job_id}").json()
        if state["status"] in ("done", "error"):
            return state
        time.sleep(0.02)
    raise AssertionError(f"ingest job {job_id} did not finish within {timeout_s}s")


def test_live_routes_are_registered():
    paths = {route.path for route in app.routes}
    assert {"/level-upload", "/contacts"} <= paths


def test_heavy_upload_routes_are_sync_so_they_run_in_the_threadpool():
    # These handlers do blocking work (pymupdf render, sequential LLM calls). They must be
    # plain sync def — FastAPI then runs them in its threadpool, so one long ingest never
    # starves the event loop (/health kept answering in the live run). async def would put
    # the blocking work back on the loop; guard against a regression to it.
    import inspect

    import api

    for handler in (api.post_ingest_upload, api.post_level_upload, api.post_inbound_reply):
        assert not inspect.iscoroutinefunction(handler), f"{handler.__name__} must be sync def"


def test_contacts_endpoint_lists_the_address_book():
    contacts = client.get("/contacts").json()
    assert isinstance(contacts, list) and contacts
    assert all({"firm_id", "trade", "email"} <= set(c) for c in contacts)


def test_level_upload_in_demo_returns_the_baked_levelling():
    files = {"files": ("reply.pdf", b"%PDF-1.4 fake", "application/pdf")}
    data = {"firm_id": "F-EL-03", "trade": "electrical"}
    resp = client.post("/level-upload", files=files, data=data)
    assert resp.status_code == 200
    levelled = resp.json()
    assert levelled and all("corrected_total" in bid for bid in levelled)


def test_ingest_upload_in_demo_returns_scope_and_untagged_tender():
    files = {"files": ("tender.pdf", b"%PDF-1.4 fake", "application/pdf")}
    body = _ingest_and_wait(files=files, data={"project_name": "Live Tender A"})
    assert body["status"] == "done" and body["job_id"] is None  # DEMO is inline, no job created
    result = body["result"]
    assert result["scope"]["packages"]     # the scope split
    assert result["tender"]["documents"]   # the tagged tender (for /dispatch routing)
    # DEMO_MODE leaves the tender untagged — classification runs only on the live path,
    # so the demo scenarios and the hero catch are untouched.
    assert all(doc["trades"] == [] for doc in result["tender"]["documents"])


def test_dispatch_send_in_demo_is_mock_and_carries_attachments():
    case = client.get("/demo/messy").json()
    scope = client.post("/ingest", json={"tender": case["tender"]}).json()
    shortlist = client.post("/shortlist", json={"scope": scope}).json()
    dispatch = client.post("/dispatch", json={
        "shortlist": shortlist, "approvals": {"electrical": ["F-EL-02"]},
        "scope": scope, "tender": case["tender"], "project_name": case["name"], "send": True,
    }).json()
    bundle = dispatch["bundles"][0]
    assert bundle["status"] == "sent_mock"  # DEMO_MODE never really sends
    kinds = {a["kind"] for a in bundle["attachments"]}
    assert "sor_sheet" in kinds and "general" in kinds  # routed, even if not written to disk


def test_refresh_write_routes_are_disabled_in_demo():
    # The refresh write path must never mutate the committed demo DB during a demo run.
    stage = client.post("/refresh/stage", json={"records": [
        {"firm_id": "some-firm", "public_flags": [{"signal_type": "winding_up", "label": "x"}]}
    ]})
    confirm = client.post("/refresh/confirm", json={})
    assert stage.status_code == 409 and confirm.status_code == 409
    # the read-only pending view is available and tolerant (empty on the demo DB)
    pending = client.get("/refresh/pending")
    assert pending.status_code == 200 and isinstance(pending.json(), list)


def test_refresh_routes_are_registered():
    paths = {route.path for route in app.routes}
    assert {"/refresh/stage", "/refresh/pending", "/refresh/confirm", "/refresh/reject"} <= paths


def test_refresh_write_refuses_a_non_live_target(tmp_path, monkeypatch):
    # Even with DEMO_MODE off, a refresh must never mutate a demo-profile DB.
    from db import seed

    demo_db = tmp_path / "demo.db"
    seed.build_database(demo_db)  # profile 'demo'
    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setenv("SITESOURCE_DB", str(demo_db))
    resp = client.post("/refresh/stage", json={"records": []})
    assert resp.status_code == 409


def test_refresh_write_applies_to_a_live_target(tmp_path, monkeypatch):
    from db import seed

    live_db = tmp_path / "live.db"
    seed.build_database(live_db, profile="live")
    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setenv("SITESOURCE_DB", str(live_db))
    resp = client.post("/refresh/stage", json={"records": [
        {"firm_id": "new-live-firm-1", "name_en": "New Live Firm Ltd", "trades": ["electrical"],
         "public_flags": [{"signal_type": "winding_up", "label": "Winding-up 2026"}]}
    ]})
    assert resp.status_code == 200 and resp.json()["staged_firms"] == 1


def test_inbound_reply_route_is_registered():
    assert "/inbound-reply" in {route.path for route in app.routes}


def test_inbound_reply_ref_path_accumulates_and_relevels(monkeypatch, tmp_path):
    from pipeline import reply_loop
    from pipeline.workspace import Workspace

    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))
    ws = Workspace()  # picks up the env
    ref2 = reply_loop.make_ref("Kwun Tong", "F-EL-02", "electrical")
    reply_loop.record_dispatch(ws, ref2, "Kwun Tong", "F-EL-02", "electrical")

    first = client.post("/inbound-reply", files={"files": ("reply.pdf", b"%PDF-1.4", "application/pdf")}, data={"ref": ref2})
    assert first.status_code == 200
    body = first.json()
    assert body["status"] == "matched" and body["reply_count"] == 1
    assert [b["firm_id"] for b in body["comparison"]] == ["F-EL-02"]

    # a second firm's reply on the same tender grows the comparison (accumulate + relevel)
    ref3 = reply_loop.make_ref("Kwun Tong", "F-EL-03", "electrical")
    reply_loop.record_dispatch(ws, ref3, "Kwun Tong", "F-EL-03", "electrical")
    second = client.post("/inbound-reply", files={"files": ("reply.pdf", b"%PDF-1.4", "application/pdf")}, data={"ref": ref3})
    body2 = second.json()
    assert body2["reply_count"] == 2
    assert {b["firm_id"] for b in body2["comparison"]} == {"F-EL-02", "F-EL-03"}


def test_inbound_reply_without_a_ref_is_unmatched(monkeypatch, tmp_path):
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))  # empty registry -> nothing to match
    resp = client.post("/inbound-reply", files={"files": ("reply.pdf", b"%PDF-1.4", "application/pdf")}, data={"ref": ""})
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "unmatched" and "manual assignment" in body["detail"]
    assert body["comparison"] == []


def test_shortlist_include_public_opens_the_pool():
    # Phase B, reached through the API: the live-engine flag adds real public firms.
    case = client.get("/demo/messy").json()
    scope = client.post("/ingest", json={"tender": case["tender"]}).json()
    default = client.post("/shortlist", json={"scope": scope}).json()
    public = client.post("/shortlist", json={"scope": scope, "include_public": True}).json()
    assert len(public["per_trade"]["electrical"]) > len(default["per_trade"]["electrical"])


def test_tender_replies_endpoint_reports_accumulated_and_outstanding(monkeypatch, tmp_path):
    # Two firms dispatched; one replies. The endpoint shows the reply, the item count, the
    # outstanding firm, and that the comparison xlsx is ready — no need to open Excel.
    from pipeline import reply_loop
    from pipeline.workspace import Workspace, tender_slug

    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))
    ws = Workspace()
    project = "Kwun Tong"
    slug = tender_slug(project)  # "kwun-tong"
    for fid in ("F-EL-02", "F-EL-03"):
        reply_loop.record_dispatch(ws, reply_loop.make_ref(project, fid, "electrical"), project, fid, "electrical")

    # Before any reply: zero replies, both outstanding, no comparison.
    empty = client.get(f"/tender/{slug}/replies").json()
    assert empty["reply_count"] == 0 and empty["last_received"] is None
    assert {o["firm_id"] for o in empty["outstanding"]} == {"F-EL-02", "F-EL-03"}
    assert empty["comparison_available"] is False
    assert client.get(f"/tender/{slug}/comparison.xlsx").status_code == 404

    # F-EL-02 replies (DEMO fixture parse + comparison written).
    ref = reply_loop.make_ref(project, "F-EL-02", "electrical")
    assert client.post("/inbound-reply", files={"files": ("r.pdf", b"%PDF-1.4", "application/pdf")}, data={"ref": ref}).status_code == 200

    body = client.get(f"/tender/{slug}/replies").json()
    assert body["tender_slug"] == slug and body["reply_count"] == 1
    assert body["last_received"] is not None
    assert [r["firm_id"] for r in body["replies"]] == ["F-EL-02"]
    assert body["replies"][0]["line_items"] >= 1
    assert {o["firm_id"] for o in body["outstanding"]} == {"F-EL-03"}  # F-EL-02 no longer outstanding
    assert body["comparison_available"] is True

    dl = client.get(f"/tender/{slug}/comparison.xlsx")
    assert dl.status_code == 200 and dl.content[:2] == b"PK"  # xlsx zip


def test_tender_replies_routes_are_registered():
    paths = {route.path for route in app.routes}
    assert {"/tender/{slug}/replies", "/tender/{slug}/comparison.xlsx"} <= paths


def test_ingest_upload_returns_the_tender_slug():
    # The client needs the server-derived slug to poll /tender/{slug}/replies.
    body = _ingest_and_wait(files={"files": ("t.pdf", b"%PDF-1.4 fake", "application/pdf")})
    assert body["result"]["tender_slug"] == "uploaded-tender"  # slug of the DEMO placeholder name


def test_shortlist_k_caps_the_ranked_list_through_the_api():
    # The live frontend sends include_public + k so a broad trade (22 external_works
    # firms on the real GI tender) does not become 22 dispatch bundles.
    case = client.get("/demo/messy").json()
    scope = client.post("/ingest", json={"tender": case["tender"]}).json()
    full = client.post("/shortlist", json={"scope": scope, "include_public": True}).json()
    capped = client.post("/shortlist", json={"scope": scope, "include_public": True, "k": 3}).json()
    for trade, cands in full["per_trade"].items():
        expected = [c["firm"]["firm_id"] for c in cands][:3]
        assert [c["firm"]["firm_id"] for c in capped["per_trade"][trade]] == expected  # head, not reshuffle


# -- deterministic xlsx reply parsing through the routes ------------------------------
#
# The realistic reply is our own dispatched SoR sheet returned with the Rate column
# filled — parsing it needs no model at all. These tests flip DEMO off (the parse
# branch only runs live) and stub the LLM parse with an assertion bomb, proving the
# xlsx path never consults the model: openpyxl is local, so everything stays offline.

_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _model_must_not_be_called(*args, **kwargs):
    raise AssertionError("the model must not be called on the xlsx reply path")


def _filled_sheet_bytes(tmp_path) -> bytes:
    """The dispatched SoR sheet for two electrical items, rates filled in by the firm."""
    from io import BytesIO

    from openpyxl import load_workbook

    from pipeline.stage_03_dispatch.attachments import generate_sor_sheet
    from schemas.models import SorItem, TradeWorkPackage

    pkg = TradeWorkPackage(
        trade="electrical", scope_summary="LV works",
        sor_items=[
            SorItem(item_ref="E-01", description="LV main switchboard", unit="no", qty=1.0),
            SorItem(item_ref="E-02", description="Sub-main cabling", unit="m", qty=100.0),
        ],
        source_refs=["Schedule of Rates"],
    )
    path = generate_sor_sheet(pkg, "Kwun Tong", tmp_path / "sor.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    header = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    rates = {"E-01": 950000, "E-02": 2500}
    for r in range(header + 1, ws.max_row + 1):
        ref = ws.cell(row=r, column=1).value
        if ref in rates:
            ws.cell(row=r, column=5, value=rates[ref])  # "Rate (HKD)"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_level_upload_parses_an_xlsx_reply_with_no_model_call(monkeypatch, tmp_path):
    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setattr("api.parse_bid_reply", _model_must_not_be_called)

    resp = client.post(
        "/level-upload",
        files={"files": ("SoR_electrical.xlsx", _filled_sheet_bytes(tmp_path), _XLSX_CT)},
        data={"firm_id": "F-EL-02", "trade": "electrical"},
    )

    assert resp.status_code == 200
    (bid,) = resp.json()
    assert bid["firm_id"] == "F-EL-02" and bid["trade"] == "electrical"
    assert bid["corrected_total"] == 1200000.0  # 1 x 950,000 + 100 x 2,500 — Layer 1 arithmetic
    assert {ir["item_ref"]: ir["rate"] for ir in bid["item_rates"]} == {"E-01": 950000.0, "E-02": 2500.0}


def test_inbound_reply_xlsx_resolves_by_ref_and_parses_deterministically(monkeypatch, tmp_path):
    from pipeline import reply_loop
    from pipeline.workspace import Workspace

    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))
    monkeypatch.setattr("api.parse_bid_reply", _model_must_not_be_called)
    ws = Workspace()
    ref = reply_loop.make_ref("Kwun Tong", "F-EL-02", "electrical")
    reply_loop.record_dispatch(ws, ref, "Kwun Tong", "F-EL-02", "electrical")

    resp = client.post(
        "/inbound-reply",
        files={"files": ("SoR_electrical.xlsx", _filled_sheet_bytes(tmp_path), _XLSX_CT)},
        data={"ref": ref},
    )

    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "matched" and body["firm_id"] == "F-EL-02"  # ref stays authoritative
    assert body["reply_count"] == 1
    assert body["comparison"][0]["corrected_total"] == 1200000.0


def test_level_upload_rejects_a_non_sor_xlsx_with_a_clear_400(monkeypatch):
    from io import BytesIO

    from openpyxl import Workbook

    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setattr("api.parse_bid_reply", _model_must_not_be_called)
    wb = Workbook()
    wb.active.append(["Colour", "Size", "Price"])  # wrong headers — not our SoR layout
    buffer = BytesIO()
    wb.save(buffer)

    resp = client.post(
        "/level-upload",
        files={"files": ("prices.xlsx", buffer.getvalue(), _XLSX_CT)},
        data={"firm_id": "F-EL-02", "trade": "electrical"},
    )

    assert resp.status_code == 400
    assert "Not a Schedule of Rates sheet" in resp.json()["detail"]


# -- /ingest-upload adopts the extracted contract name (project_name propagation) -----
#
# Live-verified inconsistency: the split extracts the real contract name into
# scope.project_name while the returned tender keeps the form default "Uploaded tender",
# so the workspace, ref registry, and dispatch all key off the placeholder. The route now
# decides ONE final name after the split (explicit form value wins; otherwise the
# extracted name is adopted) and saves the originals under it. The model seams are
# stubbed — offline, no fitz, no LLM.

_GE_NAME = "Contract No. GE/2026/14 — Ground Investigation Works"


def _stub_live_ingest(monkeypatch, extracted_name: str):
    from schemas.models import ScopePackages

    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setattr("api.extract_document", lambda data, ct: ("[page 1]\nSoR rows", []))
    monkeypatch.setattr("api.to_images", lambda data, ct, max_pages=2: ["page-png"])
    monkeypatch.setattr(
        "api.ingest_tender",
        lambda tender, images=None, doc_text="", context_text="", progress_cb=None: ScopePackages(project_name=extracted_name, packages=[]),
    )
    monkeypatch.setattr("api.classify_documents", lambda tender, imgs, per_doc_text=None: tender)


def test_ingest_upload_adopts_the_extracted_contract_name(monkeypatch, tmp_path):
    from pipeline.workspace import Workspace

    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))
    _stub_live_ingest(monkeypatch, _GE_NAME)

    # form project_name left at its default -> the split's extracted name is adopted
    body = _ingest_and_wait(files={"files": ("sr01.pdf", b"%PDF-1.4", "application/pdf")})

    assert body["status"] == "done", body
    result = body["result"]
    assert result["tender"]["project_name"] == _GE_NAME  # adopted into the returned tender
    assert result["scope"]["project_name"] == _GE_NAME
    ws = Workspace()
    assert (ws.docs_dir(_GE_NAME) / "sr01.pdf").is_file()  # originals under the final slug
    assert not ws.tender_dir("Uploaded tender").exists()   # nothing under the placeholder


def test_ingest_upload_keeps_an_explicit_project_name(monkeypatch, tmp_path):
    from pipeline.workspace import Workspace

    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))
    _stub_live_ingest(monkeypatch, _GE_NAME)

    body = _ingest_and_wait(
        files={"files": ("sr01.pdf", b"%PDF-1.4", "application/pdf")},
        data={"project_name": "My Tender 7"},
    )

    result = body["result"]
    assert result["tender"]["project_name"] == "My Tender 7"  # explicit operator value wins
    assert result["scope"]["project_name"] == "My Tender 7"   # aligned — one name end-to-end
    assert (Workspace().docs_dir("My Tender 7") / "sr01.pdf").is_file()


def test_ingest_upload_demo_keeps_the_placeholder_name():
    # DEMO path: baked fixture scope, no adoption, no workspace writes — unchanged.
    body = _ingest_and_wait(files={"files": ("tender.pdf", b"%PDF-1.4 fake", "application/pdf")})
    assert body["result"]["tender"]["project_name"] == "Uploaded tender"


# -- /ingest-upload gates item extraction to the Schedule(s) of Rates ------------------
#
# Live bug: item extraction ran over the merged text of ALL documents, so the Method of
# Measurement yielded 57 phantom sor_items. The route now classifies first and feeds only
# schedule_of_rates text to the extractor; other documents inform the trade split as
# context but never produce a line item. The model seams are stubbed (offline).

def test_ingest_upload_extracts_items_only_from_schedule_of_rates_text(monkeypatch, tmp_path):
    from schemas.models import DocType, ScopePackages, SorItem, TenderPackage, TradeWorkPackage

    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))

    # extract_document returns per-file text keyed off the filename passed via content.
    texts = {"sr01.pdf": "A1 rotary drilling m 300", "mm01.pdf": "Method: measure net in place, 57 rules"}
    monkeypatch.setattr("api.extract_document", lambda data, ct: (texts[data.decode()], []))
    monkeypatch.setattr("api.to_images", lambda data, ct, max_pages=2: [])

    # Classifier: sr01 is the Schedule of Rates, mm01 is the Method of Measurement.
    def fake_classify(tender, per_doc_images, per_doc_text=None):
        kinds = {"sr01.pdf": DocType.SCHEDULE_OF_RATES, "mm01.pdf": DocType.METHOD_OF_MEASUREMENT}
        return TenderPackage(project_name=tender.project_name, documents=[
            d.model_copy(update={"doc_type": kinds[d.filename], "trades": []}) for d in tender.documents
        ])
    monkeypatch.setattr("api.classify_documents", fake_classify)

    captured = {}

    def fake_ingest(tender, images=None, doc_text="", context_text="", progress_cb=None):
        captured["doc_text"] = doc_text
        captured["context_text"] = context_text
        return ScopePackages(project_name="GE/2026/14", packages=[TradeWorkPackage(
            trade="ground_investigation", scope_summary="GI",
            sor_items=[SorItem(item_ref="A1", description="rotary drilling", unit="m", qty=300.0)],
        )])
    monkeypatch.setattr("api.ingest_tender", fake_ingest)

    body = _ingest_and_wait(files=[
        ("files", ("sr01.pdf", b"sr01.pdf", "application/pdf")),
        ("files", ("mm01.pdf", b"mm01.pdf", "application/pdf")),
    ])

    assert body["status"] == "done", body
    # Only the SoR text reaches the priced-item extractor; the MoM is context, not items.
    assert "rotary drilling" in captured["doc_text"] and "Method: measure" not in captured["doc_text"]
    assert "Method: measure" in captured["context_text"]
    assert [i["item_ref"] for i in body["result"]["scope"]["packages"][0]["sor_items"]] == ["A1"]


def test_ingest_upload_mom_only_yields_no_line_items(monkeypatch, tmp_path):
    # An upload with NO Schedule of Rates (just a Method of Measurement) must produce a
    # scope split with zero sor_items — never phantom rows from the MoM's item-like text.
    from schemas.models import DocType, ScopePackages, TenderPackage, TradeWorkPackage

    monkeypatch.setenv("DEMO_MODE", "false")
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path))
    monkeypatch.setattr("api.extract_document", lambda data, ct: ("Method of measurement rules, 57 of them", []))
    monkeypatch.setattr("api.to_images", lambda data, ct, max_pages=2: [])
    monkeypatch.setattr("api.classify_documents", lambda tender, imgs, per_doc_text=None: TenderPackage(
        project_name=tender.project_name,
        documents=[d.model_copy(update={"doc_type": DocType.METHOD_OF_MEASUREMENT, "trades": []}) for d in tender.documents],
    ))

    seen = {}

    def fake_ingest(tender, images=None, doc_text="", context_text="", progress_cb=None):
        seen["doc_text"] = doc_text
        # The real extractor produces no items from empty SoR text; model the split it returns.
        return ScopePackages(project_name="GE/2026/14", packages=[
            TradeWorkPackage(trade="ground_investigation", scope_summary="GI", sor_items=[])
        ])
    monkeypatch.setattr("api.ingest_tender", fake_ingest)

    body = _ingest_and_wait(files={"files": ("mm01.pdf", b"x", "application/pdf")})

    assert body["status"] == "done", body
    assert seen["doc_text"] == ""  # no schedule_of_rates text fed to the extractor
    assert body["result"]["scope"]["packages"][0]["sor_items"] == []
