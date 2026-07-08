"""Every dispatched enquiry gets a sheet, anchored on ITS canonical items (return round-trip v2,
Commit 1). The sheet is enquiry-anchored — the full canonical item set, a firm's rate where its
return matched, a scope gap where not — not reply-anchored. Reads the real workbook with openpyxl."""

from openpyxl import load_workbook

from pipeline.stage_04_level.export_xlsx import export_leveling_xlsx, sheet_title
from pipeline.stage_04_level.level import level_bids
from schemas.models import (
    BidLineItem,
    BidReply,
    ScopePackages,
    SectionMeta,
    SorItem,
    TradeWorkPackage,
)


def _texts(ws):
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def _unit_items():
    return {"ground_investigation:H": [
        SorItem(item_ref="H1", description="Boreholes by rotary drilling", section="H"),
        SorItem(item_ref="H2", description="Undisturbed sampling", section="H"),
    ]}


def test_dispatched_unit_with_no_return_still_gets_a_canonical_sheet(tmp_path):
    # The Geophysical-Survey-vanished bug: a dispatched enquiry whose return is absent still gets its
    # sheet, with every canonical row present and the awaiting note — never dropped from the workbook.
    out = export_leveling_xlsx(
        [], [], path=tmp_path / "c.xlsx",
        units=["ground_investigation:H"], unit_items=_unit_items(),
        awaiting={"ground_investigation:H": ["F-A"]})
    ws = load_workbook(out)[sheet_title("ground_investigation:H")]
    texts = _texts(ws)
    assert "H1" in texts and "H2" in texts                       # every canonical row present
    assert "Boreholes by rotary drilling" in texts and "Undisturbed sampling" in texts
    assert "Awaiting reply (enquired, no priced return yet)" in texts and "F-A" in texts


def test_priced_return_shows_rates_and_unpriced_canonical_rows_as_gaps(tmp_path):
    # A correct priced return: H1 priced renders its rate; H2 unpriced is still a row, marked a scope
    # gap — the sheet is the enquiry's full item set, not just what the firm returned.
    reply = BidReply(firm_id="F-A", trade="ground_investigation:H",
                     line_items=[BidLineItem(item_ref="H1", qty=1.0, rate=100.0, amount=100.0)])
    levelled = level_bids([reply])
    out = export_leveling_xlsx(
        levelled, [reply], path=tmp_path / "c.xlsx",
        units=["ground_investigation:H"], unit_items=_unit_items())
    ws = load_workbook(out)[sheet_title("ground_investigation:H")]
    texts = _texts(ws)
    assert "H1" in texts and "H2" in texts                       # both canonical rows shown
    assert "100" in texts                                         # H1's rate rendered
    assert any("scope gap (unpriced)" in t for t in texts)       # H2 flagged unpriced for F-A


def test_return_matching_the_wrong_reformed_ref_still_anchors_on_canonical(tmp_path):
    # A routed line keeps its original ref form ("h1" / "H 1"); the sheet matches it to canonical H1
    # by normalised ref, so its rate lands on the canonical row (not a phantom extra row).
    reply = BidReply(firm_id="F-A", trade="ground_investigation:H",
                     line_items=[BidLineItem(item_ref="h1", qty=1.0, rate=100.0, amount=100.0)])
    out = export_leveling_xlsx(
        level_bids([reply]), [reply], path=tmp_path / "c.xlsx",
        units=["ground_investigation:H"], unit_items=_unit_items())
    ws = load_workbook(out)[sheet_title("ground_investigation:H")]
    texts = _texts(ws)
    assert "100" in texts and "H1" in texts                       # rate on the canonical row
    assert "h1" not in texts                                      # no phantom original-ref row appended


# -- scope-aware leveling: coverage + like-for-like valuation over the routed unit ----------
def _gi_split_scope() -> ScopePackages:
    # >3 sections -> route_units auto-splits ground_investigation into per-section units (G/H/I/J).
    return ScopePackages(project_name="GE/2026/14", packages=[TradeWorkPackage(
        trade="ground_investigation", scope_summary="GI",
        sor_items=[
            SorItem(item_ref="G4", description="Trial pit excavation", section="G"),
            SorItem(item_ref="H5A", description="Rotary drilling in rock", section="H"),
            SorItem(item_ref="H12", description="Field vane shear test in soft clay", section="H"),
            SorItem(item_ref="H13", description="Standpipe piezometer installation", section="H"),
            SorItem(item_ref="I3", description="Rotary core drilling in rock", section="I"),
            SorItem(item_ref="J1", description="Instrumentation monitoring", section="J")],
        sections=[SectionMeta(code=c, item_count=n) for c, n in [("G", 1), ("H", 3), ("I", 1), ("J", 1)]])])


def test_a_description_routed_line_is_not_double_counted_as_an_unpriced_gap():
    # A garbled ref ("H1Z") that routes to canonical H12 only by DESCRIPTION must count as returning
    # H12 — not be added to the unpriced scope (valued at the peer price) AND left priced, which would
    # double-count H12 in normalized_total and falsely mark it "not returned".
    scope = _gi_split_scope()
    a = BidReply(firm_id="F-A", trade="ground_investigation:H", line_items=[
        BidLineItem(item_ref="H1Z", description="Field vane shear test in soft clay",
                    qty=1.0, rate=100.0, amount=100.0)])
    b = BidReply(firm_id="F-B", trade="ground_investigation:H", line_items=[
        BidLineItem(item_ref="H12", description="Field vane shear test in soft clay",
                    qty=1.0, rate=200.0, amount=200.0)])
    by = {x.firm_id: x for x in level_bids([a, b], scope)}
    fa = by["F-A"]
    assert fa.corrected_total == 100.0
    assert fa.normalized_total == 100.0                       # NOT 300 — H12 not double-counted at peer price
    assert not any("H12" in g for g in fa.scope_gaps)         # H12 was priced (by description), not a gap
    assert any(g.startswith("H13") for g in fa.scope_gaps)    # H13 genuinely not returned -> a real gap


def test_a_returned_but_unpriced_gap_is_valued_at_the_peer_price_despite_ref_drift():
    # F-A returns H5A under the drifted form "H5(a)" with NO rate -> a returned-but-unpriced gap; a
    # peer priced "H5A". The gap must be valued at the peer price (matched on the normalised ref),
    # not 0 — otherwise a firm that left expensive scope unpriced ranks artificially cheap.
    scope = _gi_split_scope()
    a = BidReply(firm_id="F-A", trade="ground_investigation:H", line_items=[
        BidLineItem(item_ref="H5(a)", description="Rotary drilling in rock")])  # no rate/amount -> gap
    b = BidReply(firm_id="F-B", trade="ground_investigation:H", line_items=[
        BidLineItem(item_ref="H5A", description="Rotary drilling in rock", qty=1.0, rate=200.0, amount=200.0)])
    by = {x.firm_id: x for x in level_bids([a, b], scope)}
    fa = by["F-A"]
    assert fa.corrected_total == 0.0
    assert fa.normalized_total == 200.0                       # peer price applied despite H5(a)/H5A drift
    assert any(g.startswith("H5(a)") for g in fa.scope_gaps)  # the returned-but-unpriced line is surfaced


def test_two_dispatched_units_each_get_a_sheet_even_when_only_one_replied(tmp_path):
    # Two enquiries, one priced return: BOTH units get a sheet (the unreplied one shows its canonical
    # rows + awaiting), and the Summary tab lists the unreplied unit without crashing on empty bids.
    unit_items = {
        "ground_investigation:H": [SorItem(item_ref="H1", description="Boreholes", section="H")],
        "ground_investigation:J": [SorItem(item_ref="J1", description="Instruments", section="J")],
    }
    reply = BidReply(firm_id="F-A", trade="ground_investigation:H",
                     line_items=[BidLineItem(item_ref="H1", qty=1.0, rate=50.0, amount=50.0)])
    out = export_leveling_xlsx(
        level_bids([reply]), [reply], path=tmp_path / "c.xlsx",
        units=["ground_investigation:H", "ground_investigation:J"], unit_items=unit_items,
        awaiting={"ground_investigation:J": ["F-B"]})
    wb = load_workbook(out)
    assert sheet_title("ground_investigation:H") in wb.sheetnames
    assert sheet_title("ground_investigation:J") in wb.sheetnames  # unreplied enquiry still has a sheet
    jtext = _texts(wb[sheet_title("ground_investigation:J")])
    assert "J1" in jtext and "F-B" in jtext                        # canonical row + awaiting note
    assert "awaiting return" in _texts(wb["Summary"])              # summary handled the empty unit
