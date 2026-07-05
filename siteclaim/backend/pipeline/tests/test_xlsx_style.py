"""Prompt 4 — the shared Excel style kit (presentation only, offline)."""

from openpyxl import Workbook, load_workbook

from pipeline._xlsx_style import (
    MONEY_FORMAT,
    NAVY,
    TINT,
    autofit,
    footer_note,
    money_cell,
    style_header,
    style_totals,
    title_block,
)


def _roundtrip(wb, tmp_path):
    path = tmp_path / "styled.xlsx"
    wb.save(path)
    return load_workbook(path)


def test_kit_styles_a_workbook_without_touching_values(tmp_path):
    wb = Workbook()
    ws = wb.active
    title_block(ws, "Levelled bid comparison", ["Project: Demo Tower", "Reference: GE/2026/14"])
    ws.append(["Item", "Description", "Rate"])
    header_row = ws.max_row  # capture AFTER append — a blank spacer row advances the cursor, not max_row
    style_header(ws, header_row, 3)
    ws.append(["E-01", "Switchboard", 1200000.0])
    money_cell(ws.cell(row=ws.max_row, column=3))
    ws.append(["", "TOTAL", 1200000.0])
    totals_row = ws.max_row
    style_totals(ws, totals_row, 3)
    money_cell(ws.cell(row=totals_row, column=3))
    footer_note(ws, "Exclusions (non-comparable)")
    autofit(ws, min_row=header_row)

    got = _roundtrip(wb, tmp_path).active
    # values intact
    assert got.cell(row=header_row, column=1).value == "Item"
    assert got.cell(row=header_row + 1, column=3).value == 1200000.0
    assert got.cell(row=1, column=1).value == "Levelled bid comparison"
    # header band: navy fill + bold white + frozen panes below
    head = got.cell(row=header_row, column=1)
    assert head.fill.fgColor.rgb.endswith(NAVY)
    assert head.font.bold and str(head.font.color.rgb).endswith("FFFFFF")
    assert got.freeze_panes == f"A{header_row + 1}"
    # totals: tint + bold; money format applied
    tot = got.cell(row=totals_row, column=2)
    assert tot.fill.fgColor.rgb.endswith(TINT) and tot.font.bold
    assert got.cell(row=totals_row, column=3).number_format == MONEY_FORMAT
    # autofit set explicit widths
    assert got.column_dimensions["B"].width and got.column_dimensions["B"].width >= 9


def test_money_cell_leaves_placeholder_text_alone(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["—"])  # a rate-only line's em-dash amount stays text, never formatted
    money_cell(ws.cell(row=1, column=1))
    got = _roundtrip(wb, tmp_path).active
    assert got.cell(row=1, column=1).value == "—"
    assert got.cell(row=1, column=1).number_format != MONEY_FORMAT


def test_styled_sor_sheet_keeps_the_deterministic_roundtrip(tmp_path):
    # The outbound SoR sheet carries the letterhead but its data rows are exactly the
    # package's items — and the returned sheet still parses with NO model.
    from pipeline.stage_03_dispatch.attachments import generate_sor_sheet
    from pipeline.stage_04_level.reply_xlsx import parse_sor_xlsx
    from schemas.models import SorItem, TradeWorkPackage

    pkg = TradeWorkPackage(
        trade="electrical", scope_summary="LV works",
        sor_items=[
            SorItem(item_ref="E-01", description="Switchboard", unit="no", qty=1.0),
            SorItem(item_ref="E-02", description="Sub-main cabling", unit="m", qty=420.0),
        ],
        source_refs=[],
    )
    out = generate_sor_sheet(pkg, "Demo Tower", tmp_path / "sor.xlsx")

    got = load_workbook(out).active
    assert got.cell(row=1, column=1).value == "Schedule of Rates — Electrical — Demo Tower"
    flat = [c.value for row in got.iter_rows() for c in row if c.value is not None]
    assert "Priced return requested by the tender date." in flat

    reply = parse_sor_xlsx(out.read_bytes(), firm_id="F-X", trade="electrical")
    assert [(li.item_ref, li.description, li.unit, li.qty) for li in reply.line_items] == [
        ("E-01", "Switchboard", "no", 1.0),
        ("E-02", "Sub-main cabling", "m", 420.0),
    ]
