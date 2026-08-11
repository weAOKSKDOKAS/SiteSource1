"""Excel drops a trailing zero, and two different items arrive under one key.

THE EVIDENCE, read off the real pack (CEDD ND/2025/04). Bill 1 in the PDF render is clean —
63 unique references, 63 occurrences, 1.1 … 1.63, no gaps, no duplicates, no letter suffixes. The
same bill in the editable workbook, the one that is actually priced::

    58 unique refs · 63 total occurrences
    DUPLICATE 1.2 x2 at rows [12, 98]
    DUPLICATE 1.3 x2 at rows [14, 143]
    DUPLICATE 1.4 x2 at rows [18, 171]
    DUPLICATE 1.5 x2 at rows [23, 219]
    DUPLICATE 1.6 x2 at rows [25, 265]

and the rows say why::

    r 96:  1.19  Servicing of core and sample store for the …   87  wk
    r 98:  1.2   Handing over of core and sample store to …     -   item    <- this is 1.20
    r141:  1.29  Provision                                      -   item
    r143:  1.3   Maintenance                                    -   item    <- this is 1.30

Excel stored the reference as a NUMBER, so ``1.20`` came back as ``1.2``. Five decade boundaries,
five collisions with the real items 1.2 … 1.6 near the top of the bill.

WHY IT MATTERS MORE THAN A MISCOUNT. The row count was never wrong — 63 rows in, 63 rows out, and
the reader keeping both copies was right. The IDENTITY was wrong, at source: *"Servicing"* (28 mth)
and *"Handing over of core and sample store"* (item) are two different items arriving under one key.
Every index downstream is keyed on the reference, so it collapses them, and the total counts one
line twice.

And it contradicts a standing assumption. The workbook path is the trusted one — it *"parses with
zero model calls … read back deterministically, with no interpretation."* True of the numbers, and
NOT true of the reference itself. A silently-corrupted key in the deterministic path is worse than
a noisy one in the model path, because nothing downstream doubts it.

`item_ref_of` already recovers this where the cell's number format says two decimals were printed.
On the real file that evidence is absent. What remains is the sheet's own ordering, and it is
decidable rather than a guess: a bill numbers its items in ascending order, so a reference that goes
BACKWARDS did not come from the document.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from pipeline.stage_01_ingest import workbook as wb
from pipeline.stage_01_ingest.workbook import compare_reference_sets, restore_dropped_zero


# --- the rule on its own -------------------------------------------------------------------------
class TestTheFiveRealCollisions:
    """Each pair is a decade boundary from the real pack, with the reference that precedes it."""

    @pytest.mark.parametrize("previous,rendered,restored", [
        ("1.19", "1.2", "1.20"),
        ("1.29", "1.3", "1.30"),
        ("1.39", "1.4", "1.40"),
        ("1.49", "1.5", "1.50"),
        ("1.59", "1.6", "1.60"),
    ])
    def test_the_trailing_zero_is_restored(self, previous, rendered, restored):
        ref, note = restore_dropped_zero(rendered, previous, from_number=True)
        assert ref == restored
        assert "Restored from the sequence" in note
        assert previous in note

    def test_the_restoration_is_not_silent(self):
        _ref, note = restore_dropped_zero("1.2", "1.19", from_number=True)
        assert note, "a repaired reference must say it was repaired"
        assert "trailing zero" in note


class TestAReferenceThatIsAlreadyRight:
    """The case that matters most: rewriting a correct reference invents a defect."""

    def test_one_point_two_after_one_point_one_is_item_two(self):
        assert restore_dropped_zero("1.2", "1.1", from_number=True) == ("1.2", "")

    @pytest.mark.parametrize("previous,rendered", [
        ("1.1", "1.2"), ("1.9", "1.10"), ("1.19", "1.20"), ("1.20", "1.21"), ("2.3", "2.4"),
    ])
    def test_a_forward_step_is_never_touched(self, previous, rendered):
        assert restore_dropped_zero(rendered, previous, from_number=True) == (rendered, "")

    def test_a_reference_stored_as_text_is_authoritative(self):
        """A PDF prints what was typed and a text cell holds what was typed. Neither loses a zero,
        so neither is reconstructed."""
        assert restore_dropped_zero("1.2", "1.19", from_number=False) == ("1.2", "")

    def test_the_first_item_of_a_sheet_has_nothing_to_compare_against(self):
        assert restore_dropped_zero("1.2", "", from_number=True) == ("1.2", "")

    def test_a_different_bill_is_not_compared(self):
        """Refs are only ordered within their own bill. 2.1 after 1.63 is a new sheet, not a jump."""
        assert restore_dropped_zero("2.1", "1.63", from_number=True) == ("2.1", "")

    def test_a_genuine_repeat_is_left_for_the_duplicate_check(self):
        """Two rows under one reference may be a real duplicate in the client's file. Rewriting the
        second would hide it; the reader's own duplicate report owns that case."""
        assert restore_dropped_zero("1.2", "1.2", from_number=True) == ("1.2", "")

    def test_a_lettered_reference_is_out_of_scope(self):
        assert restore_dropped_zero("1.61A", "1.60", from_number=True) == ("1.61A", "")


class TestWhatItRefusesToGuess:

    def test_a_backwards_jump_no_zero_explains_is_reported_and_kept(self):
        """1.5 after 1.60 cannot be 1.50 (that is still backwards) and cannot be 1.500. So it is
        kept exactly as printed and said out loud — repairing on a guess is how a priced row ends
        up under the wrong item."""
        ref, note = restore_dropped_zero("1.5", "1.600", from_number=True)
        assert ref == "1.5"
        assert "numbering goes backwards" in note
        assert "Kept exactly as printed" in note

    def test_a_hundred_boundary_needs_two_zeros(self):
        assert restore_dropped_zero("1.1", "1.99", from_number=True)[0] == "1.100"

    def test_the_smallest_reconstruction_wins(self):
        """1.2 after 1.19 is 1.20, never 1.200 — both move forward, and the bill's next item is
        the one immediately after it."""
        assert restore_dropped_zero("1.2", "1.19", from_number=True)[0] == "1.20"


# --- the rule inside the reader ------------------------------------------------------------------
def _book(build) -> bytes:
    book = Workbook()
    book.remove(book.active)
    build(book)
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def _furniture(sheet, bill: int) -> None:
    sheet.append([f"Bill No. {bill}", "Description", None, None, "Quantity", "Unit", "Rate",
                  "Amount"])
    sheet.append(["Item No.", "Item Description", None, None, None, None, None, "(HK$)"])


def _bill_1(book) -> None:
    """A bill that walks its decade boundary, exactly as the real one does."""
    sheet = book.create_sheet("Bill No.1")
    _furniture(sheet, 1)
    sheet.append([1.1, "Taking over", None, None, None, "item", "-", "=G3"])
    sheet.append([1.2, "Servicing", None, None, 28, "mth", None, "=E4*G4"])
    sheet.append([1.19, "Servicing of core and sample store", None, None, 87, "wk", None, "=E5*G5"])
    # Stored as the number 1.2 — the real row 98.
    sheet.append([1.2, "Handing over of core and sample store", None, None, None, "item", "-",
                  "=G6"])
    sheet.append([1.21, "Removal", None, None, None, "item", "-", "=G7"])


class TestTheReaderKeepsTheTwoItemsApart:

    @pytest.fixture
    def read(self):
        return wb.read_workbook(_book(_bill_1))

    def test_both_items_survive_with_their_own_identity(self, read):
        refs = [i.item_ref for i in read.items]
        assert refs == ["1.1", "1.2", "1.19", "1.20", "1.21"]

    def test_the_two_colliding_items_keep_their_own_descriptions(self, read):
        by_ref = {i.item_ref: i for i in read.items}
        assert by_ref["1.2"].description == "Servicing"
        assert by_ref["1.20"].description == "Handing over of core and sample store"

    def test_the_index_no_longer_collapses_them(self, read):
        """The harm, stated as the thing that stops happening: five rows, five distinct keys."""
        assert len({i.item_ref for i in read.items}) == len(read.items) == 5

    def test_each_is_counted_once(self, read):
        """Two items under one key were summed twice. Their quantities are different, so summing
        the pair is the check."""
        by_ref = {i.item_ref: i for i in read.items}
        assert by_ref["1.2"].qty == 28
        assert by_ref["1.20"].qty is None, "a lump item carries no quantity"

    def test_the_restoration_is_reported_to_the_caller(self):
        notes: list[str] = []
        wb.read_workbook(_book(_bill_1), on_note=notes.append)
        assert any("trailing zero" in n and "1.20" in n for n in notes)

    def test_no_duplicate_is_reported_because_there_is_none(self):
        notes: list[str] = []
        wb.read_workbook(_book(_bill_1), on_note=notes.append)
        assert not any("appears" in n and "times" in n for n in notes), notes


class TestABillThatNeverLosesAZero:
    """The negative control. A bill whose references all move forward must read back verbatim."""

    def test_nothing_is_rewritten(self):
        def build(book):
            sheet = book.create_sheet("Bill No.3")
            _furniture(sheet, 3)
            for n in range(1, 10):
                sheet.append([float(f"3.{n}"), f"Item {n}", None, None, 1, "nr", None,
                              f"=E{n + 2}*G{n + 2}"])

        notes: list[str] = []
        read = wb.read_workbook(_book(build), on_note=notes.append)
        assert [i.item_ref for i in read.items] == [f"3.{n}" for n in range(1, 10)]
        assert not any("trailing zero" in n for n in notes)


# --- the independent check ------------------------------------------------------------------------
class TestTheWorkbookVsRenderGuard:
    """The pack ships both. The render carries the reference better; the workbook carries
    everything else better. A disagreement is reported, never resolved here."""

    def test_it_fires_when_the_reference_sets_disagree(self):
        notes = compare_reference_sets({"1.1", "1.2", "1.3"}, {"1.1", "1.2", "1.20"})
        assert len(notes) == 2
        assert "the PDF render of this bill carries 1 reference(s) the workbook does not: 1.20" \
            in notes[0]
        assert "the workbook carries 1 reference(s) the PDF render does not: 1.3" in notes[1]

    def test_it_names_the_dropped_zero_as_the_usual_cause(self):
        notes = compare_reference_sets({"1.1"}, {"1.1", "1.20"})
        assert "drops its trailing zero" in notes[0]

    def test_it_decides_nothing(self):
        """Which document is right about a reference is a question about the client's pack."""
        notes = compare_reference_sets({"1.3"}, {"1.30"})
        assert any("nothing here decides which" in n for n in notes)

    def test_agreement_is_stated_rather_than_left_silent(self):
        notes = compare_reference_sets({"1.1", "1.2"}, {"1.1", "1.2"})
        assert notes == ["the workbook and the PDF render agree on all 2 item reference(s)."]

    def test_a_missing_side_is_not_a_disagreement(self):
        """No render, or a render nothing could be read from, is a check that did not run."""
        assert compare_reference_sets({"1.1"}, set()) == []
        assert compare_reference_sets(set(), {"1.1"}) == []

    def test_the_real_bill_1_shape(self):
        """63 clean references in the render against the 58 the workbook carried."""
        render = {f"1.{n}" for n in range(1, 64)}
        book = render - {"1.20", "1.30", "1.40", "1.50", "1.60"}
        notes = compare_reference_sets(book, render)
        assert len(notes) == 1, "the workbook is a strict subset — nothing is only in the workbook"
        assert "carries 5 reference(s) the workbook does not" in notes[0]
        for ref in ("1.20", "1.30", "1.40", "1.50", "1.60"):
            assert ref in notes[0]
