"""Deterministic xlsx reply parsing — the dispatched SoR sheet returned with rates.

The round-trip is against the real writer (``generate_sor_sheet``): build the sheet the
subcontractor actually receives, fill the Rate column as they would, and parse it back
into a ``BidReply`` with pure openpyxl — no model, no network, no SDK import.
"""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from pipeline.stage_03_dispatch.attachments import generate_sor_sheet
from pipeline.stage_04_level.reply_xlsx import is_xlsx_upload, parse_sor_xlsx
from schemas.models import SorItem, TradeWorkPackage

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _dispatched_sheet(tmp_path) -> tuple[bytes, object]:
    """The sheet a firm receives: three items, one of them rate-only (no qty)."""
    pkg = TradeWorkPackage(
        trade="electrical",
        scope_summary="LV works",
        sor_items=[
            SorItem(item_ref="E-01", description="LV main switchboard", unit="no", qty=1.0),
            SorItem(item_ref="E-02", description="Sub-main cabling", unit="m", qty=100.0),
            SorItem(item_ref="E-03", description="Maintenance visit (rate only)", unit="visit"),
        ],
        source_refs=["Schedule of Rates"],
    )
    path = generate_sor_sheet(pkg, "Kwun Tong Commercial Tower", tmp_path / "sor.xlsx")
    return path.read_bytes(), path


def _fill_rates(path, rates: dict[str, object]) -> bytes:
    """Fill the Rate column of a dispatched sheet, exactly as a subcontractor would."""
    wb = load_workbook(path)
    ws = wb.active
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Item")
    for r in range(header_row + 1, ws.max_row + 1):
        ref = ws.cell(row=r, column=1).value
        if ref in rates:
            ws.cell(row=r, column=5, value=rates[ref])  # "Rate (HKD)"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_filled_dispatch_sheet_round_trips_into_a_bid_reply(tmp_path):
    _, path = _dispatched_sheet(tmp_path)
    filled = _fill_rates(path, {"E-01": 950000, "E-02": 2500, "E-03": 800})

    reply = parse_sor_xlsx(filled, firm_id="F-EL-02", trade="electrical")

    assert reply.firm_id == "F-EL-02" and reply.trade == "electrical"
    by_ref = {li.item_ref: li for li in reply.line_items}
    assert set(by_ref) == {"E-01", "E-02", "E-03"}
    assert by_ref["E-01"].rate == 950000.0 and by_ref["E-01"].qty == 1.0
    assert by_ref["E-02"].rate == 2500.0 and by_ref["E-02"].unit == "m"
    assert by_ref["E-03"].rate == 800.0 and by_ref["E-03"].qty is None  # rate-only line
    assert by_ref["E-01"].description == "LV main switchboard"


def test_blank_rate_rows_come_through_unpriced(tmp_path):
    _, path = _dispatched_sheet(tmp_path)
    filled = _fill_rates(path, {"E-01": 950000})  # E-02 / E-03 left blank

    reply = parse_sor_xlsx(filled)

    by_ref = {li.item_ref: li for li in reply.line_items}
    assert by_ref["E-01"].rate == 950000.0
    assert by_ref["E-02"].rate is None and by_ref["E-03"].rate is None  # unpriced, not zero


def test_text_typed_rates_and_totals_rows_are_tolerated(tmp_path):
    # A sub types "2,450" as text and appends a totals row with no item ref.
    _, path = _dispatched_sheet(tmp_path)
    filled = _fill_rates(path, {"E-02": "2,450"})
    wb = load_workbook(BytesIO(filled))
    ws = wb.active
    ws.append(["", "TOTAL", "", "", "", 245000])  # no item ref -> must be skipped
    buffer = BytesIO()
    wb.save(buffer)

    reply = parse_sor_xlsx(buffer.getvalue())

    by_ref = {li.item_ref: li for li in reply.line_items}
    assert by_ref["E-02"].rate == 2450.0                       # comma text parsed
    assert set(by_ref) == {"E-01", "E-02", "E-03"}             # totals row skipped, never a ref


def test_non_sor_xlsx_raises_a_clear_value_error():
    wb = Workbook()
    ws = wb.active
    ws.append(["Colour", "Size", "Price"])  # wrong headers — not our SoR layout
    ws.append(["red", "M", 10])
    buffer = BytesIO()
    wb.save(buffer)

    with pytest.raises(ValueError, match="Not a Schedule of Rates sheet"):
        parse_sor_xlsx(buffer.getvalue())


def test_garbage_bytes_and_empty_bytes_raise_value_error():
    with pytest.raises(ValueError, match="Could not read the xlsx"):
        parse_sor_xlsx(b"this is not a zip archive")
    with pytest.raises(ValueError, match="Empty file"):
        parse_sor_xlsx(b"")


def test_is_xlsx_upload_detects_extension_and_content_type():
    assert is_xlsx_upload("reply.xlsx", "application/octet-stream")  # extension wins
    assert is_xlsx_upload("SoR_electrical.XLSX", None)
    assert is_xlsx_upload("reply.bin", XLSX_CT)                      # content type wins
    assert not is_xlsx_upload("reply.pdf", "application/pdf")
    assert not is_xlsx_upload(None, None)
