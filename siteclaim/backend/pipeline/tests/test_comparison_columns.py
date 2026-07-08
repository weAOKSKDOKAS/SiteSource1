"""The unit comparison attributes one rate column per firm and never bleeds across firms, and notes
firms still awaited (replies-meet-enquiries, Commit 3). Reads the real workbook with openpyxl."""

from openpyxl import load_workbook

from pipeline.stage_04_level.export_xlsx import export_leveling_xlsx, sheet_title
from pipeline.stage_04_level.level import level_bids
from schemas.models import BidLineItem, BidReply


def _reply(firm, key, ref, rate):
    return BidReply(firm_id=firm, trade=key,
                    line_items=[BidLineItem(item_ref=ref, qty=1.0, rate=rate, amount=rate)])


def _texts(ws):
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def test_one_rate_column_per_firm_on_the_unit_sheet_no_bleed(tmp_path):
    a = _reply("F-A", "ground_investigation:H", "H12", 100.0)
    b = _reply("F-B", "ground_investigation:H", "H12", 120.0)
    levelled = level_bids([a, b])  # firm names fall back to the id for unknown firms
    out = export_leveling_xlsx(levelled, [a, b], path=tmp_path / "c.xlsx",
                               awaiting={"ground_investigation:H": ["F-C"]})
    ws = load_workbook(out)[sheet_title("ground_investigation:H")]
    texts = _texts(ws)
    # one rate + corrected column per replying firm — both present, attributed by name
    assert "F-A — rate" in texts and "F-A — corrected" in texts
    assert "F-B — rate" in texts and "F-B — corrected" in texts
    # the enquired-but-not-replied firm is a coverage NOTE, never a priced column
    assert "Awaiting reply (enquired, no priced return yet)" in texts and "F-C" in texts
    assert "F-C — rate" not in texts


def test_two_units_are_separate_sheets_each_with_its_own_firm(tmp_path):
    # firm A priced unit H; firm B priced unit J — never mixed into one sheet.
    a = _reply("F-A", "ground_investigation:H", "H12", 100.0)
    b = _reply("F-B", "ground_investigation:J", "J1", 300.0)
    levelled = level_bids([a, b])
    out = export_leveling_xlsx(levelled, [a, b], path=tmp_path / "c.xlsx")
    wb = load_workbook(out)
    h, j = wb[sheet_title("ground_investigation:H")], wb[sheet_title("ground_investigation:J")]
    assert "F-A — rate" in _texts(h) and "F-B — rate" not in _texts(h)  # H sheet: only F-A
    assert "F-B — rate" in _texts(j) and "F-A — rate" not in _texts(j)  # J sheet: only F-B
