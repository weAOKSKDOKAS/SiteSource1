"""Document routing and the generated SoR sheet (Phase A, build plan §5).

Turns the labels in ``bundle_doc_refs`` into real, routed :class:`BundleAttachment`
files so a subcontractor's email carries exactly what that trade needs to price:

1. **General documents** — those every trade needs (form of tender, conditions,
   general preliminaries). A :class:`TenderDocument` with an empty ``trades`` list is
   general and goes to every firm, whole.
2. **Trade-specific documents** — a :class:`TenderDocument` whose ``trades`` names this
   trade is routed only to that trade, whole.
3. **The generated SoR sheet** — a clean per-trade Excel of the priceable items
   (``sor_items``) with blank Rate/Amount columns for the subcontractor to fill. It is
   labelled an *excerpt*, with the full package available on request, so a derived
   summary never stands in for the legal document (the §5 safety rule).

Whole-file routing is the safe default the plan calls for: no page is sliced out of a
combined legal PDF here. Routing is deterministic Layer-1 Python — no model decides
which file goes to which firm.

``source_path`` is filled only when the real original exists in the tender
:class:`~pipeline.workspace.Workspace` (the live path). Offline, attachments are still
described (filename, kind, trade) with ``source_path=None``; the SoR sheet, which is
derived purely from ``sor_items``, is generated whenever a workspace is given, so even
a demo tender can produce a real, openable sheet.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from pipeline.workspace import Workspace
from schemas.models import (
    AttachmentKind,
    BundleAttachment,
    ScopePackages,
    TenderDocument,
    TenderPackage,
    TradeWorkPackage,
)


def route_documents(
    tender: Optional[TenderPackage], trade: str
) -> tuple[list[TenderDocument], list[TenderDocument]]:
    """Split ``tender``'s documents into (general, this-trade-specific)."""
    if tender is None:
        return [], []
    general = [d for d in tender.documents if not d.trades]
    specific = [d for d in tender.documents if d.trades and trade in d.trades]
    return general, specific


def generate_sor_sheet(pkg: TradeWorkPackage, project_name: str, path: Path | str) -> Path:
    """Write this trade's priceable-items sheet to ``path`` and return it."""
    from openpyxl import Workbook  # lazy — dispatch must not require openpyxl offline
    from openpyxl.styles import Font

    trade_label = pkg.trade.replace("_", " ").title()
    wb = Workbook()
    ws = wb.active
    ws.title = "SoR"

    ws.append([f"{project_name} — {trade_label}"])
    ws[ws.max_row][0].font = Font(bold=True, size=13)
    ws.append([
        "Excerpt of the priceable items for this trade. The full tender package is "
        "available on request. Please price every line; state any exclusions."
    ])
    ws.append([])

    header = ["Item", "Description", "Unit", "Qty", "Rate (HKD)", "Amount (HKD)"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for item in pkg.sor_items:
        ws.append([item.item_ref, item.description, item.unit, item.qty, "", ""])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _doc_attachment(
    doc: TenderDocument, kind: AttachmentKind, trade: Optional[str], workspace: Optional[Workspace], tender_id: str
) -> BundleAttachment:
    source_path: Optional[str] = None
    if workspace is not None:
        candidate = workspace.doc_path(tender_id, doc.filename)
        if candidate.is_file():
            source_path = str(candidate)
    label = doc.filename if kind is AttachmentKind.GENERAL else f"{doc.filename} ({trade})"
    return BundleAttachment(
        filename=doc.filename, kind=kind, trade=trade, source_path=source_path, label=label
    )


def build_attachments(
    trade: str,
    scope: Optional[ScopePackages],
    tender: Optional[TenderPackage],
    *,
    project_name: str = "",
    tender_id: str = "",
    workspace: Optional[Workspace] = None,
) -> list[BundleAttachment]:
    """Assemble the routed attachment list for one trade.

    General documents + this trade's specific documents + the generated SoR sheet.
    When ``workspace`` is given the SoR sheet is really written and its ``source_path``
    set (and each routed original's path resolved if present); otherwise attachments
    are described without touching the disk.
    """
    tender_id = tender_id or project_name
    general, specific = route_documents(tender, trade)

    attachments: list[BundleAttachment] = [
        _doc_attachment(d, AttachmentKind.GENERAL, None, workspace, tender_id) for d in general
    ]
    attachments += [
        _doc_attachment(d, AttachmentKind.TRADE_SPECIFIC, trade, workspace, tender_id) for d in specific
    ]

    pkg = None
    if scope is not None:
        pkg = next((p for p in scope.packages if p.trade == trade), None)
    if pkg is not None:
        sheet_source: Optional[str] = None
        if workspace is not None:
            sheet_source = str(generate_sor_sheet(pkg, project_name, workspace.sor_sheet_path(tender_id, trade)))
        trade_slug = trade.replace(" ", "_")
        attachments.append(
            BundleAttachment(
                filename=f"SoR_{trade_slug}.xlsx",
                kind=AttachmentKind.SOR_SHEET,
                trade=trade,
                source_path=sheet_source,
                generated=True,
                label=f"{trade} — Schedule of Rates (excerpt to price)",
            )
        )
    return attachments
