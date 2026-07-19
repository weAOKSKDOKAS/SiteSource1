"""Deterministic pricing-workbook export: sheet structure and figures equal to the persisted estimate."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from client_boq.estimate.run import assemble_estimate, load_demo_schedule
from client_boq.estimate.workbook import build_workbook


def _estimate():
    return assemble_estimate("t", 15.0, load_demo_schedule())


def _wb():
    return load_workbook(io.BytesIO(build_workbook(_estimate())))


def test_sheets_present_including_one_per_activity() -> None:
    est = _estimate()
    wb = _wb()
    assert {"WBS", "Resources", "Indirect Costs", "Flags"} <= set(wb.sheetnames)
    for act in est.activities:
        assert act.item_id in wb.sheetnames


def test_wbs_totals_and_marked_up_equal_estimate() -> None:
    est = _estimate()
    ws = _wb()["WBS"]
    cells = {}
    activity_rows = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if row[0] in {a.item_id for a in est.activities}:
            activity_rows[row[0]] = row
        elif isinstance(row[0], str):
            cells[row[0]] = row[1]

    # Top-block totals equal the persisted estimate to the cent.
    assert cells["Total"] == est.totals.total_cost
    assert cells["Profit"] == est.totals.margin_amount
    assert cells["Total Direct Costs"] == est.totals.total_direct
    assert cells["Total Indirect Costs"] == est.totals.total_indirect

    # Each activity row: Total = activity_total; Marked Up = Total × (1 + margin).
    for a in est.activities:
        r = activity_rows[a.item_id]
        assert r[5] == a.activity_total                                    # Total column
        assert r[7] == round(a.activity_total * (1 + est.totals.margin_pct / 100.0), 2)  # Marked Up


def test_activity_sheets_sum_to_their_wbs_row() -> None:
    est = _estimate()
    wb = _wb()
    for a in est.activities:
        ws = wb[a.item_id]
        total = 0.0
        for row in ws.iter_rows(min_row=6, values_only=True):   # data rows start after the header at row 5
            if row and isinstance(row[7], (int, float)):
                total += row[7]                                  # Total Cost column
        assert round(total, 2) == a.activity_total


def test_flags_sheet_lists_every_flag() -> None:
    est = _estimate()
    ws = _wb()["Flags"]
    kinds = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    assert {f.kind for f in est.flags} <= kinds
