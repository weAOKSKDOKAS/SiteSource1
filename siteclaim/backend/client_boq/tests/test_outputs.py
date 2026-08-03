"""Spec for T5 (Departure Schedule) and T6 (Letter of Qualifications), plus strategy-flag detection.

The finding that shapes all three: **both reference tenders penalise qualifying a bid.**

    "Any qualification of the tender may cause the tender to be disqualified."
        -- ND/2025/04, General Conditions of Tender GCT 4, page 6
    "Any qualification of tender or of the tender documents may cause the tender to be
     disqualified."
        -- CIC (325), Conditions of Tender 4.26, page 8

So these documents are internal working papers by default, the submission version is opt-in and
carries the tender's own words as a warning, and the app flags the rule at ingest rather than
letting it be discovered after the query cut-off has passed.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

fitz = pytest.importorskip("fitz")  # PyMuPDF


@pytest.fixture
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "workspace"))
    from api import app

    return TestClient(app)


def _pdf(pages: int, toc: list | None = None) -> bytes:
    doc = fitz.open()
    for i in range(1, pages + 1):
        doc.new_page().insert_text((72, 100), f"Binder page {i} of {pages}, with readable text",
                                   fontsize=11)
    if toc:
        doc.set_toc(toc)
    data = doc.tobytes()
    doc.close()
    return data


def _priced(client: TestClient, name: str = "outputs-demo") -> str:
    """A set carried all the way through: ingested, reviewed, approved, scoped and priced."""
    resp = client.post("/client-boq/ingest/upload", data={"project_name": name},
                       files={"files": ("binder.pdf", _pdf(12, [
                           [1, "Conditions of Tender", 1], [1, "Scope of Works", 5],
                           [1, "Pricing Schedule", 9]]), "application/pdf")})
    set_id = resp.json()["result"]["set_id"]
    client.post("/client-boq/ingest/manifest/approve", json={"set_id": set_id})
    client.post("/client-boq/ingest/split", json={"set_id": set_id})
    client.post("/client-boq/review/run", data={"project_name": name, "set_id": set_id})
    return set_id


def _items(client: TestClient, set_id: str) -> list[dict]:
    return client.get(f"/client-boq/review/register/{set_id}").json()["register"]["items"]


def _confirm_some(client: TestClient, set_id: str) -> dict:
    """Confirm a couple of departures and query one, then approve."""
    items = _items(client, set_id)
    ok = [i for i in items if i["status"] not in ("citation_failed", "unresolved")]
    decisions = {str(ok[0]["item"]): "confirmed", str(ok[1]["item"]): "confirmed",
                 str(ok[2]["item"]): "query"}
    client.post("/client-boq/review/approve",
                json={"set_id": set_id, "decisions": decisions, "approved": True})
    return {"confirmed": [ok[0]["item"], ok[1]["item"]], "queried": ok[2]["item"]}


# ---------------------------------------------------------------------------
# Strategy flags detected at ingest
# ---------------------------------------------------------------------------
def test_ingest_flags_a_tender_that_penalises_qualifications(client):
    set_id = _priced(client)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()

    assert parts["penalises_qualifications"] is True
    flag = next(f for f in parts["strategy_flags"]
                if f["kind"] == "qualifications-penalised")
    assert "disqualified" in flag["quote"]
    assert flag["clause"] and flag["page"]      # citable, so a human can check it
    assert flag["part_id"]                      # and traceable to the part it was found in


def test_other_bidding_conditions_are_surfaced_too(client):
    set_id = _priced(client)
    kinds = {f["kind"] for f in
             client.get(f"/client-boq/ingest/parts/{set_id}").json()["strategy_flags"]}
    assert {"no-alterations", "two-envelope"} <= kinds


# ---------------------------------------------------------------------------
# T5 — Departure Schedule
# ---------------------------------------------------------------------------
def test_the_schedule_lists_confirmed_departures_with_their_clause(client):
    set_id = _priced(client)
    marked = _confirm_some(client, set_id)

    md = client.get(f"/client-boq/review/{set_id}/departure-schedule").text
    assert "# Departure Schedule" in md
    assert "internal working copy" in md          # internal is the default
    for item in marked["confirmed"]:
        assert f"| {item} |" in md


def test_an_open_query_appears_marked_as_unresolved(client):
    # Decided: a queried line belongs on the schedule, flagged, because the schedule should show
    # every unresolved contractual point.
    set_id = _priced(client)
    marked = _confirm_some(client, set_id)

    md = client.get(f"/client-boq/review/{set_id}/departure-schedule").text
    assert "Subject to outstanding clarification" in md
    assert f"| {marked['queried']} |" in md
    assert "Still with the client" in md          # and the internal copy says what happens next


def test_a_line_whose_citation_failed_is_withheld_and_reported(client):
    # You cannot ask a client to amend a clause you could not locate in their own document.
    set_id = _priced(client)
    failed = [i for i in _items(client, set_id) if i["status"] == "citation_failed"]
    if not failed:
        pytest.skip("this fixture register has no citation_failed line")
    client.post("/client-boq/review/approve",
                json={"set_id": set_id, "decisions": {}, "approved": True})

    md = client.get(f"/client-boq/review/{set_id}/departure-schedule").text
    assert "Withheld from this schedule" in md
    for item in failed:
        assert f"- Item {item['item']}" in md     # named, not silently dropped


def test_the_submission_version_carries_the_tenders_own_warning(client):
    set_id = _priced(client)
    _confirm_some(client, set_id)

    internal = client.get(f"/client-boq/review/{set_id}/departure-schedule").text
    submission = client.get(
        f"/client-boq/review/{set_id}/departure-schedule?audience=submission").text

    assert "Do not submit this without a decision" not in internal
    assert "Do not submit this without a decision" in submission
    assert "may cause the tender to be disqualified" in submission   # quoted from the tender
    assert "internal working copy" not in submission


def test_the_schedule_downloads_as_xlsx(client):
    from openpyxl import load_workbook

    set_id = _priced(client)
    _confirm_some(client, set_id)

    resp = client.get(f"/client-boq/review/{set_id}/departure-schedule?format=xlsx")
    assert resp.status_code == 200 and resp.content[:2] == b"PK"
    book = load_workbook(io.BytesIO(resp.content))
    assert book.sheetnames == ["Departure Schedule", "Notes"]
    sheet = book["Departure Schedule"]
    # The header sits below the title, and below the risk banner when there is one, so find it
    # rather than assume a row.
    header = next(row for row in sheet.iter_rows(values_only=True) if row[0] == "Item")
    assert list(header[:3]) == ["Item", "Clause", "As drafted"]
    assert sheet.max_row > sheet.min_row       # and there are data rows under it

    risky = client.get(
        f"/client-boq/review/{set_id}/departure-schedule?format=xlsx&audience=submission")
    book = load_workbook(io.BytesIO(risky.content))
    banner = "\n".join(str(c.value) for row in book["Departure Schedule"].iter_rows(max_row=3)
                       for c in row if c.value)
    assert "penalises qualifying the bid" in banner


def test_bad_audience_and_format_are_refused(client):
    set_id = _priced(client)
    _confirm_some(client, set_id)
    assert client.get(
        f"/client-boq/review/{set_id}/departure-schedule?audience=public").status_code == 422
    assert client.get(
        f"/client-boq/review/{set_id}/departure-schedule?format=pdf").status_code == 422
    assert client.get("/client-boq/review/nope/departure-schedule").status_code == 404


# ---------------------------------------------------------------------------
# T6 — Letter of Qualifications
# ---------------------------------------------------------------------------
def _price(client: TestClient, set_id: str) -> None:
    client.post("/client-boq/estimate/scope", json={"set_id": set_id})
    client.post("/client-boq/estimate/scope/approve", json={"set_id": set_id, "approved": True})
    client.post("/client-boq/estimate/run", json={"set_id": set_id})


def test_qualifications_state_the_assumptions_and_their_source(client):
    set_id = _priced(client)
    _confirm_some(client, set_id)
    _price(client, set_id)

    md = client.get(f"/client-boq/estimate/{set_id}/qualifications").text
    assert "# Letter of Qualifications" in md
    assert "Qualifications and assumptions" in md
    assert "source: confirmed departure" in md    # nothing free-floating in the internal copy
    assert "Tender sum:" in md


def test_an_unanswered_query_becomes_a_priced_assumption(client):
    # This is what T4 made possible, and the honest version of what every contractor does anyway.
    set_id = _priced(client)
    _confirm_some(client, set_id)
    client.post("/client-boq/rfi", json={
        "set_id": set_id, "origin": "pricing", "clause": "4.12",
        "question": "Is the temporary access road priced by the contractor?"})
    _price(client, set_id)

    md = client.get(f"/client-boq/estimate/{set_id}/qualifications").text
    assert "temporary access road" in md
    assert "sought clarification and had no answer" in md
    assert "source: unanswered query" in md
    assert "still with the client" in md.lower()


def test_an_answered_query_is_not_carried_as_an_assumption(client):
    set_id = _priced(client)
    _confirm_some(client, set_id)
    raised = client.post("/client-boq/rfi", json={
        "set_id": set_id, "question": "Who supplies the crane?"}).json()["rfi"]["rfi_id"]
    client.post("/client-boq/rfi/answer", json={
        "set_id": set_id, "rfi_id": raised, "answer": "The contractor does."})
    _price(client, set_id)

    md = client.get(f"/client-boq/estimate/{set_id}/qualifications").text
    assert "Who supplies the crane?" not in md   # answered, so there is nothing to assume


def test_the_submission_letter_reads_as_a_letter_and_carries_the_warning(client):
    set_id = _priced(client)
    _confirm_some(client, set_id)
    _price(client, set_id)

    md = client.get(f"/client-boq/estimate/{set_id}/qualifications?audience=submission").text
    assert "Dear Sirs," in md and "Yours faithfully," in md
    assert "form part of our offer" in md
    assert "may cause the tender to be disqualified" in md
    assert "source: confirmed departure" not in md   # provenance is internal-only


def test_qualifications_warn_when_nothing_is_approved_yet(client):
    set_id = _priced(client)
    md = client.get(f"/client-boq/estimate/{set_id}/qualifications").text
    assert "Not final" in md
    assert "review register is not approved" in md


def test_qualifications_404_for_an_unknown_set(client):
    assert client.get("/client-boq/estimate/nope/qualifications").status_code == 404
    assert client.get(
        "/client-boq/estimate/nope/qualifications?audience=bad").status_code == 422


# ---------------------------------------------------------------------------
# The offer letter now points at both, instead of restating them
# ---------------------------------------------------------------------------
def test_the_offer_letter_references_the_attachments(client):
    set_id = _priced(client)
    _confirm_some(client, set_id)
    _price(client, set_id)

    md = client.get(f"/client-boq/estimate/{set_id}/letter").json()["markdown"]
    assert "subject to the accompanying **Departure Schedule** and **Letter of Qualifications**" in md
    assert "Both form part of this offer" in md


def test_the_output_routes_are_mounted(client):
    paths = set(client.app.openapi()["paths"])  # openapi(), not app.routes — CLAUDE.md trap 1
    assert {
        "/client-boq/review/{set_id}/departure-schedule",
        "/client-boq/estimate/{set_id}/qualifications",
    } <= paths
