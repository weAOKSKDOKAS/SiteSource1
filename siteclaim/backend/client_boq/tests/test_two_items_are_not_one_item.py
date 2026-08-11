"""The costing side of the dropped trailing zero: two items must stay two items.

The reader's job is the identity. This is what depends on it.

`ClientBill.index()` and `PricedBQ.index()` are both keyed on `full_ref`, so while Excel's
`1.20 → 1.2` corruption survived the read, two genuinely different items — *"Servicing"* at 28 mth
and *"Handing over of core and sample store"* as a lump — arrived under one key. The index kept
whichever came last, so a screen reading through it showed one item where the bill printed two;
the workbook sheet printed both rows, so its `SUM` counted the pair; and a rate typed against
`1.2` applied to both.

None of that was a costing bug. It was an identity bug, three layers up, and these tests hold the
line where the money is: distinct references in, distinct references priced, each summed once.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from client_boq.boq.reader import read_workbook


def _workbook(tmp_path, rows):
    """A one-sheet bill in the shape `reader.read_workbook` expects: A=ref, B=description,
    E=quantity, F=unit."""
    book = Workbook()
    book.remove(book.active)
    sheet = book.create_sheet("Bill No.1")
    sheet["A1"] = "Bill No. 1"
    sheet["B1"] = "General and Preliminaries"
    sheet["A2"] = "Item No."
    sheet["B2"] = "Item Description"
    for n, (ref, description, qty, unit) in enumerate(rows, start=4):
        sheet.cell(row=n, column=1, value=ref)
        sheet.cell(row=n, column=2, value=description)
        if qty is not None:
            sheet.cell(row=n, column=5, value=qty)
        sheet.cell(row=n, column=6, value=unit)
    path = tmp_path / "bill.xlsx"
    book.save(path)
    return path


#: The real pack's collision, minimally: item 1.2 near the top, item 1.20 after 1.19, and both
#: stored as the number 1.2 under a General format that says nothing.
COLLIDING = [
    (1.1, "Taking over", None, "item"),
    (1.2, "Servicing", 28, "mth"),
    (1.19, "Servicing of core and sample store", 87, "wk"),
    (1.2, "Handing over of core and sample store", None, "item"),
    (1.21, "Removal", None, "item"),
]


@pytest.fixture
def bill(tmp_path):
    return read_workbook(_workbook(tmp_path, COLLIDING), set_id="t", rev=0)


class TestTheBillKeepsThemApart:

    def test_five_rows_read_as_five_items(self, bill):
        assert len(bill.items) == 5

    def test_every_reference_is_distinct(self, bill):
        refs = [i.full_ref for i in bill.items]
        assert refs == ["1.1", "1.2", "1.19", "1.20", "1.21"]
        assert len(set(refs)) == 5

    def test_the_index_no_longer_loses_one_of_them(self, bill):
        """The harm, as the thing that stops happening. Keyed on the reference, the index kept
        whichever row came last and a screen showed one item where the bill printed two."""
        index = bill.index()
        assert len(index) == 5
        assert index["1.2"].description == "Servicing"
        assert index["1.20"].description == "Handing over of core and sample store"

    def test_they_carry_their_own_quantities(self, bill):
        index = bill.index()
        assert index["1.2"].qty == 28 and index["1.2"].unit == "mth"
        assert index["1.20"].qty is None and index["1.20"].unit == "item"

    def test_the_restoration_is_on_the_item_that_was_restored(self, bill):
        restored = bill.index()["1.20"]
        assert any("trailing zero" in note for note in restored.notes)
        assert not any("trailing zero" in note for note in bill.index()["1.2"].notes)


class TestARateBelongsToOneItem:

    def test_a_rate_typed_against_one_does_not_reach_the_other(self, bill):
        """The consequence that reaches money. Two items under one key meant one typed rate
        priced both — and their quantities differ, so the amounts were not even the same wrong
        number twice."""
        index = bill.index()
        assert index["1.2"] is not index["1.20"]
        assert index["1.2"].full_ref != index["1.20"].full_ref

    def test_each_is_summed_once(self, bill):
        """A `SUM` over printed rows counts what the rows carry. With distinct keys there is one
        row per item and one item per key, so the two are no longer the same money twice."""
        from collections import Counter

        counts = Counter(i.full_ref for i in bill.items)
        assert set(counts.values()) == {1}


class TestABillThatWasNeverCorruptedIsUntouched:
    """The negative control, and the case the rule must never break: a plain ascending bill."""

    def test_nothing_is_renumbered(self, tmp_path):
        rows = [(float(f"1.{n}"), f"Item {n}", 1, "nr") for n in range(1, 10)]
        bill = read_workbook(_workbook(tmp_path, rows), set_id="t", rev=0)
        assert [i.full_ref for i in bill.items] == [f"1.{n}" for n in range(1, 10)]
        assert not any("trailing zero" in n for i in bill.items for n in i.notes)

    def test_a_reference_stored_as_text_is_taken_verbatim(self, tmp_path):
        rows = [("1.19", "Nineteen", 1, "nr"), ("1.2", "Two, as typed", 1, "nr")]
        bill = read_workbook(_workbook(tmp_path, rows), set_id="t", rev=0)
        assert [i.full_ref for i in bill.items] == ["1.19", "1.2"], (
            "a text cell holds what was typed and loses no zero, so it is authoritative")
