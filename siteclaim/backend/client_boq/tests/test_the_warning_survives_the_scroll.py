"""The provisional warning gets louder, not quieter, on its way into the deliverable.

WHAT WAS WRONG, measured on a generated workbook rather than read off the source:

* The banner was ONE cell at `A3` — not merged, not wrapped, no fill, 11pt in a muted terracotta —
  on a sheet with **no freeze panes**. It left the screen on the first wheel-turn, so the director
  read a grand total with nothing in view.
* It sat on **sheet 6 of 8**, and `00 README` — the first sheet anybody opens — said nothing about
  placeholders at all.
* The per-line marking was written into **column M**, which begins past ~1,380 px of columns B..L.
  So the only place a placeholder line was marked started beyond the right edge of a normal window,
  leaving B..L byte-identical to a genuinely priced row. The comment above that line claimed
  "somebody scrolling this sheet must be able to see which numbers nobody chose without
  cross-referencing anything" — the layout defeated it.
* `line.note` already opens `"PROVISIONAL — "` (`model.PLACEHOLDER_NOTE`, applied in
  `costing.py`), and the writer prefixed it again: `PROVISIONAL — PROVISIONAL — …`.
* `priced.unpriced` and `priced.problems` are composed carefully and reached no sheet at all, so a
  line with NO RATE added nothing to column L and was counted nowhere. The total looked complete
  because the shortfall was invisible, not because it was zero.

The web UI was markedly louder than the deliverable — and the deliverable is the thing that leaves
the building.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from client_boq.boq.costing import PricedBQ, PricedRow
from client_boq.boq.model import PLACEHOLDER_NOTE


class _NoRefs:
    def get(self, _key):
        return ""

    def formula(self, _key):
        return "1"


def _row(ref: str, *, source="basis", note="", cost_basis=100.0) -> PricedRow:
    return PricedRow(full_ref=ref, description=f"item {ref}", qty=10.0, unit="m",
                     cost_basis=cost_basis, rate_to_submit=120.0, amount=1200.0,
                     source=source, note=note)


def _provisional() -> PricedBQ:
    priced = PricedBQ(
        rows=[_row("1.1"),
              _row("1.2", source="placeholder", note=PLACEHOLDER_NOTE),
              _row("1.3", source="placeholder", note=PLACEHOLDER_NOTE)],
        total=3600.0)
    priced.placeholders = ["1.2", "1.3"]
    return priced


def _sheet(priced: PricedBQ):
    from client_boq.boq import costing_workbook
    from openpyxl import Workbook

    book = Workbook()
    ws = book.active
    costing_workbook._bq_priced(ws, priced, where=_NoRefs())
    stream = io.BytesIO()
    book.save(stream)
    stream.seek(0)
    return load_workbook(stream).active


def _readme(priced: PricedBQ):
    from client_boq.boq import costing_workbook
    from openpyxl import Workbook

    book = Workbook()
    ws = book.active
    costing_workbook._readme(ws, priced)
    stream = io.BytesIO()
    book.save(stream)
    stream.seek(0)
    return load_workbook(stream).active


class TestTheBannerCannotScrollAway:

    def test_the_rows_above_the_table_are_frozen(self):
        """The single most important line in this file. Without it the warning is gone the moment
        the reader looks at row 20 of a 166-line bill."""
        ws = _sheet(_provisional())
        assert ws.freeze_panes == "A5"

    def test_the_banner_is_filled_merged_and_wrapped(self):
        ws = _sheet(_provisional())
        assert ws["A3"].fill.fgColor.rgb == "FFF8D7DA"
        assert "A3:M3" in {str(r) for r in ws.merged_cells.ranges}
        assert ws["A3"].alignment.wrap_text is True
        assert ws.row_dimensions[3].height == 30

    def test_it_is_bigger_and_redder_than_the_body_text(self):
        ws = _sheet(_provisional())
        assert ws["A3"].font.bold and ws["A3"].font.size == 12
        assert ws["A3"].font.color.rgb == "FF9C0006"

    def test_the_sheet_tab_turns_red(self):
        ws = _sheet(_provisional())
        assert (ws.sheet_properties.tabColor.rgb or "").endswith("C25539")

    def test_the_words_that_matter_are_still_the_words(self):
        ws = _sheet(_provisional())
        assert "DO NOT SUBMIT" in ws["A3"].value
        assert "chosen by nobody" in ws["A3"].value
        assert "2 lines stand on a placeholder" in ws["A3"].value

    def test_a_clean_bill_has_no_banner_and_no_red_tab(self):
        """Loud on a provisional bill; silent on a settled one. A warning that is always there is
        not a warning."""
        ws = _sheet(PricedBQ(rows=[_row("1.1")], total=1200.0))
        assert ws["A3"].value is None
        assert ws.sheet_properties.tabColor is None
        assert ws.freeze_panes == "A5", "the header freeze is unconditional"


class TestAPlaceholderLineIsVisibleWithoutScrollingRight:

    def test_the_whole_row_is_tinted(self):
        """Column M begins past ~1,380 px of B..L, so a note there is off-screen. The row tint is
        the one marking that cannot be."""
        ws = _sheet(_provisional())
        placeholder_rows = [r for r in range(5, 9) if ws.cell(row=r, column=2).value in ("1.2", "1.3")]
        assert len(placeholder_rows) == 2
        for r in placeholder_rows:
            for column in range(2, 14):
                assert ws.cell(row=r, column=column).fill.fgColor.rgb == "FFFDEDEE", \
                    f"column {column} of row {r} is not tinted"

    def test_a_priced_row_is_not_tinted(self):
        ws = _sheet(_provisional())
        clean = next(r for r in range(5, 9) if ws.cell(row=r, column=2).value == "1.1")
        assert ws.cell(row=clean, column=3).fill.fgColor.rgb != "FFFDEDEE"

    def test_the_prefix_is_not_doubled(self):
        ws = _sheet(_provisional())
        marked = next(r for r in range(5, 9) if ws.cell(row=r, column=2).value == "1.2")
        note = ws.cell(row=marked, column=13).value or ""
        assert note.count("PROVISIONAL —") == 1, note

    def test_the_warning_column_finally_has_a_header(self):
        """It was the only unlabelled column on the sheet, and the one carrying every warning."""
        ws = _sheet(_provisional())
        assert ws.cell(row=4, column=13).value == "Notes and warnings"


class TestTheFirstSheetAnybodyOpensSaysIt:

    def test_the_readme_carries_the_alarm(self):
        ws = _readme(_provisional())
        assert "DO NOT SUBMIT" in (ws["A3"].value or "")
        assert "05 BQ Priced" in ws["A3"].value
        assert ws["A3"].fill.fgColor.rgb == "FFF8D7DA"
        assert (ws.sheet_properties.tabColor.rgb or "").endswith("C25539")

    def test_it_gives_both_numbers_so_the_reader_can_see_the_shortfall(self):
        ws = _readme(_provisional())
        assert "Actually priced" in ws["A3"].value

    def test_a_clean_bill_leaves_the_readme_alone(self):
        ws = _readme(PricedBQ(rows=[_row("1.1")], total=1200.0))
        assert ws["A3"].value is None
        assert ws.sheet_properties.tabColor is None

    def test_a_readme_built_without_a_bill_still_works(self):
        """`_readme` is called before the priced bill exists in some paths; it must not require it."""
        from client_boq.boq import costing_workbook
        from openpyxl import Workbook

        ws = Workbook().active
        costing_workbook._readme(ws)
        assert ws["A1"].value == "GI Tender Costing"


class TestWhatIsNotSettledReachesTheSheet:

    def test_an_unpriced_line_is_counted_and_named(self):
        priced = PricedBQ(rows=[_row("1.1"), _row("3.9", cost_basis=None)], total=1200.0)
        priced.unpriced = ["3.9"]
        ws = _sheet(priced)
        text = " ".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
        assert "WHAT IS NOT SETTLED" in text
        assert "1 line(s) carry no rate at all" in text
        assert "3.9" in text

    def test_a_problem_string_is_printed(self):
        priced = PricedBQ(rows=[_row("1.1")], total=1200.0)
        priced.problems = ["2 items were priced from a basis nothing divides."]
        ws = _sheet(priced)
        text = " ".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
        assert "basis nothing divides" in text

    def test_a_settled_bill_gets_no_block(self):
        ws = _sheet(PricedBQ(rows=[_row("1.1")], total=1200.0))
        text = " ".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
        assert "WHAT IS NOT SETTLED" not in text


class TestItPrintsLikeADeliverable:

    def test_the_header_repeats_on_every_printed_page(self):
        ws = _sheet(_provisional())
        # openpyxl normalises the range to absolute form on the round trip.
        assert ws.print_title_rows == "$1:$4"

    def test_it_fits_the_page_width_rather_than_spilling_across_sheets(self):
        ws = _sheet(_provisional())
        assert ws.page_setup.orientation == "landscape"
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True


@pytest.mark.parametrize("count", [1, 5, 99])
def test_the_count_is_the_real_one(count):
    """99 lines and 8,890,660 was the observed figure. The banner must never round it away."""
    rows = [_row(f"1.{n}", source="placeholder", note=PLACEHOLDER_NOTE) for n in range(count)]
    priced = PricedBQ(rows=rows, total=1200.0 * count)
    priced.placeholders = [r.full_ref for r in rows]
    ws = _sheet(priced)
    assert f"{count} lines stand on a placeholder" in ws["A3"].value
