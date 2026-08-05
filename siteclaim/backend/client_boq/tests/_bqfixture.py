"""Build a bill-of-quantities workbook that reproduces every trap measured in the real one.

The reference package (CEDD ND/2025/04) is a real, digitally-signed government tender and does not
belong in this repository, so the tests run against a workbook generated here instead. It is small —
five bills, twenty-odd items — but it reproduces each specific thing that broke, or would have
broken, a naive reader:

===========================================================================================
trap                                              where it is in this fixture
===========================================================================================
item ref stored as a float, colliding under        Bill No.1 items 1.2 / 1.20  (both hold 1.2)
different number formats                           Bill No.2 items 2.1 / 2.10  (both hold 2.1)
                                                   Bill No.9 items 9.1 / 9.10  (both hold 9.1)
a ref whose stored value is wrong and is           Bill No.2 item "2.24" (holds 2.244)
masked by its format
a ref stored as a string, with a letter suffix     Bill No.1 item "1.61A" (rev 2 only)
so nothing downstream is renumbered
lettered variants stored as a bare "a"/"b"         Bill No.2 items 2.2a / 2.2b (rev 1+)
under a parent that keeps its own number
lump items: "-" in the quantity AND rate cells     Bill No.1 items 1.5, 1.6, 1.62, 1.63
a genuine zero quantity, which is NOT a lump       Bill No.2 item 2.5
description spread across three indent columns     every sheet
hard-wrapped continuation rows whose leading       Bill No.1 item 1.2, Bill No.4 item 4.18
and trailing spaces are load-bearing
a quantity stranded on the caption row above       Bill No.1 item 1.16 (rows 15/16)
its item
unit spelling variants of one unit                 "item", "item ", "Item"; "nr", "nr."
page references that exist only as page-break      every bill sheet
geometry
~76,000 stray cells past column H from an old      Bill No.4
fill-right, with the wrong bill's name in them
pre-priced items the tenderer may not alter        all of Bill No.9
client-inserted contingency and provisional sums   Grand Summary (B), (D), (E)
===========================================================================================

and across the three revisions, each of the real addenda's moves:

    rev 0 → 1   item 2.2 (91 nr) split into 2.2a (80) + 2.2b (11), parent kept as a caption
                Bill No.6 quantities 6.4/6.5/6.6 multiplied by ~2.17 — the change that was
                announced as "Updated the quantities of item nos. 6.4 - 6.6"
                every item below the split moves down a row and KEEPS ITS REFERENCE

    rev 1 → 2   a caption edited from "marine" to "land" traffic, with the three items beneath
                it untouched — a scope change an item-row diff cannot see
                a new item 1.61A inserted with a letter suffix, so 1.62 and 1.63 move rows
                without being renumbered
                item 4.18's description narrowed from "soil and ground water" to "soil" at an
                unchanged quantity

Deliberately faithful in one more respect: **nothing here marks what changed.** No fill, no bold on
a changed value, no comment, no defined name — because the real workbook has none either. Rev 1
carries a sheet-level print footer on the bills it touched (as the real one does); rev 2 carries
nothing at all (as the real one does).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

# The stray-cell block on Bill No.4. The real sheet carries roughly 76,500 of them out to column
# XFD; a few hundred is enough to prove the clamp works without making every test slow.
STRAY_COLUMNS = range(200, 320)
STRAY_ROWS = (5, 6, 7)


def _put(sheet: Any, row: int, col: int, value: Any, *, fmt: Optional[str] = None,
         caption: bool = False) -> None:
    from openpyxl.styles import Font

    cell = sheet.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    # Captions are bold + underlined in the reference workbook. It is a signal, not the rule — the
    # reader falls back on column depth — but it is what the real file does, so the fixture does it.
    cell.font = Font(name="Times New Roman", size=11, bold=caption, underline="single" if caption else None)


def _header(sheet: Any, row: int, bill_no: str, title: str) -> None:
    _put(sheet, row, 1, f"'Bill No. {bill_no}'", caption=True)
    _put(sheet, row + 1, 1, "Item No.", caption=True)
    _put(sheet, row + 1, 2, "Item Description", caption=True)
    _put(sheet, row + 1, 5, "Quantity", caption=True)
    _put(sheet, row + 1, 6, "Unit", caption=True)
    _put(sheet, row + 1, 7, " Rate", caption=True)
    _put(sheet, row + 1, 8, "Amount", caption=True)
    _put(sheet, row + 2, 7, "(HK$)", caption=True)
    _put(sheet, row + 2, 8, "(HK$)", caption=True)
    sheet.cell(row=row, column=2, value=f"' {title}'")


def _breaks(sheet: Any, *rows: int) -> None:
    from openpyxl.worksheet.pagebreak import Break

    for row in rows:
        sheet.row_breaks.append(Break(id=row))


def _item(sheet: Any, row: int, ref: Any, col: int, text: str, qty: Any, unit: str, *,
          fmt: Optional[str] = None, rate: Any = None, amount: Any = None) -> None:
    _put(sheet, row, 1, ref, fmt=fmt)
    _put(sheet, row, col, text)
    if qty is not None:
        _put(sheet, row, 5, qty)
    if unit:
        _put(sheet, row, 6, unit)
    if rate is not None:
        _put(sheet, row, 7, rate)
    if amount is not None:
        _put(sheet, row, 8, amount)


def _bill_1(workbook: Any, rev: int) -> None:
    """General and Preliminaries. Carries the ref collision, the stranded quantity, the caption
    that rev 2 rewrites, and the item rev 2 inserts."""
    sheet = workbook.create_sheet("Bill No.1")
    _header(sheet, 1, "1", "General and Preliminaries")
    _put(sheet, 5, 2, "SECTION 1 - PRELIMINARIES", caption=True)
    _put(sheet, 7, 2, "Maintenance of Traffic Flow", caption=True)
    # THE caption rev 2 edits. Items 1.5-1.7 beneath it are never touched, yet their scope moves
    # from marine to land traffic, because General Preambles 2 makes a sub-heading part of the item.
    _put(sheet, 9, 3, "Maintain land traffic flow" if rev >= 2 else "Maintain marine traffic flow",
         caption=True)
    _item(sheet, 10, 1.5, 4, "Provision of measures", "-", "item", rate="-")
    _item(sheet, 12, 1.6, 4, "Maintenance of measures", "-", "item ", rate="-")

    _put(sheet, 14, 2, "Photographs", caption=True)
    # The stranded quantity: the caption carries it, the item row below does not.
    _put(sheet, 15, 3, "Record photographs ", caption=True)
    _put(sheet, 15, 5, 1)
    _put(sheet, 15, 6, "nr.")
    _item(sheet, 16, 1.16, 4, "8R size print", None, "")

    _put(sheet, 18, 2, "Transport for the Use of the Project Manager", caption=True)
    _put(sheet, 19, 3, "Land transport for the use of the Project Manager ", caption=True)
    # 1.2 under General. Below, the SAME stored value under '0.00' is item 1.20 — a different item.
    _item(sheet, 20, 1.2, 4, "air-conditioned environmentally-friendly petrol ", 1, "nr.",
          fmt="General")
    _put(sheet, 21, 4, "private car vehicle with seating capacity of not less")
    _put(sheet, 22, 4, " than 7 seats excluding driver")
    _item(sheet, 24, 1.2, 4, "operation and maintenance of the above", 122, "nr-wk", fmt="0.00")

    row = 26
    if rev >= 2:
        _put(sheet, 26, 2, "Hoardings, Temporary Fences and Signboards", caption=True)
        _put(sheet, 27, 3, "Signboard", caption=True)
        _put(sheet, 28, 4, "minor project signboard as CEDD Drawing")
        _put(sheet, 29, 4, "Nos. C1003/1I and C1003/2B")
        # A string, not a float, and suffixed so 1.62 and 1.63 below need no renumbering.
        _item(sheet, 30, "1.61A", 4, "size A", 1, "nr")
        row = 32

    _put(sheet, row, 2, "SECTION 3 - SITE CLEARANCE", caption=True)
    _item(sheet, row + 1, 1.62, 3, "General site clearance of the Site", "-", "item", rate="-")
    _item(sheet, row + 3, 1.63, 3, "Vegetation Survey in Conservation Area", "-", "Item", rate="-")
    _breaks(sheet, 25)


def _bill_2(workbook: Any, rev: int) -> None:
    """Ground Investigation Fieldworks. Carries the a/b split, a second ref collision, the item
    whose stored reference is wrong, the three-deep caption chain, and a real zero quantity."""
    sheet = workbook.create_sheet("Bill No.2")
    _header(sheet, 1, "2", "Ground Investigation Fieldworks")
    _put(sheet, 5, 2, "SECTION 2 - GROUND INVESTIGATION", caption=True)
    _put(sheet, 7, 2, "Rigs", caption=True)
    _item(sheet, 9, 2.1, 3, "Establishment of rigs", "-", "item", fmt="General", rate="-")

    if rev >= 1:
        # The parent keeps its number and becomes a caption; the variants are bare letters. The
        # addendum calls them "item nos. 2.2a and 2.2b" — a form the spreadsheet never stores.
        _item(sheet, 11, 2.2, 3, "Moving rigs", None, "")
        _item(sheet, 12, "a", 4, "in Class A of site", 80, "nr")
        _item(sheet, 14, "b", 4, "in Class B of site", 11, "nr")
        shift = 5
    else:
        _item(sheet, 11, 2.2, 3, "Moving rigs", 91, "nr")
        shift = 0

    # Everything below the split moves down, and keeps its reference. That is the whole point.
    _item(sheet, 13 + shift, 2.3, 3, "Standing time for rigs", 455, "h")
    _put(sheet, 16 + shift, 2, "Drilling, Boring and Probing", caption=True)
    _put(sheet, 18 + shift, 2, "Drilling", caption=True)
    _put(sheet, 20 + shift, 3, "Drilling size H or N", caption=True)
    _item(sheet, 21 + shift, 2.4, 4, "vertically downwards, material other than rock, ", 2300, "m")
    _put(sheet, 22 + shift, 4, "boulder or artificial hard material ")
    # A genuine 0, not a lump: the tenderer still has to rate it in case one is ordered.
    _item(sheet, 24 + shift, 2.5, 4, " vertically downwards, rock", 0, "m")

    _put(sheet, 26 + shift, 2, " Trial Pits and Inspection Pits", caption=True)
    _put(sheet, 28 + shift, 2, " Extra over for excavation in rock", caption=True)
    # Its own cell says only "maximum depth..."; the meaning is the chain above it.
    _item(sheet, 29 + shift, 2.9, 3, " maximum depth not exceeding 3.00m", 17, "m3")
    _item(sheet, 31 + shift, 2.1, 3, " Accoustic Televiewer Test", 8, "nr", fmt="0.00")
    # Stored 2.244, displayed "2.24". The displayed form is the identity everything else cites.
    _item(sheet, 33 + shift, 2.244, 3, " Pressuremeter Test", 31, "nr", fmt="0.00")
    _breaks(sheet, 25 + shift)
    if rev >= 1:
        sheet.oddFooter.center.text = "Tender Addendum No. 1"


def _bill_4(workbook: Any, rev: int) -> None:
    """Laboratory Testing — the structurally damaged sheet, and the description rev 2 narrows."""
    sheet = workbook.create_sheet("Bill No.4")
    _header(sheet, 1, "4", "Laboratory Testing for Ground Investigation Fieldworks")
    _put(sheet, 5, 2, "SECTION 2 - GROUND INVESTIGATION, LABORATORY TESTING", caption=True)
    if rev >= 2:
        # Two rows collapsed into one, dropping " and ground water". The quantity never moves, so
        # the rate now covers half the medium it did before.
        _item(sheet, 7, 4.18, 3, "Determination of carbonate content test for soil", 16, "nr")
    else:
        _item(sheet, 7, 4.18, 3, "Determination of carbonate content test for", 16, "nr")
        _put(sheet, 8, 3, "soil and ground water ")
    _item(sheet, 10, 4.19, 3, "Determination of resistivity test for soil", 16, "nr")

    # The damage: a page-header block dragged right across the sheet years ago, still naming the
    # wrong bill. It prints fine because the print area is clamped to column H.
    for row in STRAY_ROWS:
        for col in STRAY_COLUMNS:
            sheet.cell(row=row, column=col,
                       value="'Bill No. 1'" if col % 8 == 0 else " General and Preliminaries")
    sheet.print_area = "A1:H60"
    _breaks(sheet, 54)
    if rev >= 2:
        sheet.oddFooter.center.text = "Tender Addendum No. 2"


def _bill_6(workbook: Any, rev: int) -> None:
    """Groundwater Monitoring — the quantities the first addendum doubled without saying so."""
    sheet = workbook.create_sheet("Bill No.6")
    _header(sheet, 1, "6", "Groundwater Monitoring Works")
    _put(sheet, 5, 2, "SECTION 2 - GROUND INVESTIGATION", caption=True)
    _put(sheet, 7, 2, "Recording and reporting of instrument", caption=True)
    quantities = (2451, 3546, 5996) if rev >= 1 else (1128, 1623, 2760)
    _item(sheet, 8, 6.4, 3, "Standpipe", quantities[0], "nr-wk")
    _item(sheet, 10, 6.5, 3, "Piezometer", quantities[1], "nr-wk")
    _item(sheet, 12, 6.6, 3, "Automatic groundwater monitoring device ", quantities[2], "nr-wk")
    _breaks(sheet, 54)
    if rev >= 1:
        sheet.oddFooter.center.text = "Tender Addendum No. 1"


def _bill_9(workbook: Any, rev: int) -> None:
    """Site Safety Management — priced in full by the client under the Pay for Safety Scheme, so
    nobody can win by shaving safety money. GCT App C 2.2(vi) reinstates these if altered."""
    sheet = workbook.create_sheet("Bill No.9")
    _header(sheet, 1, "9", "Site Safety Management")
    _put(sheet, 4, 2, "SECTION 28 - SITE SAFETY MANAGEMENT", caption=True)
    _put(sheet, 6, 2, "PROVIDE SAFETY OFFICER", caption=True)
    _item(sheet, 8, 9.1, 3, "Provide Safety Officer", 16, "nr-mth", fmt="General",
          rate=4860, amount=77760)
    _put(sheet, 10, 2, "SAFETY PROMOTION", caption=True)
    _item(sheet, 12, 9.9, 3, "Participate in Safety promotional campaign", "-", "sum",
          rate=10800, amount=10800)
    _item(sheet, 14, 9.1, 3, "Arrange and hold Pre-work Activities of Site", 20, "mth", fmt="0.00",
          rate=8650, amount=173000)
    _put(sheet, 15, 3, "Safety Cycle")
    _breaks(sheet, 54)


def _grand_summary(workbook: Any) -> None:
    sheet = workbook.create_sheet("Grand Summary")
    _put(sheet, 11, 1, "Page \nNo.", caption=True)
    _put(sheet, 11, 2, "Bill \nNo.", caption=True)
    _put(sheet, 11, 3, "Description", caption=True)
    _put(sheet, 11, 4, "Amount\nHK$", caption=True)
    for offset, (bill, title) in enumerate((
        ("1", "General and Preliminaries"),
        ("2", "Ground Investigation Fieldworks"),
        ("4", "Laboratory Testing for Ground Investigation Fieldworks"),
        ("6", "Groundwater Monitoring Works"),
        ("9", "Site Safety Management"),
    )):
        _put(sheet, 13 + offset, 1, f"BQ/{bill}/2")
        _put(sheet, 13 + offset, 2, int(bill))
        _put(sheet, 13 + offset, 3, title)
    _put(sheet, 13 + 4, 4, 429810)                       # Bill 9 arrives priced
    _put(sheet, 23, 3, "Sub-total above (A)")
    _put(sheet, 24, 3, "Tendered total of the Prices = (A)")
    _put(sheet, 25, 3, "Contingency sum for Defined Cost for compensation events* (B)")
    _put(sheet, 25, 4, 4342620)
    _put(sheet, 26, 3, "Contingency sum for Fee for compensation events* (C) = (B) x direct fee "
                       "percentage inserted by the tenderer in the Contract Data Part two")
    _put(sheet, 27, 3, "Provisional sum for price adjustment for inflation under Secondary "
                       "Option X1* (D)")
    _put(sheet, 27, 4, 1550000)
    _put(sheet, 28, 3, "Provisional sum for performance-tied payment item under Pay for Safety "
                       "Performance Merit Scheme (PFSPMS) under Secondary Option X20* (E)")
    _put(sheet, 28, 4, 609370)
    _put(sheet, 29, 3, "Sub-total of all contingency sums and provisional sums above * "
                       "(F) = (B) + (C) + (D) + (E) ")
    _put(sheet, 30, 3, "Forecast total of the Prices for tender evaluation purpose * (G) = (A) + (F) ")
    _put(sheet, 32, 1, "*Remarks: (1) The contingency sums, provisional sums and forecast total of "
                       "the Prices shall not form part of this contract.")


def build_bill_workbook(path: str | Path, rev: int = 0) -> Path:
    """Write a bill-of-quantities workbook at revision ``rev`` (0, 1 or 2). Returns the path."""
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    index = workbook.create_sheet("Index")
    _put(index, 16, 1, "Bill No.", caption=True)
    _put(index, 16, 3, "Description", caption=True)
    _put(index, 16, 6, "Page No.", caption=True)
    _grand_summary(workbook)
    _bill_1(workbook, rev)
    _bill_2(workbook, rev)
    _bill_4(workbook, rev)
    _bill_6(workbook, rev)
    _bill_9(workbook, rev)

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    workbook.close()
    return target
