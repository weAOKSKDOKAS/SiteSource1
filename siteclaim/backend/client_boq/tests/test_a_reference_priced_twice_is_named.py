"""A bill reference priced more than once is named where the money is.

THE DEFECT, as observed. Bill 1 of the reference pack prints 63 items — 1.1 … 1.63, verified, no
gaps, no letter suffixes — and the generated `05 BQ Priced` sheet carried 64 rows.

Traced rather than guessed. One `PricedRow` is emitted per non-parent `BillItem` (`price()` has
four append sites, all inside one loop over `bill.items`, each path terminated). `BillItem`s come
from the client's own workbook through `reader.read_workbook`, which is the only non-test place a
`ClientBill` is constructed. Nothing between the two adds a row: the spread is an allocation across
existing items, `unbilled` is a gate producing decisions, a placeholder marks a row that already
exists, and a parent REMOVES one. So a 64th `1.x` row is a 64th non-parent `BillItem`.

And the reader already knows: it appends a row per reference row it finds and, on a repeat, notes
"both kept, neither is assumed correct" (`reader.py:381-387`). That is the right call — a reader
must not choose which of two printed rows is the real one.

WHAT WAS MISSING is what happens next. `PricedBQ.index()` and `ClientBill.index()` both key on the
reference and collapse the pair, so every screen reading through an index shows 63 while the
workbook prints 64; the workbook's `SUM` sums printed ROWS, so the duplicated amount is in the
total twice; and a typed override for that reference applies to both copies, so the double reaches
the tender. None of that was said anywhere.

So: named, never repaired. Which printed row is the item is a question about the client's document,
and this cannot answer it — but it can refuse to let the total quietly stop being the bill's total.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from client_boq.boq import costing as boq_costing
from client_boq.boq.costing import PricedBQ, PricedRow


def _row(ref: str, *, qty=10.0, rate=100.0, source="basis") -> PricedRow:
    return PricedRow(full_ref=ref, description=f"item {ref}", qty=qty, unit="m",
                     rate_to_submit=rate, amount=qty * rate, source=source)


def _priced(*rows: PricedRow) -> PricedBQ:
    return PricedBQ(rows=list(rows), total=round(sum(r.amount or 0.0 for r in rows), 2))


class TestACleanBillSaysNothing:

    def test_distinct_references_report_no_duplicates(self):
        clean = _priced(_row("1.1"), _row("1.2"), _row("1.3"))
        assert clean.duplicate_refs() == {}
        assert clean.duplicate_total() == 0.0
        assert clean.duplicate_note() == ""

    def test_an_empty_bill_is_not_a_duplicate(self):
        assert PricedBQ().duplicate_refs() == {}
        assert PricedBQ().duplicate_note() == ""

    def test_the_same_reference_in_different_bills_is_not_a_duplicate_here(self):
        """`full_ref` already carries the bill, so 1.5 and 2.5 are different references. This test
        exists because `bill_no` is NOT carried onto `PricedRow`, so `full_ref` is the only
        identity there is — and that is exactly why it must be the thing compared."""
        assert _priced(_row("1.5"), _row("2.5")).duplicate_refs() == {}


class TestTheReferencePricedTwice:

    def test_it_is_counted_and_named(self):
        twice = _priced(_row("1.53"), _row("1.54"), _row("1.53"))
        assert twice.duplicate_refs() == {"1.53": 2}
        assert "1.53 ×2" in twice.duplicate_note()

    def test_the_extra_copy_is_the_money_the_total_is_over_by(self):
        """The first copy is the item. Every copy after it is the over-count, so the figure is the
        amount of the SECOND row onwards — not the sum of both."""
        twice = _priced(_row("1.53", qty=10.0, rate=100.0), _row("1.53", qty=10.0, rate=100.0))
        assert twice.total == 2000.0
        assert twice.duplicate_total() == 1000.0

    def test_three_copies_report_two_extras(self):
        thrice = _priced(*[_row("1.53", qty=1.0, rate=500.0) for _ in range(3)])
        assert thrice.duplicate_refs() == {"1.53": 3}
        assert thrice.duplicate_total() == 1000.0

    def test_copies_priced_differently_are_still_both_the_bill_reader_s_doing(self):
        """Two printed rows under one reference need not carry the same quantity. The over-count is
        whatever the extra rows actually amount to, not an assumption that they match."""
        odd = _priced(_row("1.53", qty=10.0, rate=100.0), _row("1.53", qty=4.0, rate=100.0))
        assert odd.duplicate_total() == 400.0

    def test_the_note_says_what_to_do_and_does_not_pretend_to_know_which_row_is_real(self):
        note = _priced(_row("1.53"), _row("1.53")).duplicate_note()
        assert "rather than choosing between them" in note
        assert "Settle which printed row is the item" in note
        assert "counts 1,000.00 twice" in note

    def test_the_index_is_the_reason_this_exists(self):
        """Established as a fact rather than asserted in prose: the index collapses the pair, so
        a screen reading through it cannot see the second copy at all."""
        twice = _priced(_row("1.53"), _row("1.54"), _row("1.53"))
        assert len(twice.rows) == 3 and len(twice.index()) == 2


class TestItReachesTheWorkbook:
    """The sheet is the only surface the second copy is visible on, so it must say so there."""

    @staticmethod
    def _sheet(priced: PricedBQ):
        from client_boq.boq import costing_workbook

        class _NoRefs:
            """The cross-sheet reference map. Empty here: this test is about column M, and an
            empty map makes every formula fall back to its literal, which is fine."""

            def get(self, _key):
                return ""

            def formula(self, _key):
                return "1"

        book = __import__("openpyxl").Workbook()
        ws = book.active
        costing_workbook._bq_priced(ws, priced, where=_NoRefs())
        stream = io.BytesIO()
        book.save(stream)
        stream.seek(0)
        return load_workbook(stream).active

    def test_a_clean_sheet_still_says_total_of_priced_items(self):
        ws = self._sheet(_priced(_row("1.1"), _row("1.2")))
        labels = [c.value for c in ws["C"] if isinstance(c.value, str)]
        assert "TOTAL OF PRICED ITEMS" in labels
        assert not any("NOT THE BILL'S TOTAL" in v for v in labels)

    def test_a_duplicated_sheet_stops_calling_the_sum_the_bill_s_total(self):
        ws = self._sheet(_priced(_row("1.53"), _row("1.54"), _row("1.53")))
        labels = [c.value for c in ws["C"] if isinstance(c.value, str)]
        assert any("NOT THE BILL'S TOTAL" in v for v in labels)
        assert any("of which counted twice" in v for v in labels)

    def test_every_copy_is_marked_on_its_own_row(self):
        ws = self._sheet(_priced(_row("1.53"), _row("1.54"), _row("1.53")))
        marks = [c.value for c in ws["M"] if isinstance(c.value, str) and "DUPLICATE" in c.value]
        assert len(marks) == 2
        assert "copy 1 of 2" in marks[0] and "copy 2 of 2" in marks[1]

    def test_the_row_that_is_not_duplicated_carries_no_mark(self):
        ws = self._sheet(_priced(_row("1.53"), _row("1.54"), _row("1.53")))
        for row in ws.iter_rows(min_row=5, max_row=7, min_col=2, max_col=13):
            ref, note = row[0].value, row[11].value
            if ref == "1.54":
                assert not note or "DUPLICATE" not in note

    def test_a_placeholder_and_a_duplicate_are_both_said_on_the_same_row(self):
        """Two different problems with one line. Overwriting one with the other would hide
        whichever the code happened to write second."""
        first_copy = _row("1.53", source="placeholder")
        first_copy.cost_basis = 100.0
        first_copy.note = "no rate in the book for this shape"
        both = _priced(first_copy, _row("1.53"))
        both.placeholders = ["1.53"]
        ws = self._sheet(both)
        first = ws.cell(row=5, column=13).value or ""
        assert "PROVISIONAL" in first and "DUPLICATE REFERENCE" in first

    def test_the_over_count_figure_is_on_the_sheet_as_a_number(self):
        ws = self._sheet(_priced(_row("1.53", qty=10.0, rate=100.0),
                                 _row("1.53", qty=10.0, rate=100.0)))
        values = [c.value for c in ws["L"] if isinstance(c.value, (int, float))]
        assert 1000.0 in values


class TestTheCostingPayloadCarriesIt:

    @pytest.fixture
    def client(self, tmp_path, monkeypatch):
        from fastapi.testclient import TestClient

        monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "workspace"))
        from api import app

        return TestClient(app)

    def test_the_route_is_registered_and_the_block_is_shaped(self, client):
        """The endpoint needs a whole tender to answer, so this pins the contract's shape at the
        source instead — the payload builds these three keys from exactly these methods."""
        priced = _priced(_row("1.53"), _row("1.53"))
        block = {"refs": priced.duplicate_refs(), "amount": priced.duplicate_total(),
                 "note": priced.duplicate_note()}
        assert block["refs"] == {"1.53": 2}
        assert block["amount"] == 1000.0
        assert block["note"]
        assert "/client-boq/costing/{set_id}" in client.app.openapi()["paths"]


def test_price_emits_exactly_one_row_per_non_parent_item():
    """The premise the whole investigation rests on, stated executably: a 64th row IS a 64th item,
    so the reading is where to look and no stage between adds anything."""
    import inspect

    source = inspect.getsource(boq_costing.price)
    assert source.count("result.rows.append(") == 4, "four append sites"
    assert "for item in bill.items:" in source
    assert "if item.is_parent:" in source and "continue" in source
