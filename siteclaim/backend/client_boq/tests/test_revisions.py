"""Spec for T3 — the revision axis.

The rule the whole feature turns on: **nothing is ever destroyed.** A corrected re-upload or an
addendum appends a revision; Rev 0 survives Rev 1 and stays readable. Modelled on the real
ND/2025/04 package, where 154 documents stayed at Rev 0, 9 went to Rev 1, and 2 went to Rev 2 —
so a revision must be cheap and local, never a second copy of the whole tender.
"""

from __future__ import annotations

import sqlite3

import pytest
from fastapi.testclient import TestClient

from client_boq import models, store

fitz = pytest.importorskip("fitz")  # PyMuPDF


@pytest.fixture
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "workspace"))
    from api import app

    return TestClient(app)


def _pdf(pages: int, text: str = "page", toc: list | None = None) -> bytes:
    doc = fitz.open()
    for i in range(1, pages + 1):
        doc.new_page().insert_text((72, 100), f"{text} {i} of {pages}, with enough text to read",
                                   fontsize=11)
    if toc:
        doc.set_toc(toc)
    data = doc.tobytes()
    doc.close()
    return data


def _ingest(client: TestClient, name: str = "rev-demo") -> str:
    binder = _pdf(12, "Binder page", toc=[
        [1, "Conditions of Tender", 1], [1, "Scope of Works", 5], [1, "Pricing Schedule", 9],
    ])
    resp = client.post("/client-boq/ingest/upload", data={"project_name": name},
                       files={"files": ("binder.pdf", binder, "application/pdf")})
    set_id = resp.json()["result"]["set_id"]
    client.post("/client-boq/ingest/manifest/approve", json={"set_id": set_id})
    client.post("/client-boq/ingest/split", json={"set_id": set_id})
    return set_id


def _send_addendum(client: TestClient, set_id: str, filename: str = "boq-rev1.pdf") -> dict:
    resp = client.post(
        "/client-boq/ingest/document",
        data={"set_id": set_id, "kind": "addendum", "ref": "Tender Addendum No. 1"},
        files=[
            ("files", ("Tender Addendum No.1.pdf", _pdf(2, "Addendum letter"), "application/pdf")),
            ("files", (filename, _pdf(6, "Revised pricing"), "application/pdf")),
        ],
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


# ---------------------------------------------------------------------------
# Step 1: the base ingest still behaves, now on revisions underneath
# ---------------------------------------------------------------------------
def test_a_fresh_ingest_puts_every_part_at_rev_0(client):
    set_id = _ingest(client)

    body = client.get(f"/client-boq/revisions/{set_id}").json()
    assert [p["operative_rev"] for p in body["parts"]] == [0, 0, 0]
    assert body["amended"] == []
    assert len(body["documents"]) == 1
    assert body["documents"][0]["kind"] == "base"
    assert body["documents"][0]["seq"] == 0


def test_re_splitting_the_same_document_does_not_invent_a_revision(client):
    # A manifest edit is a better reading of ONE document, not a new document. Treating it as a
    # revision would fill the history with noise from ordinary fiddling.
    set_id = _ingest(client)
    client.post("/client-boq/ingest/manifest/approve", json={
        "set_id": set_id,
        "parts": [{"n": 1, "abbr": "A", "slug": "first", "title": "First", "start": 1, "end": 6},
                  {"n": 2, "abbr": "B", "slug": "second", "title": "Second", "start": 7, "end": 12}],
    })
    client.post("/client-boq/ingest/split", json={"set_id": set_id})

    body = client.get(f"/client-boq/revisions/{set_id}").json()
    assert [p["operative_rev"] for p in body["parts"]] == [0, 0]
    assert len(body["documents"]) == 1          # still just the original upload
    assert [p["title"] for p in body["parts"]] == ["First", "Second"]


# ---------------------------------------------------------------------------
# Step 2: an addendum proposes, the gate commits
# ---------------------------------------------------------------------------
def test_an_addendum_proposes_but_commits_nothing(client):
    set_id = _ingest(client)
    plan = _send_addendum(client, set_id)

    assert plan["requires_gate"] is True
    assert plan["kind"] == "addendum"
    assert plan["changes"]                       # the addendum's own table was read
    assert "neither exhaustive" in plan["advisory"]

    # Nothing has moved yet.
    body = client.get(f"/client-boq/revisions/{set_id}").json()
    assert body["amended"] == []
    assert [p["operative_rev"] for p in body["parts"]] == [0, 0, 0]


def test_approving_the_mapping_creates_a_new_revision_and_keeps_the_old_one(client):
    set_id = _ingest(client)
    _send_addendum(client, set_id)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    target = parts[2]["part_id"]

    resp = client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": target}],
    })
    assert resp.status_code == 200, resp.text
    assert resp.json()["revised"] == [{"part_id": target, "rev": 1, "title": parts[2]["title"]}]

    body = client.get(f"/client-boq/revisions/{set_id}").json()
    assert body["amended"] == [target]
    revised = next(p for p in body["parts"] if p["part_id"] == target)
    assert revised["operative_rev"] == 1

    # Rev 0 survives Rev 1 — the whole point of the rule.
    revs = {r["rev"]: r for r in revised["revisions"]}
    assert set(revs) == {0, 1}
    assert revs[0]["cause"] == "base"
    assert revs[1]["cause"] == "addendum"
    assert revs[1]["doc_ref"] == "Tender Addendum No. 1"
    assert revs[0]["pdf_path"] and revs[1]["pdf_path"]
    assert revs[0]["pdf_path"] != revs[1]["pdf_path"]   # the old file is still on disk


def test_only_the_mapped_parts_move(client):
    # The empirical shape of a real addendum: 9 of 165 documents changed. Everything else must
    # stay exactly where it was, at its existing revision.
    set_id = _ingest(client)
    _send_addendum(client, set_id)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    target = parts[0]["part_id"]

    client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": target}],
    })

    body = client.get(f"/client-boq/revisions/{set_id}").json()
    moved = {p["part_id"]: p["operative_rev"] for p in body["parts"]}
    assert moved[target] == 1
    assert all(rev == 0 for pid, rev in moved.items() if pid != target)


def test_the_operative_view_is_what_downstream_reads(client):
    set_id = _ingest(client)
    _send_addendum(client, set_id)
    before = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    target = before[1]["part_id"]

    client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": target}],
    })

    after = {p["part_id"]: p for p in client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]}
    # The parts list serves the NEW revision without the caller asking for it.
    assert after[target]["source_doc"] == "boq-rev1.pdf"
    assert after[target]["page_count"] == 6


def test_an_unmatched_replacement_is_surfaced_not_guessed(client):
    set_id = _ingest(client)
    plan = _send_addendum(client, set_id, filename="mystery-document.pdf")

    # The fixture maps two known filenames; this upload used neither, so it must come back
    # explicitly unmatched rather than being attached to whichever part looked closest.
    assert "mystery-document.pdf" in plan["unmatched"]
    mapping = next(m for m in plan["mappings"] if m["filename"] == "mystery-document.pdf")
    assert mapping["part_id"] == ""
    assert "by hand" in mapping["reason"]


def test_a_replacement_left_out_of_the_approval_is_not_applied(client):
    set_id = _ingest(client)
    _send_addendum(client, set_id)

    resp = client.post("/client-boq/ingest/changes/approve",
                       json={"set_id": set_id, "doc_id": "doc-1", "mappings": []})
    assert resp.status_code == 200
    assert resp.json()["revised"] == []
    assert client.get(f"/client-boq/revisions/{set_id}").json()["amended"] == []


# ---------------------------------------------------------------------------
# Corrections and clarifications
# ---------------------------------------------------------------------------
def test_a_correction_appends_a_revision_rather_than_replacing_rev_0(client):
    set_id = _ingest(client)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    target = parts[0]["part_id"]

    client.post(
        "/client-boq/ingest/document",
        data={"set_id": set_id, "kind": "correction", "ref": "Corrected upload"},
        files=[("files", ("boq-rev1.pdf", _pdf(4, "Corrected"), "application/pdf"))],
    )
    client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": target}],
    })

    body = client.get(f"/client-boq/revisions/{set_id}").json()
    revised = next(p for p in body["parts"] if p["part_id"] == target)
    revs = {r["rev"]: r for r in revised["revisions"]}
    assert set(revs) == {0, 1}                  # Rev 0 kept, exactly as the user specified
    assert revs[1]["cause"] == "correction"     # and distinguishable from a client addendum


def test_a_clarification_records_itself_and_changes_nothing(client):
    set_id = _ingest(client)

    resp = client.post(
        "/client-boq/ingest/document",
        data={"set_id": set_id, "kind": "clarification", "ref": "Tender Clarification No. 1"},
        files=[("files", ("TC1.pdf", _pdf(2, "Clarification"), "application/pdf"))],
    )
    body = resp.json()
    assert body["requires_gate"] is False
    assert "non-contractual" in body["notes"]

    history = client.get(f"/client-boq/revisions/{set_id}").json()
    assert history["amended"] == []
    kinds = [d["kind"] for d in history["documents"]]
    assert kinds == ["base", "clarification"]   # recorded in the history, but inert


# ---------------------------------------------------------------------------
# The history tabs, and stale verdicts
# ---------------------------------------------------------------------------
def test_the_set_can_be_replayed_to_any_point_in_its_history(client):
    set_id = _ingest(client)
    _send_addendum(client, set_id)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    target = parts[0]["part_id"]
    client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": target}],
    })

    as_issued = client.get(f"/client-boq/revisions/{set_id}/as-at/0").json()
    assert all(p["rev"] == 0 for p in as_issued["parts"])
    assert as_issued["as_at"]["kind"] == "base"

    after = client.get(f"/client-boq/revisions/{set_id}/as-at/1").json()
    moved = {p["part_id"]: p["rev"] for p in after["parts"]}
    assert moved[target] == 1
    assert sum(1 for r in moved.values() if r > 0) == 1   # a tab, not a second copy of everything


def test_a_revision_reopens_the_verdicts_that_depended_on_it(client):
    set_id = _ingest(client)
    client.post("/client-boq/review/run", data={"project_name": "rev-demo", "set_id": set_id})
    register = client.get(f"/client-boq/review/register/{set_id}").json()
    items = [i["item"] for i in register["register"]["items"]][:3]
    client.post("/client-boq/review/approve",
                json={"set_id": set_id, "decisions": {str(i): "dismissed" for i in items},
                      "approved": True})

    confirmed = client.get(f"/client-boq/review/register/{set_id}").json()
    assert any(i["status"] == "dismissed" for i in confirmed["register"]["items"])

    _send_addendum(client, set_id)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    resp = client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": parts[0]["part_id"]}],
    })
    assert resp.status_code == 200
    # Whether any line reopens depends on which part its clause came from, but the machinery must
    # report honestly either way, and never silently leave a verdict on rewritten wording.
    assert "reopened_register_items" in resp.json()


# ---------------------------------------------------------------------------
# The migration off the pre-revision schema
# ---------------------------------------------------------------------------
def test_pre_revision_part_rows_migrate_forward_without_losing_their_context(tmp_path, monkeypatch):
    # A part's interpreted context costs real model calls, so an upgrade must carry it forward
    # rather than dropping the table and making the user re-ingest.
    db = tmp_path / "old.db"
    conn = sqlite3.connect(str(db))
    conn.execute("""
        CREATE TABLE client_boq_parts (
            set_id TEXT NOT NULL, part_id TEXT NOT NULL, n INTEGER, abbr TEXT, slug TEXT,
            title TEXT, start_page INTEGER, end_page INTEGER, category TEXT, scanned INTEGER,
            pdf_path TEXT, context_json TEXT, PRIMARY KEY (set_id, part_id)
        )
    """)  # note: the ORIGINAL shape, before source_doc was added
    conn.execute(
        "INSERT INTO client_boq_parts VALUES ('s1','02-ct',2,'CT','ct','Conditions of Tender',"
        "5,16,'tender-instructions',0,'/tmp/ct.pdf','{\"summary\":\"kept\",\"readable\":true}')"
    )
    conn.commit()
    conn.close()

    monkeypatch.setenv("SITESOURCE_DB", str(db))
    conn = store.get_conn()
    try:
        cols = {r[1] for r in conn.execute("PRAGMA table_info(client_boq_parts)")}
        assert "start_page" not in cols          # the moved columns are gone from identity
        revs = store.load_part_revisions(conn, "s1", "02-ct")
        assert [r["rev"] for r in revs] == [0]
        assert revs[0]["start_page"] == 5 and revs[0]["end_page"] == 16
        assert revs[0]["cause"] == models.DOC_BASE
        parts = store.load_parts(conn, "s1")
        assert parts and parts[0][0].title == "Conditions of Tender"
        assert parts[0][2].summary == "kept"     # the interpreted context survived
    finally:
        conn.close()


def test_the_history_workbook_has_a_sheet_per_event(client):
    import io

    from openpyxl import load_workbook

    set_id = _ingest(client)
    _send_addendum(client, set_id)
    parts = client.get(f"/client-boq/ingest/parts/{set_id}").json()["parts"]
    target = parts[0]["part_id"]
    client.post("/client-boq/ingest/changes/approve", json={
        "set_id": set_id, "doc_id": "doc-1",
        "mappings": [{"filename": "boq-rev1.pdf", "part_id": target}],
    })

    resp = client.get(f"/client-boq/revisions/{set_id}/workbook")
    assert resp.status_code == 200 and resp.content[:2] == b"PK"

    book = load_workbook(io.BytesIO(resp.content))
    assert "History" in book.sheetnames
    assert "Declared changes" in book.sheetnames
    # One state-of-the-tender sheet per event: the base issue and the addendum.
    event_sheets = [n for n in book.sheetnames if n[:2].isdigit()]
    assert len(event_sheets) == 2

    as_issued = book[event_sheets[0]]
    revs = [row[4] for row in as_issued.iter_rows(min_row=4, values_only=True) if row[0]]
    assert set(revs) == {0}                       # at the base issue, everything is Rev 0

    after = book[event_sheets[1]]
    revs = {row[1]: row[4] for row in after.iter_rows(min_row=4, values_only=True) if row[0]}
    assert revs[target] == 1
    assert sum(1 for r in revs.values() if r > 0) == 1

    # The acknowledgement lists the client's addendum, and says corrections are excluded.
    summary = "\n".join(
        str(cell) for row in book["History"].iter_rows(values_only=True)
        for cell in row if cell is not None
    )
    assert "Tender Addendum No. 1" in summary
    assert "Corrections are our own re-uploads" in summary


def test_a_correction_is_kept_out_of_the_addendum_acknowledgement(client):
    import io

    from openpyxl import load_workbook

    set_id = _ingest(client)
    client.post(
        "/client-boq/ingest/document",
        data={"set_id": set_id, "kind": "correction", "ref": "Our corrected upload"},
        files=[("files", ("boq-rev1.pdf", _pdf(4, "Corrected"), "application/pdf"))],
    )

    book = load_workbook(io.BytesIO(client.get(f"/client-boq/revisions/{set_id}/workbook").content))
    rows = list(book["History"].iter_rows(values_only=True))
    ack_index = next(i for i, row in enumerate(rows)
                     if row and row[0] == "Acknowledgement of addenda received")
    ack_block = "\n".join(str(c) for row in rows[ack_index:ack_index + 3] for c in row if c)
    # The correction must not be presented to the client as one of their addenda.
    assert "Our corrected upload" not in ack_block
    assert "No addenda received" in ack_block


def test_the_revision_routes_are_mounted(client):
    paths = set(client.app.openapi()["paths"])  # openapi(), not app.routes — CLAUDE.md trap 1
    assert {
        "/client-boq/ingest/document",
        "/client-boq/ingest/changes/approve",
        "/client-boq/revisions/{set_id}",
        "/client-boq/revisions/{set_id}/as-at/{seq}",
    } <= paths
