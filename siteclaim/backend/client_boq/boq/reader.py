"""BOQ — read the client's bill of quantities out of their Excel workbook.

Bucket: **Deterministic**. openpyxl only; no model, no network.

Nothing here is defensive programming for its own sake. Every rule below exists because the real
ND/2025/04 workbook does the thing it guards against:

* **The item reference is stored as a float.** Excel drops the trailing zero, so item ``1.20`` is
  stored as ``1.2`` — the same value as item ``1.2``. Twelve such collisions exist in Rev 2 of the
  real bill, and the ONLY thing separating each pair is ``cell.number_format`` (``General`` vs
  ``'0.00'``). So the reference is rendered THROUGH the format, never taken from ``cell.value``.
  Reading it the obvious way silently merges items that carry different rates.

* **There is no description column.** Description is spread across B, C and D, and *the column is
  the indent level*. Item 2.9's own cell reads "maximum depth not exceeding 3.00m" — meaningless
  until you read up the chain to "Extra over for excavation in rock". General Preambles 2 makes
  that contractual: the "headings, sub-headings, item descriptions ... identify the work covered".

* **One sheet is structurally corrupt.** ``Bill No.4`` reports ``max_column = 16384`` with ~76,500
  stray cells and 9,963 merged ranges, from a page-header block fill-right-dragged across the sheet
  years ago. It prints fine because the print area is clamped. Every read here clamps to column H.

* **Page references live only in the page-break geometry.** ``BQ/2/1`` appears in no cell, yet the
  Index, the Grand Summary and both addenda all cite it.

The house rule applies throughout: **no silent drops**. Anything the reader cannot do cleanly is
recorded in ``notes`` on the item or the bill — a lossy reference render, a quantity found on the
wrong row, a unit that normalises to nothing. A note is cheap; a quietly wrong quantity is not.

ONE LIMIT WORTH KNOWING, because it is the format's and not this reader's
-------------------------------------------------------------------------
Nesting between captions **in the same column cannot be recovered**, because the workbook does not
encode it. In the reference bill, ``Trial Pits and Inspection Pits``, ``Trial pits``, ``Inspection
pits`` and ``Extra over for excavation in rock`` all sit in column B, at four different levels of
meaning, distinguished on the printed page only by the reader's eye. Leading whitespace is not the
signal either — it is applied inconsistently (``' Trial Pits and Inspection Pits'`` has a leading
space, ``'Trial pits '`` does not).

So ``heading_path`` holds the SMM **section** banner (detectable, and tracked separately for exactly
this reason) plus the nearest caption at each column depth. For item 2.9 that yields "SECTION 2 -
GROUND INVESTIGATION / Extra over for excavation in rock / maximum depth not exceeding 3.00m" — the
operative pair, which is what pricing and the diff both need. The intermediate group caption is not
recoverable, and inventing a hierarchy the file does not contain would be worse than saying so.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

from client_boq.models import BillItem, ClientBill, GrandSummaryLine

# Columns, fixed on every bill sheet of the reference workbook.
COL_REF, COL_B, COL_C, COL_D, COL_QTY, COL_UNIT, COL_RATE, COL_AMOUNT = range(1, 9)
MAX_COL = COL_AMOUNT          # the clamp that keeps Bill No.4 readable
DESC_COLS = (COL_B, COL_C, COL_D)

# The literal hyphen the SMM prescribes for a lump item. Corrigendum 1/2007, Part III para 3:
# "The symbol '-' shall be inserted against the rate and quantity columns in the bills for items of
# work for which the unit of measurement is 'item'."
DASH = {"-", "–", "—"}

_SHEET_BILL = re.compile(r"^\s*bill\s*(?:no\.?)?\s*(\d+)\s*$", re.IGNORECASE)
_NUMERIC_FORMAT = re.compile(r"^0(?:\.(0+))?$")
_SUB_REF = re.compile(r"^\(?([a-z])\)?$", re.IGNORECASE)
_SUMMARY_CODE = re.compile(r"\(([A-G])\)")
# What an item reference may look like: "1.5", "2.24", "1.61A". Anything else in column A is not a
# reference — the bill sheets repeat their page header every 58 rows, and its "Item No." would
# otherwise become an item and its "Item Description" a section heading.
_REF_SHAPE = re.compile(r"^\d+(?:\.\d+)?[A-Za-z]?$")
_PAGE_HEADER_TEXT = {"item no.", "item description", "quantity", "unit", "rate", "amount",
                     "(hk$)", "description"}
# The Method of Measurement's own sections, which the bill is organised by: "SECTION 1 -
# PRELIMINARIES", "SECTION 24 - LANDSCAPE SOFTWORKS AND ESTABLISHMENT WORKS", and so on. Held apart
# from the column-depth stack because a section banner sits in the same column as the group captions
# beneath it, and would otherwise be overwritten by the first of them.
_SECTION = re.compile(r"^\s*SECTION\s+\d+", re.IGNORECASE)


def normalise_unit(raw: str) -> str:
    """Fold the spelling variants that appear in one workbook into one unit.

    The reference bill uses 15 distinct raw strings for 14 units: ``item`` / ``item `` / ``Item``
    are one unit, and ``nr`` / ``nr.`` / ``nr. `` are another. Grouping or checking by the raw
    string would treat them as different.
    """
    return raw.strip().rstrip(".").strip().lower()


def format_ref(value: Any, number_format: str) -> tuple[str, str]:
    """Render an item reference the way Excel prints it. Returns ``(rendered, note)``.

    This is the single most load-bearing function in the reader. ``1.2`` under ``General`` is item
    1.2; the same ``1.2`` under ``'0.00'`` is item **1.20**, a different item at a different rate.

    ``note`` is non-empty when the render is lossy — the reference cell held more precision than it
    displays. That is not hypothetical: ``Bill No.2!A84`` of the real workbook holds ``2.244`` under
    ``'0.00'`` and prints as "2.24". The printed form is the item's true identity (it is what the
    PDF, the Index and the addenda all use), so it is what is returned — but the discrepancy is a
    data-entry error in the client's file and the estimator should be told.
    """
    if value is None:
        return "", ""
    if isinstance(value, str):
        return value.strip(), ""
    fmt = (number_format or "General").strip().strip("'\"")
    match = _NUMERIC_FORMAT.match(fmt)
    if match:
        places = len(match.group(1) or "")
        rendered = f"{float(value):.{places}f}"
    elif float(value) == int(float(value)):
        rendered = str(int(float(value)))
    else:
        rendered = str(value)
    note = ""
    if float(rendered) != float(value):
        note = (f"reference cell holds {value!r} but displays as {rendered!r} "
                f"(number format {fmt!r}); the displayed form is used, as the PDF and the addenda do")
    return rendered, note


def _page_map(sheet: Any) -> dict[int, int]:
    """row → 1-based printed page, from the sheet's horizontal page breaks.

    The ``BQ/n/m`` reference every other document cites is not written anywhere in the workbook; it
    exists only as this geometry. A break at row *r* means *r* is the LAST row of that page.
    """
    breaks = sorted({int(b.id) for b in getattr(sheet.row_breaks, "brk", []) or [] if b.id})
    out: dict[int, int] = {}
    page, cursor = 1, 1
    for stop in breaks:
        for row in range(cursor, stop + 1):
            out[row] = page
        cursor, page = stop + 1, page + 1
    for row in range(cursor, (sheet.max_row or cursor) + 1):
        out[row] = page
    return out


def _text(cells: dict[int, Any], col: int) -> str:
    value = cells.get(col)
    return value if isinstance(value, str) else ("" if value is None else str(value))


def _number(value: Any) -> Optional[float]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


def _is_page_header(cells: dict[int, Any]) -> bool:
    """A repeated page-header row. Every printed page of a bill restates ``'Bill No. 2'`` /
    ``Item No.`` / ``Item Description`` / ``Quantity`` / ``Unit`` / ``Rate`` / ``Amount``. Left
    alone, "Item No." becomes an item and "Item Description" becomes a section heading that then
    attaches itself to every real item on the page."""
    for col in (COL_REF, *DESC_COLS):
        text = _text(cells, col).strip().strip("'").lower()
        if text.startswith("bill no") or text in _PAGE_HEADER_TEXT:
            return True
    return False


def _join_description(existing: str, addition: str) -> str:
    """Join a hard-wrapped continuation row onto the text above it.

    The author wrapped at the column width, so some continuations carry their own joining space
    (" than 7 seats excluding driver") and some do not ("soil and ground water " following
    "Determination of carbonate content test for"). Concatenating blindly produces "...forsoil...";
    always inserting a space breaks the first case. So a space goes in only when neither side
    already has one.
    """
    if existing and addition and not existing[-1].isspace() and not addition[0].isspace():
        return f"{existing} {addition}".strip()
    return f"{existing}{addition}".strip()


def _read_bill_sheet(sheet: Any, bill_no: str) -> tuple[list[BillItem], list[str]]:
    """One ``Bill No.N`` sheet → its items, plus any sheet-level notes."""
    pages = _page_map(sheet)
    items: list[BillItem] = []
    notes: list[str] = []
    headings: dict[int, str] = {}          # depth (0=B, 1=C, 2=D) → the caption in force
    section = ""                           # the SMM section banner, tracked apart from the stack
    open_item: Optional[BillItem] = None
    open_desc_col = 0
    open_row = 0
    caption_row, caption_depth = -2, -1    # for joining a caption wrapped across two rows
    last_numeric_ref = ""
    # A quantity/unit stranded on the caption row directly above an item. Exactly one item in the
    # reference bill is written this way (1.16, rows 84/85), and it prints correctly because the
    # rows are visually adjacent — so a reader that only looks at the item's own row loses it.
    stranded: dict[str, Any] = {}

    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row or 1,
                                     min_col=1, max_col=MAX_COL):
        row = row_cells[0].row
        cells = {cell.column: cell.value for cell in row_cells}
        ref_cell = row_cells[COL_REF - 1]

        raw_ref, ref_note = format_ref(cells.get(COL_REF), ref_cell.number_format)
        qty_cell = cells.get(COL_QTY)
        unit_raw = _text(cells, COL_UNIT)
        desc_col = next((c for c in DESC_COLS if _text(cells, c).strip()), 0)
        blank = not raw_ref and not desc_col and qty_cell is None

        if blank:
            open_item, open_desc_col = None, 0
            stranded = {}
            continue

        if _is_page_header(cells):
            open_item, open_desc_col = None, 0
            stranded = {}
            continue

        if raw_ref and not _SUB_REF.match(raw_ref) and not _REF_SHAPE.match(raw_ref):
            notes.append(f"{sheet.title} row {row}: column A holds {raw_ref!r}, which is not an item "
                         f"reference; the row was skipped")
            continue

        if raw_ref:
            sub = _SUB_REF.match(raw_ref)
            if sub and last_numeric_ref:
                # A lettered variant: "a" / "b" under a parent that keeps its own number. The
                # spreadsheet never stores "2.2a" anywhere — the addendum's own wording ("item nos.
                # 2.2a and 2.2b") is the concatenation, so it is rebuilt here. Unqualified, the bare
                # "a" in Bill No.2 and the bare "a" in Bill No.3 would collide.
                sub_ref = sub.group(1).lower()
                item_ref, full_ref = last_numeric_ref, f"{last_numeric_ref}{sub_ref}"
                for parent in items:
                    if parent.full_ref == last_numeric_ref:
                        parent.is_parent = True
            else:
                sub_ref = ""
                item_ref = full_ref = raw_ref
                last_numeric_ref = raw_ref

            qty, lump = None, False
            if isinstance(qty_cell, str) and qty_cell.strip() in DASH:
                lump = True
            else:
                qty = _number(qty_cell)

            item_notes = [ref_note] if ref_note else []
            if qty is None and not lump and stranded:
                qty, lump = stranded.get("qty"), stranded.get("lump", False)
                unit_raw = unit_raw or stranded.get("unit", "")
                item_notes.append(
                    f"quantity and unit were on the caption row above (row {stranded['row']}), "
                    f"not on the item's own row; adopted, because the printed page reads them "
                    f"together")
            stranded = {}

            rate = _number(cells.get(COL_RATE))
            amount = _number(cells.get(COL_AMOUNT))
            unit = normalise_unit(unit_raw)
            if not lump and qty is None:
                item_notes.append("no quantity and no '-' in the quantity cell")
            if unit_raw.strip() and not unit:
                item_notes.append(f"unit {unit_raw!r} normalises to nothing")

            item = BillItem(
                bill_no=bill_no, item_ref=item_ref, sub_ref=sub_ref, full_ref=full_ref,
                heading_path=([section] if section else []) +
                             [headings[d] for d in sorted(headings) if headings[d].strip()],
                description=_text(cells, desc_col).strip() if desc_col else "",
                unit_raw=unit_raw, unit=unit, qty=qty, lump=lump,
                client_rate=rate, client_amount=amount,
                pre_priced=rate is not None or amount is not None,
                page_ref=f"BQ/{bill_no}/{pages.get(row, 1)}", sheet=sheet.title, row=row,
                notes=item_notes,
            )
            items.append(item)
            open_item, open_desc_col, open_row = item, desc_col, row
            continue

        if not desc_col:
            continue

        text = _text(cells, desc_col)
        font = row_cells[desc_col - 1].font
        emphasised = bool(getattr(font, "bold", False) or getattr(font, "underline", None))
        contiguous = open_item is not None and row == open_row + 1

        if open_item is not None and contiguous and not emphasised and desc_col == open_desc_col:
            # A continuation of the item's own description, hard-wrapped by the author at the
            # column width.
            open_item.description = _join_description(open_item.description, text)
            open_row = row
            continue

        depth = DESC_COLS.index(desc_col)
        # A caption is wrapped across rows just as an item description is: "SECTION 24 - LANDSCAPE
        # SOFTWORKS AND " / "ESTABLISHMENT WORKS" is one heading on two rows, not two headings.
        wrapped = caption_row == row - 1 and caption_depth == depth

        if _SECTION.match(text) or (wrapped and section and caption_depth == depth and not headings):
            section = _join_description(section, text) if wrapped else text.strip()
            headings = {}
            open_item, open_desc_col = None, 0
            stranded = {}
            caption_row, caption_depth = row, depth
            continue

        headings[depth] = _join_description(headings.get(depth, ""), text) if wrapped else text.strip()
        for deeper in [d for d in headings if d > depth]:
            headings.pop(deeper)
        open_item, open_desc_col = None, 0
        caption_row, caption_depth = row, depth
        # A caption that carries a quantity is the 1.16 case: hold it for the next item row.
        if cells.get(COL_QTY) is not None:
            qty_value = cells.get(COL_QTY)
            stranded = {
                "row": row, "unit": _text(cells, COL_UNIT),
                "lump": isinstance(qty_value, str) and qty_value.strip() in DASH,
                "qty": _number(qty_value),
            }
        if not emphasised:
            notes.append(
                f"{sheet.title} row {row}: {text.strip()[:60]!r} read as a heading, but it carries "
                f"no bold or underline — the usual signal on this sheet")

    return items, notes


def _read_grand_summary(sheet: Any) -> list[GrandSummaryLine]:
    """The Grand Summary. (B), (D) and (E) arrive filled by the client and are reinstated if a
    tenderer alters them (GCT App C 2.5); everything else is ours to compute."""
    lines: list[GrandSummaryLine] = []
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row or 1, min_col=1, max_col=8):
        cells = {cell.column: cell.value for cell in row_cells}
        label = next((_text(cells, c) for c in (COL_C, COL_B, COL_D) if _text(cells, c).strip()), "")
        if not label.strip():
            continue
        amount = next((_number(cells.get(c)) for c in (4, COL_AMOUNT) if _number(cells.get(c)) is not None), None)
        bill_no = _number(cells.get(COL_B))
        code = ""
        found = _SUMMARY_CODE.search(label)
        if found:
            code = found.group(1)
        elif bill_no is not None:
            code = str(int(bill_no))
        lines.append(GrandSummaryLine(
            label=" ".join(label.split()), code=code, amount=amount,
            client_inserted=amount is not None,
        ))
    return lines


def read_workbook(path: str | Path, *, set_id: str = "", rev: int = 0) -> ClientBill:
    """Read a client bill-of-quantities workbook into a :class:`ClientBill`.

    Reads values, not formulas (``data_only=True``): the reference workbook holds only 16 formulas
    in total, 14 of them in the one bill the client pre-priced, so a formula is the exception and a
    cached value is what the tenderer sees.
    """
    import openpyxl                                       # lazy: this is the only openpyxl user here

    source = Path(path)
    workbook = openpyxl.load_workbook(source, data_only=True)
    items: list[BillItem] = []
    notes: list[str] = []
    summary: list[GrandSummaryLine] = []

    for name in workbook.sheetnames:
        sheet = workbook[name]
        match = _SHEET_BILL.match(name)
        if match:
            sheet_items, sheet_notes = _read_bill_sheet(sheet, match.group(1))
            items.extend(sheet_items)
            notes.extend(sheet_notes)
            if (sheet.max_column or 0) > MAX_COL:
                notes.append(
                    f"{name}: the sheet reports {sheet.max_column} columns; everything past column "
                    f"H was ignored. The reference workbook carries ~76,500 stray cells on one "
                    f"sheet from an old fill-right, and the print area excludes them too")
        elif name.strip().lower() in {"grand summary", "grandsummary", "summary"}:
            summary = _read_grand_summary(sheet)

    seen: dict[str, str] = {}
    for item in items:
        key = f"{item.bill_no}:{item.full_ref}"
        if key in seen:
            notes.append(f"duplicate item reference {item.full_ref!r} in Bill No.{item.bill_no} "
                         f"(rows {seen[key]} and {item.row}); both kept, neither is assumed correct")
        seen[key] = str(item.row)

    workbook.close()
    return ClientBill(set_id=set_id, rev=rev, source_file=source.name, items=items,
                      summary=summary, notes=notes)
