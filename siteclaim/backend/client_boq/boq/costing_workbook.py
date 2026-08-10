"""BOQ — the costing workbook. The deliverable, and a live model rather than a report.

Bucket: **Deterministic output.** Eight sheets, **with their formulas intact**, so the estimator
changes a blue cell on ``01 Inputs`` and every rate on ``05 BQ Priced`` moves in Excel.

WHY FORMULAS AND NOT VALUES
---------------------------
Every other workbook this repo writes emits dead numbers, and that is right for a record of something
already decided. This one is different: it is the estimator's working model, handed back. A report
tells him what the app thinks; a model lets him disagree with it in the tool he already trusts, on a
train, without the app running.

It also means the app never has to be the only place the arithmetic exists — which is the honest
answer to "what if I want to do something the app cannot".

WHY IT IS GENERATED FROM THE MODEL
----------------------------------
The sheets are laid out from :class:`CostingModel`, not from a fixed template. Add a production band
and a row appears; delete a plant line and the plant block shrinks; re-point a driver and sheet 04
follows. The cross-sheet references are computed as the rows are written, so they stay correct
whatever shape the model has been edited into.

THE COLOUR CONVENTION, from the reference workbook's own README:

    BLUE   a hardcoded input you set
    BLACK  a formula — do not overwrite
    GREEN  a link to another sheet
    YELLOW a key assumption, to be confirmed on the register before submission

A NOTE ON TESTING
-----------------
``openpyxl`` writes formulas but never evaluates them, so nothing here can be checked by reading the
numbers back. The numbers are proven in Python against the same model; this module is proven
structurally — the sheets exist, the references point where they should, and at default settings the
formulas match the reference workbook cell for cell.
"""

from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from client_boq.boq import empirical
from client_boq.boq.assumptions import Register
from client_boq.boq.buildup import Buildup, Spread
from client_boq.boq.costing import PricedBQ
from client_boq.boq.model import (
    CHARGE_CONTRACT_DAY,
    CHARGE_GFT,
    CHARGE_PRELIM,
    CHARGE_RIG_DAY,
    CostingModel,
)
from client_boq.boq.programme import Programme

# The convention, as fonts. Blue is yours to change; black is arithmetic; green points elsewhere.
INPUT = Font(color="FF0000CC")
FORMULA = Font(color="FF000000")
LINK = Font(color="FF107C41")
TITLE = Font(bold=True, size=14)
HEADING = Font(bold=True, size=11)
LABEL = Font(bold=True, color="FF44546A")
NOTE = Font(italic=True, size=9, color="FF7F7F7F")
KEY_ASSUMPTION = PatternFill("solid", fgColor="FFFFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="FFEFEFEF")

MONEY = "#,##0.00"
NUMBER = "#,##0.00"
PERCENT = "0.0%"

# What the "Charged to" column says. Three day-costs, and a fourth thing that is not a day-cost at
# all: a preliminary is billed as its own item and must never be folded into either total.
CHARGE_LABELS = {
    CHARGE_RIG_DAY: "per rig-day",
    CHARGE_CONTRACT_DAY: "per contract-day",
    CHARGE_GFT: "per GFT-day",
    CHARGE_PRELIM: "billed as its own item",
}

SHEETS = ("00 README", "01 Inputs", "02 Production", "03 Resource Rates", "04 Item Buildup",
          "05 BQ Priced", "06 Assumptions Register", "07 Empirical Basis")


class _Ref:
    """Where a value ended up, so a later sheet can point at it instead of repeating it."""

    def __init__(self) -> None:
        self._at: dict[str, str] = {}

    def put(self, key: str, sheet: str, cell: str) -> None:
        self._at[key] = f"'{sheet}'!{cell}"

    def get(self, key: str) -> str:
        return self._at.get(key, "")

    def formula(self, key: str) -> str:
        target = self.get(key)
        return f"={target}" if target else "=0"


def build_workbook(model: CostingModel, programme: Programme, spread: Spread, buildup: Buildup,
                   priced: PricedBQ, register: Register, *,
                   contract_reference: str = "") -> bytes:
    """The eight sheets, wired together. Returns the .xlsx bytes."""
    book = Workbook()
    book.remove(book.active)
    where = _Ref()

    _readme(book.create_sheet(SHEETS[0]))
    _inputs(book.create_sheet(SHEETS[1]), model, where, contract_reference)
    _production(book.create_sheet(SHEETS[2]), programme, model, where)
    _resource_rates(book.create_sheet(SHEETS[3]), model, spread, where)
    _item_buildup(book.create_sheet(SHEETS[4]), model, buildup, where)
    _bq_priced(book.create_sheet(SHEETS[5]), priced, where)
    _register(book.create_sheet(SHEETS[6]), register)
    _empirical(book.create_sheet(SHEETS[7]))

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
def _readme(ws) -> None:
    ws["A1"] = "GI Tender Costing"
    ws["A1"].font = TITLE
    ws["A2"] = ("SiteSource — pricing engine output. Bottom-up cost build-up for ground "
                "investigation tenders.")
    ws["A2"].font = NOTE

    rows = [
        ("PURPOSE", "Convert a bill of quantities into defensible unit rates via a bottom-up cost "
                    "build-up, with every assumption recorded for human review before submission."),
        ("HOW TO USE", "1. Every BLUE cell on 01 Inputs is yours to change. Everything else "
                       "calculates."),
        ("", "2. Read 02 Production to check the derived programme is sane — days, rigs, "
             "standing time."),
        ("", "3. Review every line on 06 Assumptions Register and set the Status column. "
             "That is the human gate."),
        ("", "4. Only the RATE TO SUBMIT column on 05 BQ Priced goes into the client's bill."),
        ("COLOUR", "BLUE = an input you set · BLACK = a formula · GREEN = a link to another sheet"),
        ("", "YELLOW fill = a key assumption, to be confirmed on the register before submission."),
        ("CRITICAL", "This is a POINT ESTIMATE on a wide distribution. The as-built data on 07 "
                     "shows a 4.0x spread between the 10th and 90th percentile hole."),
        ("", "02 Production carries that band through. Do not submit the P50 rate without reading "
             "the P90 column."),
    ]
    for n, (head, text) in enumerate(rows, start=4):
        ws.cell(row=n, column=1, value=head).font = LABEL
        ws.cell(row=n, column=2, value=text).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 100


# ---------------------------------------------------------------------------
def _inputs(ws, model: CostingModel, where: _Ref, contract_reference: str) -> None:
    """01 Inputs — every lever in one place, and the only sheet with typed numbers on it."""
    ws["A1"] = "01 — Inputs"
    ws["A1"].font = TITLE
    ws["A2"] = "Edit the BLUE cells. Yellow fill marks an assumption to confirm before submission."
    ws["A2"].font = NOTE
    for n, head in enumerate(["Input", "Value", "Unit", "Basis / note"], start=2):
        cell = ws.cell(row=4, column=n, value=head)
        cell.font = HEADING
        cell.fill = HEADER_FILL

    row = 5

    def block(title: str) -> None:
        nonlocal row
        ws.cell(row=row, column=2, value=title).font = LABEL
        row += 1

    def scalar(key: str, label: str, unit: str = "", note: str = "",
               fmt: Optional[str] = None, key_assumption: bool = False) -> None:
        nonlocal row
        ws.cell(row=row, column=2, value=label)
        cell = ws.cell(row=row, column=3, value=model.value(key))
        cell.font = INPUT
        if fmt:
            cell.number_format = fmt
        if key_assumption:
            cell.fill = KEY_ASSUMPTION
        ws.cell(row=row, column=4, value=unit)
        ws.cell(row=row, column=5, value=note).font = NOTE
        where.put(f"inputs.{key}", ws.title, f"$C${row}")
        row += 1

    block("CONTRACT")
    ws.cell(row=row, column=2, value="Contract reference")
    ws.cell(row=row, column=3, value=contract_reference or model.name).font = INPUT
    row += 1
    scalar("contract_period_months", "Contract period", "month")
    scalar("working_days_per_month", "Working days per month", "day", "6-day week, HK civils norm.")

    block("PRODUCTION BANDS  (a lookup, not a regression — see 07 for why)")
    for n, head in enumerate(["Band", "Lower bound", "m/work-day", "n holes",
                              "calibration depth"], start=2):
        ws.cell(row=row, column=n, value=head).font = HEADING
    row += 1
    first_band_row = row
    for band in model.bands.sorted_bands():
        ws.cell(row=row, column=2, value=band.label)
        for col, value, fmt in ((3, band.lower, PERCENT), (4, band.rate, None),
                                (5, band.holes, None), (6, band.calibration_depth_m, None)):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = INPUT
            if fmt:
                cell.number_format = fmt
        row += 1
    last_band_row = row - 1
    where.put("bands.lower", ws.title, f"$C${first_band_row}:$C${last_band_row}")
    where.put("bands.rate", ws.title, f"$D${first_band_row}:$D${last_band_row}")
    where.put("bands.holes", ws.title, f"$E${first_band_row}:$E${last_band_row}")
    where.put("bands.depth", ws.title, f"$F${first_band_row}:$F${last_band_row}")
    where.put("bands.label", ws.title, f"$B${first_band_row}:$B${last_band_row}")
    ws.cell(row=row, column=2,
            value="All-in blended rates: total metres over total work-days, including per-hole "
                  "set-up. Refit whenever the corpus changes.").font = NOTE
    row += 2

    block("PRODUCTIVITY")
    scalar("residual_site_factor", "Residual site factor", "x",
           "What is left AFTER the rock-fraction band. 1.00 means the band explains the site. "
           "Raise it only for a named reason, recorded on 06.", key_assumption=True)
    scalar("calendar_to_work_day", "Calendar : work-day ratio", "x",
           "Observed 1.18 / 1.18 / 1.36. Weather lands here — it costs programme, not production.")
    scalar("standing_allowance", "Standing / idle allowance", "of work-days",
           "Observed 18%–36%. Drives the standing time item.", fmt=PERCENT, key_assumption=True)
    scalar("p90_multiplier", "Productivity band — P90 multiplier", "x")
    scalar("p10_multiplier", "Productivity band — P10 multiplier", "x")
    scalar("hours_per_day", "Hours per working day", "h")

    block("SPLIT-RATE CROSS-CHECK  (Method B — not the pricing basis)")
    scalar("setup_days_per_hole", "Fixed set-up time per hole", "work-day",
           "Regression intercept. Used only for the cross-check on 02.")
    scalar("soil_m_per_day", "Soil production rate", "m/work-day")
    scalar("rock_m_per_day", "Rock production rate", "m/work-day")

    block("COMMERCIAL")
    for key, label, note in (
        ("margin", "Direct margin", "Taken on the selling price, not added to cost."),
        ("overhead_local", "Local overhead", "Site + local office establishment."),
        ("overhead_regional", "Regional overhead", ""),
        ("overhead_international", "International overhead", ""),
        ("nec_fee", "NEC fee percentage",
         "Applies to compensation-event Defined Cost, NOT to the BQ rates."),
        ("risk_loading", "Risk / contingency loading",
         "Priced-in allowance for the productivity spread. See the P90 column on 02."),
    ):
        scalar(key, label, "", note, fmt=PERCENT, key_assumption=key == "margin")

    block("ORGANISATION")
    scalar("site_count", "Number of sites", "nr",
           "The site team is carried per site, not per rig.")
    scalar("site_team_per_site", "Site teams per site", "nr",
           "A coefficient: 1.0 dedicated, 0.5 shared with another contract. Not rounded up.")
    scalar("gft_ratio", "One GFT supervises", "rigs",
           "The GFT manages RIGS. A different resource from the site team, which manages a site.")

    block("SPREAD  (per day — the resources standing on site)")
    for n, head in enumerate(["Resource", "Multiplier", "Rate ($/day)", "Charged to"], start=2):
        ws.cell(row=row, column=n, value=head).font = HEADING
    row += 1
    for line in model.spread:
        ws.cell(row=row, column=2, value=line.label)
        ws.cell(row=row, column=3, value=line.multiplier).font = INPUT
        rate = ws.cell(row=row, column=4, value=line.rate)
        rate.font = INPUT
        rate.number_format = MONEY
        ws.cell(row=row, column=5, value=CHARGE_LABELS.get(line.charge, "per contract-day"))
        ws.cell(row=row, column=6, value=line.note).font = NOTE
        where.put(f"spread.{line.key}.multiplier", ws.title, f"$C${row}")
        where.put(f"spread.{line.key}.rate", ws.title, f"$D${row}")
        row += 1
    row += 1

    block("MATERIALS")
    for key, label, unit in (
        ("mazier_interval_m", "Mazier sample interval", "m"),
        ("mazier_sample_length_m", "Mazier sample length", "m"),
        ("soil_tube_cost", "Soil tube — unit cost", "$/nr"),
        ("core_box_cost", "Wooden core box — unit cost", "$/nr"),
        ("soil_box_capacity_m", "Soil box capacity", "m/box"),
        ("rock_box_capacity_m", "Rock box capacity", "m/box"),
        ("grout_hole_diameter_m", "Grout hole diameter", "m"),
        ("grout_cost_per_litre", "Cement grout — unit cost", "$/L"),
    ):
        scalar(key, label, unit)

    block("MOBILISATION")
    for key, label, unit in (
        ("crane_lorry_rate", "Crane lorry", "$/day"),
        ("crane_lorry_days", "Crane lorry days", "day"),
        ("truck_rate", "Truck (equipment)", "$/day"),
        ("truck_days", "Truck days", "day"),
        ("survey_per_location", "Surveying / setting out", "$/location"),
    ):
        scalar(key, label, unit)

    block("LABORATORY  (subcontract buy rates)")
    for lab in model.laboratory:
        ws.cell(row=row, column=2, value=lab.label)
        cell = ws.cell(row=row, column=3, value=lab.rate)
        cell.font = INPUT
        cell.number_format = MONEY
        cell.fill = KEY_ASSUMPTION
        ws.cell(row=row, column=4, value="$/test")
        ws.cell(row=row, column=5, value=lab.note).font = NOTE
        where.put(f"lab.{lab.key}", ws.title, f"$C${row}")
        row += 1

    _widths(ws, {"B": 40, "C": 14, "D": 16, "E": 62, "F": 40})


# ---------------------------------------------------------------------------
def _production(ws, programme: Programme, model: CostingModel, where: _Ref) -> None:
    """02 Production — quantities to work-days. The quantities are typed; the rest calculates."""
    ws["A1"] = "02 — Production model"
    ws["A1"].font = TITLE
    ws["A2"] = "P50 is the pricing basis. P10 and P90 show the exposure."
    ws["A2"].font = NOTE
    for n, head in enumerate(["", "P10 (fast)", "P50 (pricing basis)", "P90 (slow)", "Unit",
                              "Note"], start=2):
        ws.cell(row=4, column=n, value=head).font = HEADING

    q = programme.quantities
    row = 5
    ws.cell(row=row, column=2, value="QUANTITIES  (from the bill)").font = LABEL
    row += 1
    for key, label, value, unit in (
        ("holes", "Number of drillholes / rig moves", q.holes, "nr"),
        ("soil_m", "Soil drilling (material other than rock)", q.soil_m, "m"),
        ("rock_m", "Rock drilling", q.rock_m, "m"),
        ("hard_m", "Artificial hard material / boulder", q.hard_m, "m"),
    ):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=4, value=value).font = INPUT
        ws.cell(row=row, column=5, value=unit)
        where.put(f"q.{key}", ws.title, f"$D${row}")
        row += 1

    total_row = row
    ws.cell(row=row, column=2, value="Total drilled length")
    ws.cell(row=row, column=4,
            value=f"={where.get('q.soil_m')}+{where.get('q.rock_m')}+{where.get('q.hard_m')}")
    ws.cell(row=row, column=5, value="m")
    where.put("q.total_m", ws.title, f"$D${total_row}")
    row += 2

    ws.cell(row=row, column=2, value="BAND SELECTION  (derived — nothing typed here)").font = LABEL
    row += 1
    fraction_row = row
    ws.cell(row=row, column=2, value="Rock fraction of the works")
    ws.cell(row=row, column=4,
            value=f"=({where.get('q.rock_m')}+{where.get('q.hard_m')})/{where.get('q.total_m')}")
    ws.cell(row=row, column=4).number_format = PERCENT
    ws.cell(row=row, column=6, value="Rock and hard material as a share of everything drilled. "
                                     "Read from the bill.").font = NOTE
    where.put("band.fraction", ws.title, f"$D${fraction_row}")
    row += 1

    index_row = row
    ws.cell(row=row, column=2, value="Band index")
    ws.cell(row=row, column=4,
            value=f"=MATCH({where.get('band.fraction')},{where.get('bands.lower')},1)")
    where.put("band.index", ws.title, f"$D${index_row}")
    row += 1

    for key, label, source, unit in (
        ("band.label", "Band selected", "bands.label", ""),
        ("band.rate", "Band base production rate", "bands.rate", "m/work-day"),
        ("band.holes", "Band sample size", "bands.holes", "holes"),
        ("band.depth", "Band calibration mean hole depth", "bands.depth", "m"),
    ):
        ws.cell(row=row, column=2, value=label)
        cell = ws.cell(row=row, column=4,
                       value=f"=INDEX({where.get(source)},{where.get('band.index')})")
        cell.font = LINK
        ws.cell(row=row, column=5, value=unit)
        where.put(key, ws.title, f"$D${row}")
        row += 1

    ws.cell(row=row, column=2, value="This tender — mean hole depth")
    ws.cell(row=row, column=4, value=f"={where.get('q.total_m')}/{where.get('q.holes')}")
    ws.cell(row=row, column=5, value="m")
    where.put("mean_depth", ws.title, f"$D${row}")
    row += 1

    ws.cell(row=row, column=2, value="Depth departure from the band")
    ws.cell(row=row, column=4, value=f"={where.get('mean_depth')}/{where.get('band.depth')}-1")
    ws.cell(row=row, column=4).number_format = PERCENT
    ws.cell(row=row, column=6,
            value=f"Over {model.method.depth_departure_threshold:.0%} either way means the band is "
                  f"being extrapolated. Record it on 06.").font = NOTE
    where.put("depth_departure", ws.title, f"$D${row}")
    row += 2

    ws.cell(row=row, column=2,
            value="DERIVED PROGRAMME  (Method A — banded, the pricing basis)").font = LABEL
    row += 1
    days_row = row
    ws.cell(row=row, column=2, value="TOTAL RIG WORK-DAYS").font = HEADING
    ws.cell(row=row, column=3,
            value=f"=$D${days_row}*{where.get('inputs.p10_multiplier')}")
    ws.cell(row=row, column=4,
            value=(f"={where.get('q.total_m')}/{where.get('band.rate')}"
                   f"*{where.get('inputs.residual_site_factor')}"))
    ws.cell(row=row, column=5,
            value=f"=$D${days_row}*{where.get('inputs.p90_multiplier')}")
    ws.cell(row=row, column=6, value="work-day")
    where.put("work_days", ws.title, f"$D${days_row}")
    where.put("work_days_p90", ws.title, f"$E${days_row}")
    row += 1

    ws.cell(row=row, column=2, value="Implied blended rate")
    ws.cell(row=row, column=4, value=f"={where.get('q.total_m')}/{where.get('work_days')}")
    ws.cell(row=row, column=6, value="m/work-day")
    row += 2

    ws.cell(row=row, column=2,
            value="CROSS-CHECK  (Method B — split rates. If A and B diverge, do not "
                  "price.)").font = LABEL
    row += 1
    setup_row = row
    for key, label, formula in (
        ("b.setup", "Fixed set-up time (all holes)",
         f"={where.get('q.holes')}*{where.get('inputs.setup_days_per_hole')}"),
        ("b.soil", "Soil drilling time",
         f"={where.get('q.soil_m')}/{where.get('inputs.soil_m_per_day')}"),
        ("b.rock", "Rock drilling time",
         f"=({where.get('q.rock_m')}+{where.get('q.hard_m')})/"
         f"{where.get('inputs.rock_m_per_day')}"),
    ):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=5, value="work-day")
        where.put(key, ws.title, f"$D${row}")
        row += 1

    ws.cell(row=row, column=2, value="Method B total")
    ws.cell(row=row, column=4,
            value=(f"=SUM($D${setup_row}:$D${row - 1})"
                   f"*{where.get('inputs.residual_site_factor')}"))
    where.put("b.total", ws.title, f"$D${row}")
    row += 1

    ws.cell(row=row, column=2, value="Divergence A against B")
    ws.cell(row=row, column=4, value=f"={where.get('b.total')}/{where.get('work_days')}-1")
    ws.cell(row=row, column=4).number_format = PERCENT
    where.put("divergence", ws.title, f"$D${row}")
    row += 1

    ws.cell(row=row, column=2, value="Convergence check")
    ws.cell(row=row, column=4, value=(
        f'=IF(ABS({where.get("divergence")})>{model.method.divergent_threshold},'
        f'"DIVERGENT — investigate before pricing",'
        f'IF(ABS({where.get("divergence")})>{model.method.marginal_threshold},'
        f'"marginal","converged"))'))
    row += 1

    ws.cell(row=row, column=2, value="Allocation scale (A ÷ B)")
    ws.cell(row=row, column=4, value=f"={where.get('work_days')}/{where.get('b.total')}")
    ws.cell(row=row, column=6,
            value="The band sets the TOTAL. The split sets how it divides between set-up, soil and "
                  "rock. This reconciles them.").font = NOTE
    where.put("allocation", ws.title, f"$D${row}")
    row += 2

    ws.cell(row=row, column=2, value="PROGRAMME AND RESOURCES").font = LABEL
    row += 1
    ws.cell(row=row, column=2, value="Calendar days (weather lands here, not in the rate)")
    ws.cell(row=row, column=4,
            value=f"={where.get('work_days')}*{where.get('inputs.calendar_to_work_day')}")
    ws.cell(row=row, column=6, value="cal-day")
    row += 1

    ws.cell(row=row, column=2, value="Work-days available in the contract period, per rig")
    ws.cell(row=row, column=4,
            value=(f"={where.get('inputs.contract_period_months')}"
                   f"*{where.get('inputs.working_days_per_month')}"))
    where.put("days_available", ws.title, f"$D${row}")
    row += 1

    ws.cell(row=row, column=2, value="RIGS REQUIRED (concurrent)").font = HEADING
    ws.cell(row=row, column=4, value=f"={where.get('work_days')}/{where.get('days_available')}")
    ws.cell(row=row, column=6, value="Round up. Drives Bill 1 time-related preliminaries.")
    where.put("rigs", ws.title, f"$D${row}")
    row += 1

    ws.cell(row=row, column=2, value="Standing time")
    ws.cell(row=row, column=4,
            value=(f"={where.get('work_days')}*{where.get('inputs.standing_allowance')}"
                   f"*{where.get('inputs.hours_per_day')}"))
    ws.cell(row=row, column=6, value="hour")
    where.put("standing_hours", ws.title, f"$D${row}")
    row += 2

    ws.cell(row=row, column=2, value="DERIVED MATERIAL QUANTITIES").font = LABEL
    row += 1
    for key, label, formula, unit in (
        ("m.samples", "Mazier samples",
         f"=ROUND({where.get('q.soil_m')}/{where.get('inputs.mazier_interval_m')},0)", "nr"),
        ("m.in_tubes", "Soil retained in tubes", None, "m"),
        ("m.for_boxing", "Soil remaining for boxing", None, "m"),
        ("m.soil_boxes", "Soil core boxes", None, "nr"),
        ("m.rock_boxes", "Rock core boxes",
         f"=ROUNDUP(({where.get('q.rock_m')}+{where.get('q.hard_m')})/"
         f"{where.get('inputs.rock_box_capacity_m')},0)", "nr"),
        ("m.boxes", "Total core boxes", None, "nr"),
        ("m.grout", "Backfill grout volume",
         f"=PI()*({where.get('inputs.grout_hole_diameter_m')}/2)^2*"
         f"{where.get('q.total_m')}*1000", "L"),
    ):
        ws.cell(row=row, column=2, value=label)
        if formula is None:
            formula = {
                "m.in_tubes": f"={where.get('m.samples')}*"
                              f"{where.get('inputs.mazier_sample_length_m')}",
                "m.for_boxing": f"={where.get('q.soil_m')}-{where.get('m.in_tubes')}",
                "m.soil_boxes": f"=ROUNDUP({where.get('m.for_boxing')}/"
                                f"{where.get('inputs.soil_box_capacity_m')},0)",
                "m.boxes": f"={where.get('m.soil_boxes')}+{where.get('m.rock_boxes')}",
            }[key]
        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=5, value=unit)
        where.put(key, ws.title, f"$D${row}")
        row += 1

    _widths(ws, {"B": 48, "C": 15, "D": 18, "E": 15, "F": 14, "G": 60})


# ---------------------------------------------------------------------------
def _resource_rates(ws, model: CostingModel, spread: Spread, where: _Ref) -> None:
    """03 Resource Rates — the two day-costs, kept apart."""
    ws["A1"] = "03 — Resource rates"
    ws["A1"].font = TITLE
    ws["A2"] = ("Daily cost of one drilling spread. A scales with the rig count; B does not.")
    ws["A2"].font = NOTE
    for n, head in enumerate(["Resource", "Multiplier", "Rate ($/day)", "Cost ($/day)", "Note"],
                             start=2):
        ws.cell(row=4, column=n, value=head).font = HEADING

    row = 5
    rig_rows: list[int] = []
    team_rows: list[int] = []
    gft_rows: list[int] = []

    for charge, title, bucket in ((CHARGE_RIG_DAY, "A — PER RIG-DAY  (scales with rigs)", rig_rows),
                                  (CHARGE_CONTRACT_DAY,
                                   "B — PER CONTRACT-DAY  (the SITE team: per site, not per rig)",
                                   team_rows),
                                  (CHARGE_GFT,
                                   "B2 — PER GFT-DAY  (the GFT: one per gft_ratio RIGS)",
                                   gft_rows)):
        ws.cell(row=row, column=2, value=title).font = LABEL
        row += 1
        for line in model.spread:
            if line.charge != charge:
                continue
            ws.cell(row=row, column=2, value=line.label)
            ws.cell(row=row, column=3,
                    value=where.formula(f"spread.{line.key}.multiplier")).font = LINK
            ws.cell(row=row, column=4, value=where.formula(f"spread.{line.key}.rate")).font = LINK
            cost = ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
            cost.number_format = MONEY
            ws.cell(row=row, column=6, value=line.note).font = NOTE
            bucket.append(row)
            row += 1

        total = ws.cell(row=row, column=2, value={
            CHARGE_RIG_DAY: "A — COST PER RIG-DAY",
            CHARGE_CONTRACT_DAY: "B — COST PER CONTRACT-DAY",
            CHARGE_GFT: "B2 — COST PER GFT-DAY",
        }[charge])
        total.font = HEADING
        cell = ws.cell(row=row, column=5, value=_sum_of(bucket))
        cell.font = HEADING
        cell.number_format = MONEY
        where.put({CHARGE_RIG_DAY: "rig_day", CHARGE_CONTRACT_DAY: "contract_day",
                   CHARGE_GFT: "gft_day"}[charge], ws.title, f"$E${row}")
        row += 2

    # C — the preliminaries. Deliberately below the two totals and outside both, because these are
    # billed as their own items: rolling them into a day-cost would charge the site office inside
    # every metre drilled AND again on the line the client asked for it on.
    prelims = model.prelims()
    if prelims:
        ws.cell(row=row, column=2,
                value="C — PRELIMINARIES  (billed as their own items, in neither total above)"
                ).font = LABEL
        row += 1
        for line in prelims:
            ws.cell(row=row, column=2, value=line.label)
            ws.cell(row=row, column=3,
                    value=where.formula(f"spread.{line.key}.multiplier")).font = LINK
            ws.cell(row=row, column=4, value=where.formula(f"spread.{line.key}.rate")).font = LINK
            cost = ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
            cost.number_format = MONEY
            ws.cell(row=row, column=6,
                    value=f"{line.unit}. {line.note}".strip()).font = NOTE
            # 05 BQ Priced points here, so changing one of these re-prices every line that uses it
            # with the app shut. That is the difference between a model and a report.
            where.put(f"prelim.{line.key}", ws.title, f"$E${row}")
            row += 1
        row += 1

    # THE TWO COUNTS, ARRIVED AT DIFFERENTLY. Sites x a coefficient for the site team — no ROUNDUP,
    # because half a team is a team shared with another contract and rounding invents one nobody
    # employs. ROUNDUP for the GFT, because it is a whole team per group of rigs.
    ws.cell(row=row, column=2, value="Number of sites")
    ws.cell(row=row, column=5, value=where.formula("inputs.site_count")).font = LINK
    where.put("site_count", ws.title, f"$E${row}")
    row += 1

    ws.cell(row=row, column=2, value="Site teams per site")
    ws.cell(row=row, column=5, value=where.formula("inputs.site_team_per_site")).font = LINK
    where.put("per_site", ws.title, f"$E${row}")
    row += 1

    ws.cell(row=row, column=2, value="Site teams carried")
    ws.cell(row=row, column=5, value=f"={where.get('site_count')}*{where.get('per_site')}")
    ws.cell(row=row, column=6,
            value="Independent of the rig count — the site team manages a site.").font = NOTE
    where.put("teams", ws.title, f"$E${row}")
    row += 1

    ws.cell(row=row, column=2, value="One GFT supervises this many rigs")
    ws.cell(row=row, column=5, value=where.formula("inputs.gft_ratio")).font = LINK
    where.put("gft_ratio", ws.title, f"$E${row}")
    row += 1

    ws.cell(row=row, column=2, value="GFTs required")
    ws.cell(row=row, column=5,
            value=f"=ROUNDUP({where.get('rigs')}/{where.get('gft_ratio')},0)")
    where.put("gfts", ws.title, f"$E${row}")
    row += 1

    ws.cell(row=row, column=2, value="Rig cost — full programme (P50)")
    ws.cell(row=row, column=5,
            value=f"={where.get('rig_day')}*{where.get('work_days')}").number_format = MONEY
    row += 1

    ws.cell(row=row, column=2, value="Rig cost — full programme (P90)")
    ws.cell(row=row, column=5,
            value=f"={where.get('rig_day')}*{where.get('work_days_p90')}").number_format = MONEY
    ws.cell(row=row, column=6,
            value="Exposure if productivity lands at the slow end of the as-built range.").font = NOTE
    row += 1

    ws.cell(row=row, column=2, value="Site team cost — full programme")
    ws.cell(row=row, column=5,
            value=(f"={where.get('contract_day')}*{where.get('teams')}"
                   f"*{where.get('days_available')}")).number_format = MONEY
    where.put("team_cost", ws.title, f"$E${row}")
    row += 1

    ws.cell(row=row, column=2, value="GFT cost — full programme")
    ws.cell(row=row, column=5,
            value=(f"={where.get('gft_day')}*{where.get('gfts')}"
                   f"*{where.get('days_available')}")).number_format = MONEY
    ws.cell(row=row, column=6,
            value="Zero until the GFT day-rate is entered — supervising the rigs is not free."
            ).font = NOTE
    where.put("gft_cost", ws.title, f"$E${row}")

    _widths(ws, {"B": 44, "C": 13, "D": 15, "E": 16, "F": 62})


def _sum_of(rows: list[int]) -> str:
    if not rows:
        return "=0"
    return f"=SUM(E{rows[0]}:E{rows[-1]})"


# ---------------------------------------------------------------------------
def _item_buildup(ws, model: CostingModel, buildup: Buildup, where: _Ref) -> None:
    """04 Item Buildup — the spread allocated to each kind of item, then the mark-up chain."""
    ws["A1"] = "04 — Item build-up"
    ws["A1"].font = TITLE
    ws["A2"] = "Allocates the spread cost and the materials to the priced item types."
    ws["A2"].font = NOTE
    for n, head in enumerate(["Item basis", "Days / qty", "Unit cost", "Total cost",
                              "Cost per unit", "Derivation"], start=2):
        ws.cell(row=4, column=n, value=head).font = HEADING

    row = 5
    first = row
    for line in buildup.rows:
        ws.cell(row=row, column=2, value=line.label)
        ws.cell(row=row, column=3, value=line.quantity).number_format = NUMBER
        if line.unit_cost is not None:
            ws.cell(row=row, column=4, value=line.unit_cost).number_format = MONEY
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}" if line.unit_cost is not None
                else line.total_cost).number_format = MONEY
        if line.cost_per_unit is not None:
            ws.cell(row=row, column=6,
                    value=(f"=E{row}/{line.divisor}" if line.divisor else f"=E{row}")
                    ).number_format = MONEY
        ws.cell(row=row, column=7, value=line.derivation or line.note).font = NOTE
        where.put(f"basis.{line.key}", ws.title, f"$F${row}")
        row += 1

    ws.cell(row=row, column=2, value="TOTAL DIRECT COST").font = HEADING
    ws.cell(row=row, column=5, value=f"=SUM(E{first}:E{row - 1})").number_format = MONEY
    row += 2

    ws.cell(row=row, column=2, value="MARK-UP CHAIN  (applied to every rate on 05)").font = LABEL
    row += 1
    factor_rows: list[int] = []
    for step in buildup.markup_steps:
        ws.cell(row=row, column=2, value=step["label"])
        ws.cell(row=row, column=3, value=step["rate"]).number_format = PERCENT
        ws.cell(row=row, column=6,
                value=(f"=1/(1-C{row})" if step["kind"] == "on_selling" else f"=1+C{row}")
                ).number_format = "0.0000"
        ws.cell(row=row, column=7,
                value=("Taken on the selling price, not added to cost — 10% is ×1.111, not ×1.100."
                       if step["kind"] == "on_selling" else "A loading added to cost.")).font = NOTE
        factor_rows.append(row)
        row += 1

    ws.cell(row=row, column=2, value="COMBINED SELLING FACTOR").font = HEADING
    product = "*".join(f"F{n}" for n in factor_rows) or "1"
    cell = ws.cell(row=row, column=6, value=f"={product}")
    cell.font = HEADING
    cell.number_format = "0.0000"
    where.put("selling_factor", ws.title, f"$F${row}")

    _widths(ws, {"B": 34, "C": 14, "D": 14, "E": 16, "F": 15, "G": 72})


# ---------------------------------------------------------------------------
def _bq_priced(ws, priced: PricedBQ, where: _Ref) -> None:
    """05 BQ Priced — the deliverable. Column K is what goes into the client's bill."""
    ws["A1"] = "05 — BQ priced"
    ws["A1"].font = TITLE
    ws["A2"] = ("THE DELIVERABLE. Column K is the rate to transcribe. Every cell here is a formula "
                "or a stated input — nothing is written by a model.")
    ws["A2"].font = NOTE
    if priced.placeholders:
        ws["A3"] = (f"⚠ DO NOT SUBMIT. {len(priced.placeholders)} lines stand on a placeholder — a "
                    f"stand-in for the SHAPE of a line, not an estimate of it. "
                    f"{priced.placeholder_total:,.0f} of the total below was chosen by nobody.")
        ws["A3"].font = Font(bold=True, color="FFC25539")
    headers = ["BQ item", "Description", "Quantity", "Unit", "Cost basis ($/unit)", "Direct cost",
               "Selling factor", "Rate (raw)", "Rate (rounded)", "RATE TO SUBMIT", "Amount"]
    for n, head in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=n, value=head)
        cell.font = HEADING
        cell.fill = HEADER_FILL

    row = 5
    first = row
    for line in priced.rows:
        ws.cell(row=row, column=2, value=line.full_ref)
        ws.cell(row=row, column=3, value=line.description).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=4, value="-" if line.lump else line.qty)
        ws.cell(row=row, column=5, value=line.unit)

        if line.source == "client":
            ws.cell(row=row, column=11, value=line.rate_to_submit).number_format = MONEY
            ws.cell(row=row, column=12, value=line.amount).number_format = MONEY
            ws.cell(row=row, column=13, value=line.note).font = NOTE
        elif line.source == "typed":
            # A rate somebody entered. It is a SELLING rate — the commercial judgement is already
            # in it — so no factor is applied, and it is written as an input rather than a formula
            # because there is no build-up behind it to point at.
            ws.cell(row=row, column=11, value=line.rate_to_submit).font = INPUT
            ws.cell(row=row, column=12,
                    value=f"=IF(ISNUMBER(D{row}),D{row}*K{row},K{row})").number_format = MONEY
            ws.cell(row=row, column=13, value=line.note).font = NOTE
        elif line.cost_basis is None:
            bad = ws.cell(row=row, column=11, value="NO RATE")
            bad.font = Font(bold=True, color="FFC25539")
            ws.cell(row=row, column=13, value=line.note).font = NOTE
        else:
            basis = where.get(f"basis.{line.basis_key}") if line.basis_key else ""
            lab = where.get(f"lab.{line.lab_key}") if line.lab_key else ""
            # A preliminary points at its resource on 03 and converts there, so changing the site
            # office rate in Excel re-prices every line that uses it. Written as an unbracketed
            # ratio of days so the conversion is legible in the formula bar rather than a constant.
            prelim = ""
            if line.prelim_key:
                target = where.get(f"prelim.{line.prelim_key}")
                if target:
                    prelim = target if line.prelim_days == 1.0 else f"{target}*{line.prelim_days}"
            source = basis or lab or prelim
            ws.cell(row=row, column=6, value=f"={source}" if source else line.cost_basis
                    ).font = LINK if source else INPUT
            ws.cell(row=row, column=7,
                    value=f"=IF(ISNUMBER(D{row}),D{row}*F{row},F{row})").number_format = MONEY
            ws.cell(row=row, column=8, value=where.formula("selling_factor")).font = LINK
            ws.cell(row=row, column=9, value=f"=F{row}*H{row}").number_format = MONEY
            ws.cell(row=row, column=10, value=(
                f'=IF(I{row}>=1000,ROUND(I{row},-2),'
                f'IF(I{row}>=100,ROUND(I{row},-1),ROUND(I{row},0)))')).number_format = MONEY
            submit = ws.cell(row=row, column=11, value=line.rate_to_submit)
            submit.font = INPUT
            submit.number_format = MONEY
            ws.cell(row=row, column=12,
                    value=f"=IF(ISNUMBER(D{row}),D{row}*K{row},K{row})").number_format = MONEY
            # A stand-in gets the same formulas as a real rate — the arithmetic is not the problem —
            # and is called out in red on its own row. Somebody scrolling this sheet must be able to
            # see which numbers nobody chose without cross-referencing anything.
            if line.source == "placeholder":
                mark = ws.cell(row=row, column=13, value=f"PROVISIONAL — {line.note}")
                mark.font = Font(bold=True, color="FFC25539")
        row += 1

    ws.cell(row=row, column=3, value="TOTAL OF PRICED ITEMS").font = HEADING
    cell = ws.cell(row=row, column=12, value=f"=SUM(L{first}:L{row - 1})")
    cell.font = HEADING
    cell.number_format = MONEY
    if priced.placeholders:
        ws.cell(row=row, column=13,
                value=(f"PROVISIONAL. {len(priced.placeholders)} of these lines stand on a "
                       f"placeholder, worth {priced.placeholder_total:,.0f} of this total. "
                       f"Actually priced: {priced.total - priced.placeholder_total:,.0f}.")
                ).font = Font(bold=True, color="FFC25539")
    row += 2
    ws.cell(row=row, column=3,
            value="Column K is BLUE because it is yours: the rounded rate beside it is only a "
                  "proposal, and the amount follows whatever you actually submit.").font = NOTE

    _widths(ws, {"B": 10, "C": 56, "D": 11, "E": 8, "F": 16, "G": 15, "H": 13, "I": 14, "J": 14,
                 "K": 16, "L": 16, "M": 60})


# ---------------------------------------------------------------------------
def _register(ws, register: Register) -> None:
    """06 Assumptions Register — the human gate."""
    ws["A1"] = "06 — Assumptions register"
    ws["A1"].font = TITLE
    ws["A2"] = ("THE HUMAN GATE. Every assumption behind the rates. Nothing is submitted until "
                "every row has a Status.")
    ws["A2"].font = NOTE
    headers = ["#", "Assumption", "Value", "Basis", "Source", "Confidence", "Status",
               "Reviewer", "Comment"]
    for n, head in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=n, value=head)
        cell.font = HEADING
        cell.fill = HEADER_FILL

    row = 5
    first = row
    for n, entry in enumerate(register.rows, start=1):
        ws.cell(row=row, column=2, value=n)
        ws.cell(row=row, column=3, value=entry.label)
        ws.cell(row=row, column=4, value=entry.value)
        ws.cell(row=row, column=5, value=entry.basis).alignment = Alignment(wrap_text=True,
                                                                           vertical="top")
        ws.cell(row=row, column=6, value=entry.source)
        confidence = ws.cell(row=row, column=7, value=entry.confidence)
        if entry.confidence == "Low":
            confidence.fill = KEY_ASSUMPTION
        status = ws.cell(row=row, column=8, value=entry.status)
        status.font = INPUT
        ws.cell(row=row, column=9, value=entry.reviewed_by)
        ws.cell(row=row, column=10, value=entry.comment)
        row += 1

    row += 1
    ws.cell(row=row, column=3, value="STATUS VALUES: Accepted / Revised / Rejected — blank means "
                                     "not yet reviewed.").font = NOTE
    row += 1
    ws.cell(row=row, column=3, value="Rows outstanding").font = HEADING
    ws.cell(row=row, column=4, value=f"=COUNTBLANK(H{first}:H{first + len(register.rows) - 1})")
    row += 1
    ws.cell(row=row, column=3, value="SUBMISSION GATE").font = HEADING
    ws.cell(row=row, column=4, value=(
        f'=IF(COUNTBLANK(H{first}:H{first + len(register.rows) - 1})=0,'
        f'"CLEARED","NOT CLEARED")')).font = HEADING

    _widths(ws, {"B": 5, "C": 34, "D": 22, "E": 78, "F": 14, "G": 12, "H": 12, "I": 12, "J": 40})


# ---------------------------------------------------------------------------
def _empirical(ws) -> None:
    """07 Empirical Basis — the as-built data. Read-only reference."""
    ws["A1"] = "07 — Empirical basis"
    ws["A1"].font = TITLE
    ws["A2"] = ("As-built data behind the production rates. READ-ONLY reference — the numbers the "
                "bands were fitted from.")
    ws["A2"].font = NOTE

    totals = empirical.corpus_totals()
    row = 4
    ws.cell(row=row, column=2,
            value=f"Source: {totals['holes']} as-built drillholes across "
                  f"{len(empirical.SOURCE_CONTRACTS)} completed HK ground investigation "
                  f"contracts.").font = LABEL
    row += 2

    for n, head in enumerate(["Project", "Holes", "Total (m)", "Soil (m)", "Rock (m)", "Rock %",
                              "Work-days", "Blended m/work-day"], start=2):
        ws.cell(row=row, column=n, value=head).font = HEADING
    row += 1
    for contract in empirical.SOURCE_CONTRACTS:
        ws.cell(row=row, column=2, value=contract.name)
        ws.cell(row=row, column=3, value=contract.holes)
        ws.cell(row=row, column=4, value=contract.total_m)
        ws.cell(row=row, column=5, value=contract.soil_m)
        ws.cell(row=row, column=6, value=contract.rock_m)
        ws.cell(row=row, column=7, value=round(contract.rock_fraction, 4)).number_format = PERCENT
        ws.cell(row=row, column=8, value=contract.work_days)
        ws.cell(row=row, column=9, value=round(contract.blended_rate, 2))
        row += 1
    ws.cell(row=row, column=2, value="ALL").font = HEADING
    for n, key in enumerate(("holes", "total_m", "soil_m", "rock_m"), start=3):
        ws.cell(row=row, column=n, value=totals[key]).font = HEADING
    ws.cell(row=row, column=7, value=totals["rock_fraction"]).number_format = PERCENT
    ws.cell(row=row, column=8, value=totals["work_days"]).font = HEADING
    ws.cell(row=row, column=9, value=totals["blended_rate"]).font = HEADING
    row += 2

    ws.cell(row=row, column=2, value="PRODUCTION BY ROCK FRACTION — the basis of the bands").font = LABEL
    row += 1
    for n, head in enumerate(["Rock fraction", "Holes", "Mean depth (m)", "m per work-day"],
                             start=2):
        ws.cell(row=row, column=n, value=head).font = HEADING
    row += 1
    for band in empirical.DEFAULT_BANDS.sorted_bands():
        ws.cell(row=row, column=2, value=band.label)
        ws.cell(row=row, column=3, value=band.holes)
        ws.cell(row=row, column=4, value=band.calibration_depth_m)
        ws.cell(row=row, column=5, value=band.rate)
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="PER-HOLE PRODUCTION SPREAD").font = LABEL
    row += 1
    for n, head in enumerate(["P10", "Median", "P90", "P90 / P10"], start=3):
        ws.cell(row=row, column=n, value=head).font = HEADING
    row += 1
    ws.cell(row=row, column=2, value="m per work-day, a single hole")
    for n, value in enumerate((empirical.SPREAD.p10, empirical.SPREAD.median, empirical.SPREAD.p90,
                               round(empirical.SPREAD.ratio, 1)), start=3):
        ws.cell(row=row, column=n, value=value)
    row += 2

    ws.cell(row=row, column=2, value="THE FITTED MODEL  (retained only as the Method B cross-check)"
            ).font = LABEL
    row += 1
    for label, value in (("Formula", empirical.FITTED.formula()),
                         ("R-squared", empirical.FITTED.r_squared),
                         ("R-squared with project dummies",
                          empirical.FITTED.r_squared_with_project_dummies),
                         ("Median absolute error (work-days)",
                          empirical.FITTED.median_absolute_error_days)):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=value)
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="WHAT THE DATA ACTUALLY SAYS").font = LABEL
    row += 1
    for finding in empirical.FINDINGS:
        ws.cell(row=row, column=2, value=finding).alignment = Alignment(wrap_text=True,
                                                                       vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 1

    _widths(ws, {"B": 42, "C": 12, "D": 16, "E": 14, "F": 12, "G": 10, "H": 12, "I": 20})


def _widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
