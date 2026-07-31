"""T5 — the Departure Schedule: every contract term we are not accepting as drafted.

Wholly deterministic. Each row traces to a register line whose verdict a human wrote at the
review gate; nothing here decides anything. Sits beside the Letter of Qualifications, which is a
different document doing a different job: **departures are about contract TERMS, qualifications
are about SCOPE and price assumptions**, and issuers expect them apart because different people
negotiate each.

What goes on it:

* ``confirmed`` lines — departures we are formally pressing.
* ``query`` lines — raised with the client and not yet answered, marked as such. Included by
  decision: the schedule should show every unresolved contractual point, and an open query is
  unresolved. They also appear in the Letter of Qualifications as priced assumptions, because
  the two documents answer different questions about the same fact.
* ``citation_failed`` lines — **never**. You cannot ask a client to amend a clause whose citation
  could not be verified against the document. That falls out of the s08 guard and is the one
  thing this generator refuses to do.
"""

from __future__ import annotations

import io

from client_boq import models, store
from client_boq.outputs import AUDIENCE_INTERNAL, AUDIENCE_SUBMISSION
from client_boq.outputs.common import other_flags, qualification_warning, strategy_flags

COLUMNS = ["Item", "Clause", "As drafted", "Our departure", "Reason", "Proposed amendment",
           "Status"]


def collect(set_id: str) -> dict:
    """The rows, plus everything a document needs to caveat itself honestly."""
    conn = store.get_conn()
    try:
        record = store.load_set(conn, set_id)
        register = store.load_register(conn, set_id)
        flags = strategy_flags(conn, set_id)
        rfis = {r.register_item: r for r in store.load_rfis(conn, set_id) if r.register_item}
    finally:
        conn.close()
    if record is None or register is None:
        return {}

    rows, excluded = [], []
    for item in register.items:
        if item.status == models.STATUS_CITATION_FAILED:
            # Deliberately not exported, and deliberately reported rather than dropped in silence.
            excluded.append({
                "item": item.item, "clause": item.clause,
                "why": item.citation_note or "the citation could not be verified",
            })
            continue
        if item.status not in (models.STATUS_CONFIRMED, models.STATUS_QUERY):
            continue
        rfi = rfis.get(item.item)
        rows.append({
            "item": item.item,
            "clause": item.clause or item.clause_area,
            "as_drafted": item.cited_text,
            "our_departure": item.proposed_position,
            "reason": item.rationale,
            "proposed_amendment": item.amendment_proposal,
            "status": (
                "Subject to outstanding clarification"
                if item.status == models.STATUS_QUERY else "Departure"
            ),
            "query_ref": (rfi.rfi_id if rfi else ""),
            "query_status": (rfi.status if rfi else ""),
        })

    counts = {}
    for item in register.items:
        counts[item.status] = counts.get(item.status, 0) + 1
    return {
        "set_id": set_id, "project": record["name"], "rows": rows, "excluded": excluded,
        "flags": flags, "counts": counts, "approved": register.approved,
    }


def render_markdown(set_id: str, audience: str = AUDIENCE_INTERNAL) -> str:
    data = collect(set_id)
    if not data:
        return ""
    internal = audience == AUDIENCE_INTERNAL
    lines = [
        f"# Departure Schedule{'  (internal working copy)' if internal else ''}",
        "",
        f"**Project:** {data['project']}",
        f"**Departures listed:** {len(data['rows'])}",
        "",
    ]
    if not data["approved"]:
        lines += ["> **The review register for this set is not approved.** These verdicts are not "
                  "final.", ""]
    lines += qualification_warning(data["flags"], audience)

    if not data["rows"]:
        lines += ["No departures. Every reviewed term was accepted as drafted.", ""]
    else:
        lines += ["| " + " | ".join(COLUMNS) + " |",
                  "|" + "|".join("---" for _ in COLUMNS) + "|"]
        for row in data["rows"]:
            cells = [
                str(row["item"]), row["clause"], _cell(row["as_drafted"]),
                _cell(row["our_departure"]), _cell(row["reason"]),
                _cell(row["proposed_amendment"]), row["status"],
            ]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")

    if internal:
        queried = [r for r in data["rows"] if r["query_ref"]]
        if queried:
            lines += ["## Still with the client", ""]
            for row in queried:
                lines.append(
                    f"- Item {row['item']} ({row['clause']}) — query {row['query_ref']}, "
                    f"{row['query_status']}. If it is still open at freeze it becomes a priced "
                    f"assumption in the Letter of Qualifications."
                )
            lines.append("")
        if data["excluded"]:
            lines += [
                "## Withheld from this schedule", "",
                "These lines cite a clause whose citation could not be verified against the "
                "document. Asking a client to amend a clause we cannot locate is worse than "
                "saying nothing, so they are excluded and listed here instead.", "",
            ]
            for row in data["excluded"]:
                lines.append(f"- Item {row['item']} ({row['clause'] or 'no clause'}) — {row['why']}")
            lines.append("")
        lines += [f"**Register status counts:** "
                  + ", ".join(f"{k}: {v}" for k, v in sorted(data["counts"].items())), ""]
        lines += other_flags(data["flags"])
    return "\n".join(lines)


def _cell(text: str) -> str:
    """One markdown table cell: no pipes, no newlines, never empty."""
    clean = (text or "").replace("|", "\\|").replace("\n", " ").strip()
    return clean or "—"


def render_xlsx(set_id: str, audience: str = AUDIENCE_INTERNAL) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = collect(set_id)
    book = Workbook()
    sheet = book.active
    sheet.title = "Departure Schedule"
    if not data:
        return _save(book)

    sheet["A1"] = f"Departure Schedule — {data['project']}"
    sheet["A1"].font = Font(bold=True, size=14)
    row = 2
    warning = [ln for ln in qualification_warning(data["flags"], audience) if ln.strip()]
    if warning:
        sheet.cell(row=row, column=1,
                   value="RISK: this tender penalises qualifying the bid. See the notes sheet.")
        sheet.cell(row=row, column=1).font = Font(bold=True, color="FF9C0006")
        row += 1
    row += 1

    header = PatternFill("solid", fgColor="FFEFEFEF")
    for column, heading in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=row, column=column, value=heading)
        cell.font = Font(bold=True)
        cell.fill = header
    row += 1
    for entry in data["rows"]:
        values = [entry["item"], entry["clause"], entry["as_drafted"], entry["our_departure"],
                  entry["reason"], entry["proposed_amendment"], entry["status"]]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    for column, width in enumerate([6, 14, 44, 40, 40, 44, 26], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    notes = book.create_sheet("Notes")
    notes["A1"] = "Tender conditions affecting submission"
    notes["A1"].font = Font(bold=True, size=12)
    line = 3
    for flag in data["flags"]:
        notes.cell(row=line, column=1, value=flag["kind"])
        notes.cell(row=line, column=2, value=flag.get("clause", ""))
        notes.cell(row=line, column=3, value=flag.get("page"))
        notes.cell(row=line, column=4, value=flag["quote"]).alignment = Alignment(wrap_text=True)
        line += 1
    if data["excluded"]:
        line += 1
        notes.cell(row=line, column=1, value="Withheld (citation unverified)").font = Font(bold=True)
        line += 1
        for entry in data["excluded"]:
            notes.cell(row=line, column=1, value=entry["item"])
            notes.cell(row=line, column=2, value=entry["clause"])
            notes.cell(row=line, column=4, value=entry["why"])
            line += 1
    for column, width in enumerate([28, 14, 8, 90], start=1):
        notes.column_dimensions[get_column_letter(column)].width = width
    return _save(book)


def _save(book) -> bytes:
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
