"""The revision history as an .xlsx workbook — one worksheet per document event.

Deterministic assembly, no model. The worksheets are the history's tabs: **As issued**, then one
per addendum or correction, each listing every part at the revision it stood at that point. That
mirrors how the tender itself arrives — the real ND/2025/04 package is a base issue plus two
addenda — and it doubles as the evidence behind the addendum-acknowledgement returnable, which
must state which revision of each document was actually priced.

Follows the openpyxl patterns already used by ``estimate/workbook.py``.
"""

from __future__ import annotations

import io

from client_boq import store
from client_boq.models import DOC_ADDENDUM, DOC_CLARIFICATION

_HEADER_FILL = "FFEFEFEF"
_AMENDED_FILL = "FFFFF3CD"   # a part that moved at this event


def _autosize(sheet, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_header(sheet, row: int, headings: list[str]) -> int:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for column, heading in enumerate(headings, start=1):
        cell = sheet.cell(row=row, column=column, value=heading)
        cell.font = Font(bold=True)
        cell.fill = fill
    return row + 1


def build_history_workbook(set_id: str) -> bytes:
    """The full revision history of a set as a workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    conn = store.get_conn()
    try:
        record = store.load_set(conn, set_id)
        documents = store.list_documents(conn, set_id)
        operative = store.load_parts(conn, set_id)
        as_at = {
            doc["seq"]: store.load_parts_as_at(conn, set_id, doc["seq"]) for doc in documents
        }
        revisions = {
            spec.part_id: store.load_part_revisions(conn, set_id, spec.part_id)
            for spec, _p, _c in operative
        }
        changes = conn.execute(
            "SELECT doc_id, part_id, kind, pages, description FROM client_boq_changes "
            "WHERE set_id = ? ORDER BY change_id",
            (set_id,),
        ).fetchall()
    finally:
        conn.close()

    book = Workbook()
    name = (record or {}).get("name", set_id)

    # -- Summary: what came in, and what it moved ---------------------------
    sheet = book.active
    sheet.title = "History"
    sheet["A1"] = f"Revision history — {name}"
    sheet["A1"].font = Font(bold=True, size=14)
    row = _write_header(sheet, 3, ["Seq", "Document", "Kind", "Reference", "Received",
                                   "Parts amended"])
    for doc in documents:
        moved = sum(
            1 for revs in revisions.values()
            for rev in revs if rev["doc_id"] == doc["doc_id"] and rev["rev"] > 0
        )
        sheet.cell(row=row, column=1, value=doc["seq"])
        sheet.cell(row=row, column=2, value=doc["filename"])
        sheet.cell(row=row, column=3, value=doc["kind"])
        sheet.cell(row=row, column=4, value=doc["ref"])
        sheet.cell(row=row, column=5, value=doc["received_at"])
        sheet.cell(row=row, column=6, value=moved)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Acknowledgement of addenda received").font = Font(bold=True)
    row += 1
    addenda = [d for d in documents if d["kind"] == DOC_ADDENDUM]
    if addenda:
        for doc in addenda:
            sheet.cell(row=row, column=1, value=doc["ref"] or doc["filename"])
            sheet.cell(row=row, column=2, value=f"received {doc['received_at']}")
            row += 1
    else:
        sheet.cell(row=row, column=1, value="No addenda received.")
        row += 1
    row += 1
    sheet.cell(
        row=row, column=1,
        value=("Corrections are our own re-uploads and are deliberately excluded from the "
               "acknowledgement above. Clarifications change no document."),
    ).font = Font(italic=True)
    _autosize(sheet, [6, 42, 14, 28, 22, 16])

    # -- One sheet per event: the set as it stood then ----------------------
    highlight = PatternFill("solid", fgColor=_AMENDED_FILL)
    for doc in documents:
        if doc["kind"] == DOC_CLARIFICATION:
            continue  # it changed nothing, so it is not a state of the tender
        label = f"{doc['seq']:02d} {doc['ref'] or doc['kind']}"[:31]
        tab = book.create_sheet(title=label)
        tab["A1"] = f"The tender as at: {doc['ref'] or doc['filename']} ({doc['kind']})"
        tab["A1"].font = Font(bold=True, size=12)
        line = _write_header(tab, 3, ["#", "Part", "Title", "Category", "Rev", "Pages",
                                      "Source document"])
        for spec, _path, _ctx in as_at.get(doc["seq"], []):
            introduced = any(
                rev["rev"] == spec.rev and rev["doc_id"] == doc["doc_id"]
                for rev in revisions.get(spec.part_id, [])
            )
            values = [spec.n, spec.part_id, spec.title, spec.category, spec.rev,
                      f"{spec.start}-{spec.end}", spec.source_doc]
            for column, value in enumerate(values, start=1):
                cell = tab.cell(row=line, column=column, value=value)
                if introduced:
                    cell.fill = highlight   # this event is what moved this part
            line += 1
        _autosize(tab, [5, 14, 46, 22, 6, 12, 34])

    # -- The declared changes, marked as advisory ---------------------------
    tab = book.create_sheet(title="Declared changes")
    tab["A1"] = "What each addendum says it changed"
    tab["A1"].font = Font(bold=True, size=12)
    tab["A2"] = ("Reproduced from each addendum's own table. An addendum's remarks may be neither "
                 "exhaustive nor accurate — the replacement pages are the authority.")
    tab["A2"].font = Font(italic=True)
    line = _write_header(tab, 4, ["Document", "Applied to part", "Kind", "Pages", "Description"])
    for change in changes:
        ref = next((d["ref"] or d["filename"] for d in documents
                    if d["doc_id"] == change["doc_id"]), change["doc_id"])
        for column, value in enumerate(
            [ref, change["part_id"] or "(not mapped)", change["kind"], change["pages"],
             change["description"]], start=1,
        ):
            tab.cell(row=line, column=column, value=value)
        line += 1
    if not changes:
        tab.cell(row=line, column=1, value="No addendum changes recorded.")
    _autosize(tab, [28, 18, 18, 26, 90])

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
