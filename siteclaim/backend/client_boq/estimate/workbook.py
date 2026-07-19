"""Deterministic pricing-workbook export (.xlsx) from a persisted estimate — no AI, no macros.

Mirrors the reference model's sheet structure: a WBS summary, a Resources sheet, one sheet per direct
activity (the cost trace), an Indirect Costs sheet, and a Flags sheet. Every figure equals the
persisted estimate exactly (plain values + formatting; no formulas required). Built with openpyxl.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from client_boq.estimate import money
from client_boq.models import Estimate, RateRow
from client_boq.rates import load_rates, rate_index

_MONEY_FMT = "#,##0.00"
_BOLD = Font(bold=True)


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _BOLD


def _marked_up(amount: float, margin_pct: float) -> float:
    return money(amount * (1 + margin_pct / 100.0))


def _safe_sheet_title(item_id: str, used: set[str]) -> str:
    # Excel sheet titles: <=31 chars, no []:*?/\ , unique.
    base = "".join(ch for ch in (item_id or "ACT") if ch not in "[]:*?/\\")[:31] or "ACT"
    title, n = base, 1
    while title in used:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def _wbs_sheet(wb: Workbook, estimate: Estimate) -> None:
    ws = wb.active
    ws.title = "WBS"
    t = estimate.totals
    ws.cell(row=1, column=1, value="Total").font = _BOLD
    ws.cell(row=1, column=2, value=t.total_cost).number_format = _MONEY_FMT
    ws.cell(row=2, column=1, value="Mark-Up Percentage").font = _BOLD
    ws.cell(row=2, column=2, value=t.margin_pct)
    ws.cell(row=3, column=1, value="Profit").font = _BOLD
    ws.cell(row=3, column=2, value=t.margin_amount).number_format = _MONEY_FMT

    r = 4
    ws.cell(row=r, column=1, value="Total Indirect Costs").font = _BOLD
    ws.cell(row=r, column=2, value=t.total_indirect).number_format = _MONEY_FMT
    for ind in estimate.indirects:
        r += 1
        ws.cell(row=r, column=1, value=f"   {ind.label} ({ind.basis})")
        ws.cell(row=r, column=2, value=ind.amount).number_format = _MONEY_FMT
    r += 1
    ws.cell(row=r, column=1, value="Total Direct Costs").font = _BOLD
    ws.cell(row=r, column=2, value=t.total_direct).number_format = _MONEY_FMT

    # Activity summary block.
    r += 2
    _header_row(ws, r, ["TASK ID", "DESCRIPTION", "Quantity", "Rate", "UoM", "Total", "Priced Item?", "Marked Up Price"])
    for act in estimate.activities:
        r += 1
        ws.cell(row=r, column=1, value=act.item_id)
        ws.cell(row=r, column=2, value=act.description)
        ws.cell(row=r, column=3, value="")
        ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5, value=act.unit)
        ws.cell(row=r, column=6, value=act.activity_total).number_format = _MONEY_FMT
        ws.cell(row=r, column=7, value="Y" if act.activity_total > 0 else "N")
        ws.cell(row=r, column=8, value=_marked_up(act.activity_total, t.margin_pct)).number_format = _MONEY_FMT
    _autosize(ws, [16, 40, 10, 12, 8, 16, 12, 16])


def _resources_sheet(wb: Workbook, estimate: Estimate) -> None:
    ws = wb.create_sheet("Resources")
    _header_row(ws, 1, ["Resource ID", "Description", "Type", "Unit", "Rate", "Source"])
    idx = rate_index(load_rates())
    seen: dict[str, None] = {}
    r = 1
    for act in estimate.activities:
        for line in act.lines:
            key = f"{line.resource_ref}|{line.rate_source}"
            if key in seen:
                continue
            seen[key] = None
            row = idx.get(line.resource_ref)
            if line.rate_source == "csv" and row is not None:
                desc, rtype, unit = row.description, row.category, row.unit
            else:
                desc, rtype, unit = line.description, line.rate_source, line.unit
            r += 1
            ws.cell(row=r, column=1, value=line.resource_ref)
            ws.cell(row=r, column=2, value=desc)
            ws.cell(row=r, column=3, value=rtype)
            ws.cell(row=r, column=4, value=unit)
            ws.cell(row=r, column=5, value=line.rate).number_format = _MONEY_FMT
            ws.cell(row=r, column=6, value=line.rate_source)
    _autosize(ws, [16, 40, 12, 10, 14, 10])


def _activity_sheet(wb: Workbook, act, used_titles: set[str]) -> None:
    ws = wb.create_sheet(_safe_sheet_title(act.item_id, used_titles))
    ws.cell(row=1, column=1, value="Activity").font = _BOLD
    ws.cell(row=1, column=2, value=f"{act.item_id} — {act.description}")
    ws.cell(row=2, column=1, value="Unit").font = _BOLD
    ws.cell(row=2, column=2, value=act.unit)
    ws.cell(row=3, column=1, value="Activity Total").font = _BOLD
    ws.cell(row=3, column=2, value=act.activity_total).number_format = _MONEY_FMT

    _header_row(ws, 5, ["Line No", "Description", "Resource", "Resource Type", "Unit", "Qty", "Resource Rate", "Total Cost"])
    r = 5
    for n, line in enumerate(act.lines, start=1):
        r += 1
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=line.description)
        ws.cell(row=r, column=3, value=line.resource_ref)
        ws.cell(row=r, column=4, value=line.rate_source)
        ws.cell(row=r, column=5, value=line.unit)
        ws.cell(row=r, column=6, value=line.qty)
        ws.cell(row=r, column=7, value=line.rate).number_format = _MONEY_FMT
        ws.cell(row=r, column=8, value=line.amount).number_format = _MONEY_FMT
    _autosize(ws, [8, 36, 16, 14, 8, 10, 14, 16])


def _indirects_sheet(wb: Workbook, estimate: Estimate) -> None:
    ws = wb.create_sheet("Indirect Costs")
    _header_row(ws, 1, ["Item ID", "Label", "Basis", "Inputs", "Amount"])
    r = 1
    for ind in estimate.indirects:
        r += 1
        ws.cell(row=r, column=1, value=ind.item_id)
        ws.cell(row=r, column=2, value=ind.label)
        ws.cell(row=r, column=3, value=ind.basis)
        ws.cell(row=r, column=4, value=ind.detail)
        ws.cell(row=r, column=5, value=ind.amount).number_format = _MONEY_FMT
    _autosize(ws, [10, 32, 16, 40, 16])


def _flags_sheet(wb: Workbook, estimate: Estimate) -> None:
    ws = wb.create_sheet("Flags")
    _header_row(ws, 1, ["Flag", "Item ID", "Message"])
    r = 1
    for f in estimate.flags:
        r += 1
        ws.cell(row=r, column=1, value=f.kind)
        ws.cell(row=r, column=2, value=f.item_id)
        ws.cell(row=r, column=3, value=f.message)
    _autosize(ws, [22, 12, 60])


def build_workbook(estimate: Estimate) -> bytes:
    """Build the pricing workbook for a persisted estimate and return the .xlsx bytes."""
    wb = Workbook()
    _wbs_sheet(wb, estimate)
    _resources_sheet(wb, estimate)
    used_titles = {"WBS", "Resources"}
    for act in estimate.activities:
        _activity_sheet(wb, act, used_titles)
    _indirects_sheet(wb, estimate)
    _flags_sheet(wb, estimate)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
