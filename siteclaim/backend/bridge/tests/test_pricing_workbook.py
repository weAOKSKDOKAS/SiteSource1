"""The enquiry carries BOTH: a sliced PDF to read, and a workbook to price.

Proven on the real pack: the sliced bill arrives correctly and is the right document to READ — it
is the bill as issued, with its own item numbers and quantities. It is the wrong document to PRICE.
A print PDF has no form fields, so it was "filled in" in a viewer, nothing was written to the file,
and the return came back with Quantity and Unit intact and Rate and Amount entirely empty. The
engine read it correctly: line items, no rates, scope gap (unpriced). A subcontractor will hit
exactly this.

Three things pinned here:

* the SEAM — both attachments reach the Gmail call with their bytes intact, through the real bridge
  dispatch path, exactly as `test_dispatch_attachment_seam.py` established for one;
* the EQUIVALENCE — the workbook's rows are the same items, descriptions, quantities and units the
  sliced PDF states. A divergence here is a firm pricing the wrong quantities;
* the ROUND TRIP — rates typed into that workbook come back through the reply path and land on the
  comparison, with no model call.

**Fixtures, not the real pack.** The pdfs are written by pymupdf.
"""

import base64
import email as _email

import pytest
from fastapi.testclient import TestClient

from bridge import parts as parts_mod
from pipeline.stage_03_dispatch.drafts import (
    PRICING_WORKBOOK,
    assemble_firm_attachments,
    create_gmail_drafts,
    plan_for_firms,
)
from pipeline.stage_03_dispatch.relevant_docs import PRICED_RETURN
from pipeline.tests.test_gmail_client import StubService
from pipeline.workspace import Workspace
from schemas.models import ScopePackages, SorItem, TradeWorkPackage

fitz = pytest.importorskip("fitz")

SET_ID = "gi-2026-14"
PKG = "ground_investigation:G"

ITEMS = [
    ("G1", "Cable percussion borehole", "m", 120.0),
    ("G2", "Rotary core drilling", "m", 80.0),
    ("G3", "Standpipe piezometer", "nr", 12.0),
]

BILL_PAGES = [
    "SECTION G : DRILLING AND SAMPLING\n"
    + "\n".join(f"{ref}   {desc}   {unit}   {qty:g}" for ref, desc, unit, qty in ITEMS),
    "SECTION H : INSTRUMENTATION\nH1   Inclinometer casing   m   60",
]


@pytest.fixture
def client():
    import api

    return TestClient(api.app)


@pytest.fixture
def make_pdf(tmp_path):
    def _make(name: str, pages: list[str]) -> str:
        doc = fitz.open()
        for body in pages:
            page = doc.new_page()
            y = 70
            for line in body.splitlines():
                page.insert_text((55, y), line, fontsize=10)
                y += 14
        path = tmp_path / name
        doc.save(str(path))
        doc.close()
        return str(path)

    return _make


class StubClient:
    def complete_json(self, *, system, user, target_model, **_kw):
        return ScopePackages(project_name="GI/2026/14", packages=[TradeWorkPackage(
            trade="ground_investigation", scope_summary="Ground investigation",
            sor_items=[SorItem(item_ref=ref, description=desc, unit=unit, qty=qty, section="G")
                       for ref, desc, unit, qty in ITEMS]
            + [SorItem(item_ref="H1", description="Inclinometer casing", unit="m", qty=60.0,
                       section="H")],
        )])


@pytest.fixture
def dispatched(client, make_set, part_spec, make_pdf, tmp_path, monkeypatch):
    """A bridge set, split for real, then dispatched — the workbook is written by `build_dispatch`."""
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "ws"))
    monkeypatch.setattr("pipeline.stage_01_ingest.ingest.LLMClient", lambda *a, **k: StubClient())
    specs = [part_spec(1, "SR", "Schedule of Rates", "pricing", end=2)]
    make_set(SET_ID, "Contract No. GI/2026/14", specs,
             pdf_paths={specs[0].part_id: make_pdf("bill.pdf", BILL_PAGES)})
    parts_mod.confirm_bill_parts(SET_ID, [specs[0].part_id])
    split = client.post(f"/bridge/{SET_ID}/scope").json()
    scope = _dispatched_scope(split["scope"])
    _generate_the_workbook(scope)
    return scope


def _dispatched_scope(scope_json: dict) -> ScopePackages:
    from pipeline.routing.split import route_units

    scope = ScopePackages.model_validate(scope_json)
    units = route_units(scope, split_keys={p.trade for p in scope.packages})
    return ScopePackages(
        project_name=scope.project_name,
        packages=[u["package"].model_copy(update={"trade": u["package_key"]}) for u in units],
    )


def _generate_the_workbook(scope: ScopePackages) -> None:
    """What `build_dispatch` does on the live path, called directly so the test needs no shortlist.

    `assemble_bundle_attachments` writes `sor_sheet_path(tender_id, trade)` whenever a scope AND a
    workspace are given — which `/dispatch/drafts` always passes outside DEMO. This is that call,
    not a stand-in for it.
    """
    from pipeline.stage_03_dispatch.attachments import build_attachments

    ws = Workspace()
    for pkg in scope.packages:
        build_attachments(pkg.trade, scope, None, project_name=SET_ID, tender_id=SET_ID,
                          workspace=ws)


def _plan(scope):
    return plan_for_firms(scope, {PKG: ["F1"]}, tender_id=SET_ID, workspace=Workspace())[PKG]


# -- the plan carries both --------------------------------------------------------------------------
def test_the_plan_carries_the_slice_and_the_workbook(dispatched):
    plan = _plan(dispatched)
    priced = next(a for a in plan.attachments if PRICED_RETURN in a.flags)
    book = next(a for a in plan.attachments if PRICING_WORKBOOK in a.flags)

    assert priced.mode == "sliced" and priced.out_filename.endswith(".pdf")
    assert book.mode == "generated" and book.out_filename.endswith(".xlsx")
    # The pair reads as a pair, and the workbook sits directly after the slice.
    assert book.out_filename == priced.out_filename[:-4] + "_to_price.xlsx"
    assert plan.attachments.index(book) == plan.attachments.index(priced) + 1
    assert "PRICE and return" in book.reason


def test_the_slice_keeps_the_priced_return_flag(dispatched):
    """The gate protects the priced return by that flag; adding a second file must not move it."""
    plan = _plan(dispatched)
    flagged = [a for a in plan.attachments if PRICED_RETURN in a.flags]
    assert len(flagged) == 1 and flagged[0].mode == "sliced"
    assert PRICED_RETURN not in next(a for a in plan.attachments if PRICING_WORKBOOK in a.flags).flags


def test_no_second_workbook_when_the_priced_return_already_is_one(tmp_path, monkeypatch):
    """With no bill PDF to slice, the priced return IS the workbook — do not attach it twice."""
    monkeypatch.setenv("SITESOURCE_WORKDIR", str(tmp_path / "ws2"))
    scope = ScopePackages(project_name="X", packages=[TradeWorkPackage(
        trade="electrical", scope_summary="LV", sor_items=[
            SorItem(item_ref="E1", description="Switchboard", unit="nr", qty=2.0, section="E")])])
    plan = plan_for_firms(scope, {"electrical": ["F1"]}, tender_id="x-1",
                          workspace=Workspace())["electrical"]
    books = [a for a in plan.attachments if a.mode == "generated"]
    assert len(books) == 1
    assert PRICED_RETURN in books[0].flags and PRICING_WORKBOOK not in books[0].flags


def test_the_gate_preview_shows_both_before_anything_is_drafted(client, dispatched, monkeypatch):
    """`/dispatch/plan` is what the operator reads before pressing Compose.

    It generates the workbook itself rather than waiting for `build_dispatch`, or it would list one
    attachment where the enquiry sends two — the operator approving a set they were not shown.
    """
    monkeypatch.setenv("DEMO_MODE", "false")
    plans = client.post("/dispatch/plan", json={
        "scope": dispatched.model_dump(),
        "approvals": {PKG: ["F1"]},
        "project_name": SET_ID,
    }).json()

    atts = [a for p in plans for a in p["attachments"]]
    priced = next(a for a in atts if "priced_return" in a["flags"])
    book = next(a for a in atts if "pricing_workbook" in a["flags"])
    assert priced["mode"] == "sliced"
    assert book["out_filename"].endswith("_to_price.xlsx")
    assert "PRICE and return" in book["reason"]      # the role is stated, not left to the filename


def test_the_email_body_says_which_file_to_fill_in():
    """One plain line, deterministic — a model will not phrase it the same way twice, and this is
    the instruction the whole return path depends on."""
    from pipeline.stage_03_dispatch.dispatch import _PRICING_NOTE, _with_pricing_note
    from schemas.models import AttachmentKind, BundleAttachment

    sheet = BundleAttachment(filename="SoR_x.xlsx", kind=AttachmentKind.SOR_SHEET, trade="x",
                             source_path="/tmp/SoR_x.xlsx", generated=True, label="x")
    body = "Dear Firm,\n\nPlease price the package.\n\nKind regards,\nBuying Team"

    out = _with_pricing_note(body, [sheet])
    assert _PRICING_NOTE in out
    assert out.index(_PRICING_NOTE) < out.index("Kind regards")   # above the sign-off, not under it
    assert _with_pricing_note(out, [sheet]) == out                # idempotent — never doubled

    # Nothing to fill in, nothing claimed. A described-but-unwritten sheet is not an enclosure.
    described = sheet.model_copy(update={"source_path": None})
    assert _with_pricing_note(body, [described]) == body
    assert _with_pricing_note(body, []) == body


# -- the equivalence: same scope in both documents ----------------------------------------------------
def test_the_workbook_rows_are_the_items_the_slice_states(dispatched):
    """A firm pricing the workbook must be pricing the scope the PDF states.

    Item refs, descriptions, units and quantities — all four, because a divergence in any of them
    is a firm pricing the wrong thing while both documents look right.
    """
    from pipeline.stage_04_level.reply_xlsx import parse_sor_xlsx

    ws = Workspace()
    plan = _plan(dispatched)
    atts = {a["filename"]: a for a in assemble_firm_attachments(plan, ws, SET_ID, PKG)}
    book = next(a for a in atts.values() if a["filename"].endswith("_to_price.xlsx"))

    parsed = parse_sor_xlsx(base64.b64decode(book["content_b64"]))
    rows = {li.item_ref: li for li in parsed.line_items}

    # The unit's own items — section G only. H1 belongs to another enquiry and must not be here.
    assert set(rows) == {ref for ref, _d, _u, _q in ITEMS}
    for ref, desc, unit, qty in ITEMS:
        assert rows[ref].description == desc
        assert rows[ref].unit == unit
        assert rows[ref].qty == qty
        assert rows[ref].rate is None                 # blank to price — we never pre-fill a rate

    # And the slice states the same refs: same section, same page.
    slice_pdf_bytes = base64.b64decode(
        next(a for a in atts.values() if a["filename"].endswith(".pdf"))["content_b64"])
    with fitz.open(stream=slice_pdf_bytes, filetype="pdf") as doc:
        text = "\n".join(page.get_text() for page in doc)
    for ref, desc, _u, _q in ITEMS:
        assert ref in text and desc.split()[0] in text
    assert "Inclinometer" not in text                 # section H is not in this enquiry


# -- the seam: both arrive at Gmail -------------------------------------------------------------------
def test_both_attachments_reach_the_gmail_call_with_their_bytes(dispatched):
    ws = Workspace()
    plan = _plan(dispatched)
    atts = assemble_firm_attachments(plan, ws, SET_ID, PKG)
    assert len(atts) == len(plan.attachments)        # nothing skipped

    svc = StubService(draft_result={"id": "d-1"})
    drafted, failed = create_gmail_drafts([{
        "firm_id": "F1", "to": "f1@example.com", "subject": "RFQ [SiteSource Ref: r-1]",
        "body": "…", "ref": "r-1", "attachments": atts,
    }], service=svc)
    assert drafted == ["F1"] and failed == []

    (name, payload), = svc.calls
    msg = _email.message_from_bytes(base64.urlsafe_b64decode(payload["body"]["message"]["raw"]))
    arrived = {p.get_filename(): p.get_payload(decode=True) for p in msg.walk() if p.get_filename()}

    assert set(arrived) == {a["filename"] for a in atts}
    for a in atts:
        assert arrived[a["filename"]] == base64.b64decode(a["content_b64"])
    # Named, so a regression says WHICH went missing.
    pdf = next(n for n in arrived if n.endswith(".pdf"))
    xlsx = next(n for n in arrived if n.endswith("_to_price.xlsx"))
    assert arrived[pdf].startswith(b"%PDF") and arrived[xlsx].startswith(b"PK")


# -- the round trip: rates typed in that workbook reach the comparison ---------------------------------
def test_rates_typed_into_the_generated_workbook_land_on_the_comparison(dispatched, monkeypatch):
    """The whole point of sending it. Deterministic — `parse_sor_xlsx`, zero model calls."""
    import io

    import api
    from openpyxl import load_workbook
    from pipeline import reply_loop

    ws = Workspace()
    plan = _plan(dispatched)
    atts = assemble_firm_attachments(plan, ws, SET_ID, PKG)
    book = next(a for a in atts if a["filename"].endswith("_to_price.xlsx"))

    # A subcontractor fills the Rate column and sends the file back.
    wb = load_workbook(io.BytesIO(base64.b64decode(book["content_b64"])))
    sheet = wb.active
    rates = {"G1": 850.0, "G2": 1450.0, "G3": 2100.0}
    header_row = next(r for r in range(1, sheet.max_row + 1)
                      if str(sheet.cell(r, 1).value or "").strip().lower() == "item")
    for r in range(header_row + 1, sheet.max_row + 1):
        ref = str(sheet.cell(r, 1).value or "").strip()
        if ref in rates:
            sheet.cell(r, 5).value = rates[ref]       # the Rate (HKD) column
    buf = io.BytesIO()
    wb.save(buf)
    returned = buf.getvalue()

    # Leveling reads the firm register (`store.firm_profile`), and this suite's conftest points
    # `SITESOURCE_DB` at an empty temp file. A COPY of the committed demo DB, never the file
    # itself: `sitesource.db` is committed, and nothing in a test may write to it (trap 3b).
    import shutil
    from pathlib import Path as _Path

    real = _Path(__file__).resolve().parents[2] / "db" / "sitesource.db"
    if real.is_file():
        copy = _Path(str(ws.root)) / "firms-copy.db"
        shutil.copyfile(real, copy)
        monkeypatch.setenv("SITESOURCE_DB", str(copy))
    monkeypatch.setenv("DEMO_MODE", "false")
    reply_loop.record_dispatch(ws, "REF-1", SET_ID, "F1", PKG)

    sheets, images = api._read_reply_files([("priced.xlsx", None, returned)])
    assert images == [], "a workbook return must never reach the vision path"

    out = api.process_inbound_reply("REF-1", sheets, images)
    assert out.status == "matched"
    priced = {li.item_ref: li.rate for c in out.comparison for li in c.item_rates}
    for ref, rate in rates.items():
        assert priced.get(ref) == rate


def test_a_returned_pdf_still_goes_through_the_vision_path(dispatched):
    """The other return format must not regress: a PDF reply is still rasterised for the model."""
    import api

    doc = fitz.open()
    doc.new_page().insert_text((50, 60), "G1  Cable percussion borehole  m  120  850.00")
    pdf_bytes = doc.tobytes()
    doc.close()

    sheets, images = api._read_reply_files([("return.pdf", "application/pdf", pdf_bytes)])
    assert sheets == [] and len(images) == 1          # not the deterministic path — vision, as before
